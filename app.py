# Pythagorean: distance = √[(Δlat×111)^2 + (Δlon×111×cos(lat))^2] km gives ~11 km for eps=0.1.
# https://stackoverflow.com/questions/34579213/dbscan-for-clustering-of-geographic-location-data
import streamlit as st
import pandas as pd
import numpy as np
import folium
from folium.plugins import MarkerCluster, MeasureControl, Geocoder, Fullscreen
from streamlit_folium import folium_static
import matplotlib.pyplot as plt
import requests
import json
try:
    import msgspec  # 2-3x faster than orjson for large GeoJSON files
    _HAS_MSGSPEC = True
except ImportError:
    _HAS_MSGSPEC = False
try:
    import orjson as _orjson  # 5-15x faster JSON parsing (pip install orjson)
    _HAS_ORJSON = True
except ImportError:
    _HAS_ORJSON = False
import copy
import os
import glob
from math import radians, cos, sin, asin, sqrt
import hashlib
import threading
from io import BytesIO
from datetime import datetime

# ── HAVERSINE DISTANCE ───────────────────────────────────────────────────────
def haversine_distance(lat1, lon1, lat2, lon2):
    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    return 2 * asin(sqrt(a)) * 6371

# ── CALCULATE MD5 CHECKSUM ───────────────────────────────────────────────────
def calculate_md5_checksum(file_obj):
    """Calculate MD5 checksum of a file or file-like object."""
    md5_hash = hashlib.md5()
    try:
        # Handle both file paths (str) and file-like objects (Streamlit UploadedFile)
        if isinstance(file_obj, str):
            # It's a file path
            with open(file_obj, 'rb') as f:
                for chunk in iter(lambda: f.read(4096), b''):
                    md5_hash.update(chunk)
        else:
            # It's a file-like object (Streamlit UploadedFile)
            file_obj.seek(0)  # Reset to beginning
            while True:
                chunk = file_obj.read(4096)
                if not chunk:
                    break
                md5_hash.update(chunk)
            file_obj.seek(0)  # Reset to beginning for potential re-reading
        return md5_hash.hexdigest()
    except Exception as e:
        return None

# ── PROGRESS DIRECTORY ────────────────────────────────────────────────────────
PROGRESS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'progress')
os.makedirs(PROGRESS_DIR, exist_ok=True)

# ── SAVE PROGRESS TO JSON ────────────────────────────────────────────────────
def save_progress(file_md5, file_name):
    """Save current session progress to a JSON file."""
    now = datetime.now()
    save_data = {
        'file_md5': file_md5,
        'file_name': file_name,
        'saved_at': now.strftime('%Y-%m-%d %H:%M:%S'),
        'form_state': {
            'rural_clusters': st.session_state.get('rural_clusters', 0),
            'urban_clusters': st.session_state.get('urban_clusters', 0),
            'non_tribal_clusters': st.session_state.get('non_tribal_clusters', 0),
            'non_mp_clusters': st.session_state.get('non_mp_clusters', 0),
            'mp_cities': st.session_state.get('mp_cities', 0),
            'tribal_clusters': st.session_state.get('tribal_clusters', 0),
            'million_plus_clusters': st.session_state.get('million_plus_clusters', 0),
            'confirm_tribal': st.session_state.get('confirm_tribal', False),
            'confirm_urban': st.session_state.get('confirm_urban', False),
            'confirm_mp': st.session_state.get('confirm_mp', False),
            'manual_entry_mode': st.session_state.get('manual_entry_mode', False),
            'mp_bi1c_values': st.session_state.get('mp_bi1c_values', []),
            'mp_clusters_selected': st.session_state.get('mp_clusters_selected', []),
            'eps_input': st.session_state.get('eps_input', '3.0'),
            'min_neighbors_input': st.session_state.get('min_neighbors_input', '10'),
        },
        'operations': {
            'steps_tracker': st.session_state.get('steps_tracker', []),
            'removed_cns': st.session_state.get('removed_cns', []),
            'merge_operations': [list(m) for m in st.session_state.get('merge_operations', [])],
            'break_operations': st.session_state.get('break_operations', []),
            'move_operations': st.session_state.get('move_operations', []),
            'reconsidered_cns': list(st.session_state.get('reconsidered_cns', set())),
            'remarks_dict': {str(k): v for k, v in st.session_state.get('remarks_dict', {}).items()},
            'districts_to_remove': st.session_state.get('_saved_districts_to_remove', []),
        }
    }
    filename = f"progress_{now.strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(PROGRESS_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(save_data, f, indent=2, ensure_ascii=False)
    return filepath, filename

# ── LOAD SAVED PROGRESS FILES ────────────────────────────────────────────────
def get_saved_progress_files(file_md5=None):
    """Get list of saved progress files, optionally filtered by MD5."""
    files = glob.glob(os.path.join(PROGRESS_DIR, 'progress_*.json'))
    results = []
    for fpath in sorted(files, reverse=True):
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            entry = {
                'path': fpath,
                'filename': os.path.basename(fpath),
                'file_md5': data.get('file_md5', ''),
                'file_name': data.get('file_name', ''),
                'saved_at': data.get('saved_at', ''),
                'steps_count': len(data.get('operations', {}).get('steps_tracker', [])),
            }
            if file_md5 is None or entry['file_md5'] == file_md5:
                results.append(entry)
        except Exception:
            continue
    return results

# ── LOAD PROGRESS FROM JSON ──────────────────────────────────────────────────
def load_progress(filepath):
    """Stage progress data for restore on next rerun (avoids widget-key conflict)."""
    with open(filepath, 'r', encoding='utf-8') as f:
        data = json.load(f)
    # Store entire payload; the top-of-script flush will apply it before any widgets render
    st.session_state['_pending_restore'] = data
    return data

# ── UNDO TO SERIAL NUMBER ────────────────────────────────────────────────────
def undo_to_sno(sno):
    """Undo all steps from S.No onwards. Rebuilds operation lists from remaining steps."""
    # Keep only steps with S.No < sno
    remaining = [s for s in st.session_state.steps_tracker if s.get('S.No', 0) < sno]
    
    # Clear all operation lists
    st.session_state.removed_cns = []
    st.session_state.merge_operations = []
    st.session_state.break_operations = []
    st.session_state.move_operations = []
    st.session_state.reconsidered_cns = set()
    st.session_state.remarks_dict = {}
    st.session_state._saved_districts_to_remove = []
    st.session_state.mp_bi1c_values = []
    st.session_state.mp_clusters_selected = []
    # Reset widget keys so multiselects/forms pick up the cleared state
    for _wk in ('districts_to_remove', 'mp_bi1c_multiselect'):
        if _wk in st.session_state:
            del st.session_state[_wk]
    
    # Rebuild from remaining steps
    for step in remaining:
        op_type = step.get('_op_type', '')
        op_data = step.get('_op_data', {})
        
        if op_type == 'allocation':
            # Restore allocation values in session state
            for key in ['rural_clusters', 'urban_clusters', 'non_tribal_clusters',
                        'non_mp_clusters', 'mp_cities', 'tribal_clusters', 'million_plus_clusters']:
                if key in op_data:
                    st.session_state[key] = op_data[key]
        elif op_type == 'select_mp':
            st.session_state.mp_bi1c_values = op_data.get('mp_bi1c_values', [])
            st.session_state.mp_clusters_selected = op_data.get('mp_clusters_selected', [])
        elif op_type == 'remove_districts':
            st.session_state._saved_districts_to_remove = op_data.get('districts', [])
        elif op_type == 'remove_cn':
            st.session_state.removed_cns.extend(op_data.get('cns', []))
        elif op_type == 'merge':
            st.session_state.merge_operations.append(
                (op_data['source'], op_data['target'])
            )
        elif op_type == 'break':
            st.session_state.break_operations.append(op_data.get('break_op', {}))
        elif op_type == 'move':
            move_op = op_data.get('move_op', {})
            if move_op:
                st.session_state.move_operations.append(move_op)
            # Restore remarks for moved rows
            for idx_str, remark in op_data.get('remarks', {}).items():
                st.session_state.remarks_dict[int(idx_str)] = remark
            # Restore reconsidered CN
            rcn = op_data.get('reconsidered_cn', '')
            if rcn:
                st.session_state.reconsidered_cns.add(rcn)
        elif op_type == 'reconsider':
            cn = op_data.get('cn', '')
            if cn:
                st.session_state.reconsidered_cns.add(cn)
            # Restore remarks
            for idx_str, remark in op_data.get('remarks', {}).items():
                st.session_state.remarks_dict[int(idx_str)] = remark
    
    st.session_state.steps_tracker = remaining
    # Reset DBSCAN labels to force recalculation
    st.session_state.dbscan_labels = None
    st.session_state.last_dbscan_cluster = None

# ── PDF REPORT GENERATION ────────────────────────────────────────────────────
def generate_pdf_report(
    parent_md5, uploaded_md5, output_md5, state_name,
    confirm_tribal, confirm_urban, confirm_mp,
    mp_cities_count, eps_val, min_neighbors_val,
    allocation_summary_1, allocation_summary_2, crosstab_df,
    regular_clusters, irregular_clusters,
    steps_tracker, removed_cns, reconsidered_cns, merge_operations,
    break_operations=None,
    move_operations=None,
    response_rate_103=0.0, response_rate_102_103=0.0,
    total_clusters=0, total_records=0, cluster_summary=None,
    mp_bi1c_values=None
):
    """Generate a professional PDF report with all session details."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm, cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = BytesIO()
    now = datetime.now()
    watermark_text = output_md5 or "NO-CHECKSUM"

    # ── Custom canvas with watermark ─────────────────────────────────────
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus.doctemplate import BaseDocTemplate, PageTemplate, Frame

    class WatermarkDocTemplate(BaseDocTemplate):
        def __init__(self, filename, watermark, **kwargs):
            self._watermark = watermark
            super().__init__(filename, **kwargs)

        def afterPage(self):
            """Draw slanted watermark text with low opacity across every page."""
            c = self.canv
            c.saveState()
            c.setFont("Helvetica", 10)
            c.setFillColorRGB(0.85, 0.85, 0.85)  # light gray
            width, height = A4
            # Tile watermark diagonally
            for y in range(-100, int(height) + 100, 60):
                for x in range(-200, int(width) + 200, 280):
                    c.saveState()
                    c.translate(x, y)
                    c.rotate(35)
                    c.drawString(0, 0, self._watermark)
                    c.restoreState()
            # Footer
            c.setFont("Helvetica", 7)
            c.setFillColorRGB(0.5, 0.5, 0.5)
            c.drawCentredString(width / 2, 12 * mm, f"{state_name} | Cluster Validator Report  |  Generated: {now.strftime('%Y-%m-%d %H:%M:%S')}  |  Output MD5: {output_md5 or 'N/A'}")
            c.restoreState()

    frame = Frame(1.5 * cm, 1.5 * cm, A4[0] - 3 * cm, A4[1] - 3 * cm, id='main')
    template = PageTemplate(id='main', frames=[frame])
    doc = WatermarkDocTemplate(buf, watermark=watermark_text, pagesize=A4)
    doc.addPageTemplates([template])

    styles = getSampleStyleSheet()
    # Custom styles
    styles.add(ParagraphStyle(name='ReportTitle', parent=styles['Title'], fontSize=22, textColor=colors.HexColor('#1a1a2e'), spaceAfter=6))
    styles.add(ParagraphStyle(name='SectionHead', parent=styles['Heading2'], fontSize=14, textColor=colors.HexColor('#16213e'), spaceBefore=14, spaceAfter=6,
                              borderWidth=0, borderPadding=0, borderColor=colors.HexColor('#0f3460')))
    styles.add(ParagraphStyle(name='SubHead', parent=styles['Heading3'], fontSize=11, textColor=colors.HexColor('#0f3460'), spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle(name='BodySmall', parent=styles['Normal'], fontSize=9, leading=12))
    styles.add(ParagraphStyle(name='CheckItem', parent=styles['Normal'], fontSize=9, leading=14, leftIndent=12))
    styles.add(ParagraphStyle(name='CenterSmall', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray))

    story = []

    # Colors
    hdr_bg = colors.HexColor('#16213e')
    hdr_fg = colors.white
    row_alt = colors.HexColor('#f0f4ff')
    accent = colors.HexColor('#0f3460')

    # ── Helper: DataFrame to reportlab table ─────────────────────────────
    def df_to_table(dataframe, col_widths=None):
        # Strip emoji prefixes for PDF
        clean_cols = []
        for c in dataframe.columns:
            clean = c
            for prefix in ['🔴', '🟢', '#']:
                clean = clean.replace(prefix, '')
            clean_cols.append(clean.strip())
        
        data = [clean_cols] + dataframe.values.tolist()
        # Convert all cells to strings
        data = [[str(cell) for cell in row] for row in data]
        
        if col_widths:
            t = Table(data, colWidths=col_widths, repeatRows=1)
        else:
            t = Table(data, repeatRows=1)
        
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 7.5),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]
        # Bold the last row if it looks like a totals row (District table)
        if len(data) > 2:
            last_row_first = str(data[-1][0]).upper()
            if 'TOTAL' in last_row_first:
                style_cmds.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
                style_cmds.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e8edf5')))
        
        t.setStyle(TableStyle(style_cmds))
        return t

    # ═══════════════════════════════════════════════════════════════════════
    # PAGE 1: TITLE + DATA INTEGRITY
    # ═══════════════════════════════════════════════════════════════════════
    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(f"{state_name} | Cluster Validator Report", styles['ReportTitle']))
    story.append(HRFlowable(width="100%", thickness=2, color=accent, spaceAfter=6))
    story.append(Paragraph(f"Generated: {now.strftime('%A, %d %B %Y at %H:%M:%S')}", styles['CenterSmall']))
    story.append(Spacer(1, 0.8 * cm))

    # ── Section 1: Data Integrity ────────────────────────────────────────
    story.append(Paragraph("1. Data Integrity", styles['SectionHead']))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e0e0e0'), spaceAfter=8))

    # Checksums table
    story.append(Paragraph("Checksums", styles['SubHead']))
    checksum_data = [
        ['Checksum', 'Value', 'Description'],
        ['Parent MD5', str(parent_md5 or 'N/A'), 'Source reference from Parent MD5 column'],
        ['Uploaded File MD5', str(uploaded_md5 or 'N/A'), 'Integrity of file uploaded to session'],
        ['Output File MD5', str(output_md5 or 'N/A'), 'MD5 of the downloaded/exported dataset'],
    ]
    checksum_data = [[str(c) for c in row] for row in checksum_data]
    chk_table = Table(checksum_data, colWidths=[90, 200, 180], repeatRows=1)
    chk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('FONTNAME', (1, 1), (1, -1), 'Courier'),
    ]))
    story.append(chk_table)
    story.append(Spacer(1, 0.5 * cm))

    # Million Plus Cluster Selection
    story.append(Paragraph("Million Plus (MP) Cluster Selection", styles['SubHead']))
    if mp_bi1c_values and len(mp_bi1c_values) > 0:
        mp_list = ', '.join(sorted(set(str(v) for v in mp_bi1c_values)))
        story.append(Paragraph(
            f"<b>Selected bi1c values indicating MP clusters:</b><br/>{mp_list}",
            styles['BodySmall']
        ))
    else:
        story.append(Paragraph("<i>No specific MP cluster values selected.</i>", styles['BodySmall']))
    story.append(Spacer(1, 0.4 * cm))

    # Classification confirmations
    story.append(Paragraph("Data Classification Confirmations", styles['SubHead']))
    story.append(Paragraph(f"<b>TRI- Keyword Present:</b> {'Yes' if confirm_tribal else 'No'}", styles['CheckItem']))
    story.append(Paragraph(f"<b>Urban Keyword Present:</b> {'Yes' if confirm_urban else 'No'}", styles['CheckItem']))
    story.append(Paragraph(f"<b>Urban & Mc Keywords Present:</b> {'Yes' if confirm_mp else 'No'}", styles['CheckItem']))
    story.append(Spacer(1, 0.6 * cm))

    # ── Cluster Distribution Histograms ──────────────────────────────────
    if cluster_summary is not None and not cluster_summary.empty:
        story.append(Paragraph("Cluster Distribution Analysis", styles['SubHead']))
        
        # Generate histogram images
        from io import BytesIO as BytesIOForImages
        hist_img_overall = BytesIOForImages()
        hist_img_status103 = BytesIOForImages()
        
        try:
            # Create histogram for overall_count
            fig_overall, ax_overall = plt.subplots(figsize=(4.5, 3.5), dpi=100)
            ax_overall.hist(cluster_summary['overall_count'], bins=20, color='#3498db', edgecolor='black', alpha=0.7)
            ax_overall.set_title('Overall Distribution', fontsize=11, fontweight='bold')
            ax_overall.set_xlabel('Overall Count', fontsize=10)
            ax_overall.set_ylabel('Frequency', fontsize=10)
            ax_overall.grid(True, alpha=0.3, linestyle='--')
            plt.tight_layout()
            fig_overall.savefig(hist_img_overall, format='png', dpi=100, bbox_inches='tight')
            hist_img_overall.seek(0)
            plt.close(fig_overall)
            
            # Create histogram for status_103_count
            fig_status103, ax_status103 = plt.subplots(figsize=(4.5, 3.5), dpi=100)
            ax_status103.hist(cluster_summary['status_103_count'], bins=20, color='#2ecc71', edgecolor='black', alpha=0.7)
            ax_status103.set_title('Status 103 (Completed) Distribution', fontsize=11, fontweight='bold')
            ax_status103.set_xlabel('Status 103 Count', fontsize=10)
            ax_status103.set_ylabel('Frequency', fontsize=10)
            ax_status103.grid(True, alpha=0.3, linestyle='--')
            plt.tight_layout()
            fig_status103.savefig(hist_img_status103, format='png', dpi=100, bbox_inches='tight')
            hist_img_status103.seek(0)
            plt.close(fig_status103)
            
            # Create two-column layout for histograms
            from reportlab.platypus import Image
            hist_width = (A4[0] - 3 * cm) / 2 - 0.2 * cm
            hist_height = 4.5 * cm
            
            hist_overall_img = Image(hist_img_overall, width=hist_width, height=hist_height)
            hist_status103_img = Image(hist_img_status103, width=hist_width, height=hist_height)
            
            hist_table = Table([[hist_overall_img, hist_status103_img]], colWidths=[hist_width, hist_width])
            hist_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            story.append(hist_table)
            story.append(Spacer(1, 0.3 * cm))
        except Exception as e:
            story.append(Paragraph(f"<i>Could not generate histograms: {str(e)}</i>", styles['BodySmall']))
            story.append(Spacer(1, 0.3 * cm))

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 2: FINALISED CLUSTER CORRECTION DETAILS
    # ═══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("2. Finalised Cluster Correction Details", styles['SectionHead']))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e0e0e0'), spaceAfter=8))

    # Parameters
    params_data = [
        ['Parameter', 'Value'],
        ['Total Clusters', str(total_clusters)],
        ['Total Records', str(total_records)],
        ['Number of MP Cities in the state', str(mp_cities_count)],
        ['DBSCAN EPS (km)', str(eps_val)],
        ['DBSCAN MIN Neighbors', str(min_neighbors_val)],
        ['Response Rate (Status 103 only)', f"{response_rate_103:.2f}%"],
        ['Response Rate (Status 102+103)', f"{response_rate_102_103:.2f}%"],
    ]
    params_table = Table(params_data, colWidths=[160, 160], repeatRows=1)
    params_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(params_table)
    story.append(Spacer(1, 0.4 * cm))

    # Rural/Urban Split
    story.append(Paragraph("Rural / Urban Split", styles['SubHead']))
    story.append(df_to_table(allocation_summary_1))
    story.append(Spacer(1, 0.4 * cm))

    # Further Classification
    story.append(Paragraph("Further Classification", styles['SubHead']))
    story.append(df_to_table(allocation_summary_2))
    story.append(Spacer(1, 0.4 * cm))

    # Breakdown by District
    story.append(Paragraph("Breakdown by District", styles['SubHead']))
    n_cols = len(crosstab_df.columns)
    avail_width = A4[0] - 3 * cm
    first_col_w = avail_width * 0.22
    remaining = avail_width - first_col_w
    other_w = remaining / max(n_cols - 1, 1)
    col_ws = [first_col_w] + [other_w] * (n_cols - 1)
    story.append(df_to_table(crosstab_df, col_widths=col_ws))
    story.append(Spacer(1, 0.6 * cm))

    # ── Regular Clusters ─────────────────────────────────────────────────
    story.append(Paragraph("Regular Clusters (Status_103: 25–40)", styles['SubHead']))
    story.append(Paragraph(
        "Clusters whose Status 103 (completed interviews) count falls within the "
        "expected range of <b>25 to 40</b>.",
        styles['BodySmall']
    ))
    story.append(Spacer(1, 0.2 * cm))
    if regular_clusters is not None and not regular_clusters.empty:
        # Remove bi1c and bi1c_1 columns for PDF display
        regular_clusters_pdf = regular_clusters.drop(columns=[col for col in ['bi1c', 'bi1c_1'] if col in regular_clusters.columns])
        story.append(df_to_table(regular_clusters_pdf))
    else:
        story.append(Paragraph("No regular clusters.", styles['BodySmall']))
    story.append(Spacer(1, 0.5 * cm))

    # ── Irregular Clusters ───────────────────────────────────────────────
    story.append(Paragraph("Irregular Clusters (Status_103: &lt;25 or &gt;40)", styles['SubHead']))
    story.append(Paragraph(
        "Clusters whose Status 103 count is <b>below 25</b> or <b>above 40</b>, "
        "indicating under- or over-coverage requiring review.",
        styles['BodySmall']
    ))
    story.append(Spacer(1, 0.2 * cm))
    if irregular_clusters is not None and not irregular_clusters.empty:
        # Remove bi1c and bi1c_1 columns for PDF display
        irregular_clusters_pdf = irregular_clusters.drop(columns=[col for col in ['bi1c', 'bi1c_1'] if col in irregular_clusters.columns])
        story.append(df_to_table(irregular_clusters_pdf))
    else:
        story.append(Paragraph("No irregular clusters.", styles['BodySmall']))
    story.append(Spacer(1, 0.6 * cm))

    # ═══════════════════════════════════════════════════════════════════════
    # SECTION 3: ACTION TRACKER
    # ═══════════════════════════════════════════════════════════════════════
    story.append(Paragraph("3. Action Tracker", styles['SectionHead']))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e0e0e0'), spaceAfter=8))

    if steps_tracker and len(steps_tracker) > 0:
        import pandas as _pd
        steps_df = _pd.DataFrame(steps_tracker)
        display_cols = []
        if 'S.No' in steps_df.columns:
            display_cols.append('S.No')
        display_cols.append('Step')
        display_cols.append('Details')
        if 'Remark' in steps_df.columns:
            display_cols.append('Remark')
        steps_df = steps_df[display_cols]
        story.append(df_to_table(steps_df))
    else:
        story.append(Paragraph("No actions performed during this session.", styles['BodySmall']))

    story.append(Spacer(1, 0.5 * cm))

    # ── Removed Clusters ─────────────────────────────────────────────────
    story.append(Paragraph("Removed Clusters", styles['SubHead']))
    if removed_cns and len(removed_cns) > 0:
        removed_list = sorted(set(str(cn) for cn in removed_cns))
        rem_data = [['#', 'Cluster (CN)']]
        for i, cn in enumerate(removed_list, 1):
            rem_data.append([str(i), cn])
        rem_table = Table(rem_data, colWidths=[40, 200], repeatRows=1)
        rem_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(rem_table)
    else:
        story.append(Paragraph("No clusters removed.", styles['BodySmall']))

    story.append(Spacer(1, 0.4 * cm))

    # ── Reconsidered Clusters ────────────────────────────────────────────
    story.append(Paragraph("Reconsidered Clusters", styles['SubHead']))
    if reconsidered_cns and len(reconsidered_cns) > 0:
        recon_list = sorted(set(str(cn) for cn in reconsidered_cns))
        rec_data = [['#', 'Cluster (CN)']]
        for i, cn in enumerate(recon_list, 1):
            rec_data.append([str(i), cn])
        rec_table = Table(rec_data, colWidths=[40, 200], repeatRows=1)
        rec_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(rec_table)
    else:
        story.append(Paragraph("No clusters reconsidered.", styles['BodySmall']))

    story.append(Spacer(1, 0.4 * cm))

    # ── Merged Clusters ──────────────────────────────────────────────────
    story.append(Paragraph("Merged Clusters", styles['SubHead']))
    if merge_operations and len(merge_operations) > 0:
        merge_data = [['#', 'Source CN', '', 'Target CN']]
        for i, (src, tgt) in enumerate(merge_operations, 1):
            merge_data.append([str(i), str(src), u'\u2192', str(tgt)])  # →
        merge_table = Table(merge_data, colWidths=[40, 120, 30, 120], repeatRows=1)
        merge_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (2, 1), (2, -1), colors.HexColor('#d62728')),
        ]))
        story.append(merge_table)
    else:
        story.append(Paragraph("No clusters merged.", styles['BodySmall']))

    story.append(Spacer(1, 0.4 * cm))

    # ── Break Clusters (newly formed) ────────────────────────────────────
    story.append(Paragraph("Break Clusters (Newly Formed)", styles['SubHead']))
    if break_operations and len(break_operations) > 0:
        break_data = [['#', 'Source CN', 'DBSCAN ID', '', 'New CN', 'Overrides']]
        for i, bop in enumerate(break_operations, 1):
            overrides = []
            if bop.get('bi1a'): overrides.append(f"bi1a={bop['bi1a']}")
            if bop.get('bi1b'): overrides.append(f"bi1b={bop['bi1b']}")
            if bop.get('bi1c'): overrides.append(f"bi1c={bop['bi1c']}")
            if bop.get('bi1c_1'): overrides.append(f"bi1c_1={bop['bi1c_1']}")
            if bop.get('lat'): overrides.append(f"lat={bop['lat']}")
            if bop.get('lon'): overrides.append(f"lon={bop['lon']}")
            override_str = ', '.join(overrides) if overrides else '-'
            # Support both new format (dbscan_ids list) and old format (dbscan_cluster_id)
            if 'dbscan_ids' in bop:
                dbscan_str = ', '.join(str(x) for x in bop['dbscan_ids'])
            else:
                dbscan_str = str(bop['dbscan_cluster_id'])
            break_data.append([
                str(i), str(bop['source_cn']), dbscan_str,
                u'\u2192', str(bop['new_cn']), override_str
            ])
        break_table = Table(break_data, colWidths=[30, 80, 60, 20, 80, 200], repeatRows=1)
        break_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (-1, 1), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('FONTNAME', (3, 1), (3, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (3, 1), (3, -1), colors.HexColor('#2ca02c')),
        ]))
        story.append(break_table)
    else:
        story.append(Paragraph("No clusters broken.", styles['BodySmall']))

    story.append(Spacer(1, 0.4 * cm))

    # ── Moved Records ────────────────────────────────────────────────────
    story.append(Paragraph("Moved Records", styles['SubHead']))
    if move_operations and len(move_operations) > 0:
        move_data = [['#', 'Source CN', '', 'Target CN', 'Rows Moved']]
        for i, mop in enumerate(move_operations, 1):
            move_data.append([
                str(i), str(mop['source_cn']), u'\u2192',
                str(mop['target_cn']), str(len(mop['indices']))
            ])
        move_table = Table(move_data, colWidths=[30, 100, 20, 100, 70], repeatRows=1)
        move_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), hdr_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), hdr_fg),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, row_alt]),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (2, 1), (2, -1), colors.HexColor('#ff7f0e')),
        ]))
        story.append(move_table)
    else:
        story.append(Paragraph("No records moved.", styles['BodySmall']))

    # Build PDF
    doc.build(story)
    buf.seek(0)
    return buf.getvalue()

# ── GENERATE NAME CORRECTION FORM (EXCEL) ───────────────────────────────────
def generate_name_correction_form(df, df_processed=None, min_neighbors_val='10'):
    """
    Generate an Excel file with Name Correction Form with two-level headers.
    Sections:
      1. Server Data    – CN, bi1a, bi1b, bi1c, bi1c_1, Rural-Urban, TYPE
      2. Basic Info     – is_Regular, Overall Count, Completed Interviews,
                          Adolescent Count (103), Gender Ratio (103),
                          n_splits, Noise   (auto-filled from df / df_processed)
      3. Your Correction– Is Habitation, Block Name, Correct Gram Panchayat Name,
                          Correct Village Name, Remark
      4. Issues         – Undesirable Split, HH Address Not Aligned,
                          Questionable Date, Not Aligning with LGD Maps,
                          Not Consistent with Monitoring forms,
                          Too much Non-response, Notes   (checkbox ☐/☑ + free text)
    Returns BytesIO object containing the Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        st.error("openpyxl library not installed. Please install it with: pip install openpyxl")
        return None

    # ── Column layout ──────────────────────────────────────────────────────
    # Section 1: Server Data  (cols 1-7)
    server_headers = ['CN', 'bi1a', 'bi1b', 'bi1c', 'bi1c_1', 'Rural-Urban', 'TYPE']
    # Section 2: Basic Info   (cols 8-17)
    basic_headers  = [
        'is_Regular\n(Yes/No)',
        'Total Individuals\nContacted\nOVERALL COUNT',
        'Total Completed\nInterviews\n(Status 103)',
        'Response Rate\n(%)',
        'Adult Count\n(103)',
        'Adolescent\nCount (103)',
        'Gender Ratio\n(103)\nM:F',
        'Missing\nGeocodes\n(lat/lon)',
        f'n_splits\n(DBSCAN {eps_val}km)',
        f'Noise\n({eps_val}km | min_NEIGHBOUR:{min_neighbors_val})',
    ]
    # Section 3: Your Correction (cols 17-23)
    correction_headers = [
        'Is the selected\nvillage name in\nthe server a\nhabitation name?',
        'Is the selected\ncluster Tribal?',
        'Block Name',
        'Correct Gram\nPanchayat Name',
        'Correct\nVillage Name',
        'VILCODE11',
        'Remark',
    ]
    # Section 4: Issues (cols 24-33)
    issue_headers = [
        'Cluster split into\nmany small groups\nthat are far from\neach other -\nUNDESIRABLE SPLIT',
        'Many interviews\nwere far apart and\ncould not form a\nproper SUB-cluster -\nTOO MUCH NOISE',
        'Presence of Duplicates or Reinterviews',
        'Gender\nBias',
        'Household addresses\ndo not match the\nexpected location',
        'Questionable\nDate',
        'Not Aligning\nwith LGD Maps',
        'Not Consistent\nwith Monitoring\nForms',
        'Too much\nNon-response',
        'Notes',
        'Total\nIssues',
    ]
    # Section 5: Weightage Details (cols 32-36)
    weightage_headers = [
        'District Rural\nPopulation',
        'Gram Panchayat\nPopulation',
        'No of GPs in\nthe District',
        'No of Villages\nin the GP',
        'No of Households\nin the Village',
    ]

    all_headers = server_headers + basic_headers + correction_headers + issue_headers + weightage_headers

    # Column index ranges (1-based)
    server_start,     server_end     = 1,  len(server_headers)
    basic_start,      basic_end      = server_end + 1,  server_end + len(basic_headers)
    correction_start, correction_end = basic_end  + 1,  basic_end  + len(correction_headers)
    issue_start,      issue_end      = correction_end + 1, correction_end + len(issue_headers)
    weightage_start,  weightage_end  = issue_end + 1,  issue_end + len(weightage_headers)
    notes_col        = issue_end - 1   # second-to-last in issues = Notes (free text)
    total_issues_col = issue_end       # last in issues = Total Issues (formula)
    missing_geocodes_col = basic_start + 7  # Position of missing_geocodes in basic section (8th column)

    # ── Colour palette ─────────────────────────────────────────────────────
    server_fill     = PatternFill(start_color="366092", end_color="366092", fill_type="solid")   # dark blue
    basic_fill      = PatternFill(start_color="375623", end_color="375623", fill_type="solid")   # dark green
    correction_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")   # dark orange
    issue_fill      = PatternFill(start_color="7B2C2C", end_color="7B2C2C", fill_type="solid")   # dark red
    weightage_fill  = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")   # dark grey

    # Sub-header fills (lighter versions)
    server_sub_fill     = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    basic_sub_fill      = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    correction_sub_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    issue_sub_fill      = PatternFill(start_color="FF7C7C", end_color="FF7C7C", fill_type="solid")
    weightage_sub_fill  = PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid")

    # Data-row background tints
    basic_data_fill      = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # light green
    correction_data_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")   # light orange
    issue_data_fill      = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")   # light red
    weightage_data_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")   # light blue

    white_font = Font(bold=True, color="FFFFFF")
    dark_font  = Font(bold=True, color="000000")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin')
    )

    # ── Workbook / worksheet ───────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Name Correction"

    def _col_letter(n):
        return get_column_letter(n)

    # also expose as bare function name for use inside loops
    from openpyxl.utils import get_column_letter

    # ── Row 1: Section headers (merged) ────────────────────────────────────
    section_defs = [
        ("Server Data",    server_start,     server_end,     server_fill),
        ("Basic Info",     basic_start,      basic_end,      basic_fill),
        ("Your Correction",correction_start, correction_end, correction_fill),
        ("Issues",         issue_start,      issue_end,      issue_fill),
        ("Weightage Details", weightage_start, weightage_end, weightage_fill),
    ]
    for label, c_start, c_end, fill in section_defs:
        start_letter = _col_letter(c_start)
        end_letter   = _col_letter(c_end)
        ws.merge_cells(f'{start_letter}1:{end_letter}1')
        cell = ws[f'{start_letter}1']
        cell.value = label
        cell.font  = Font(bold=True, size=12, color="FFFFFF")
        cell.fill  = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        # Apply border to all merged cells in row 1
        for c in range(c_start, c_end + 1):
            ws.cell(row=1, column=c).border = thin_border

    # ── Row 2: Sub-column headers ───────────────────────────────────────────
    sub_fill_map = {}
    for c in range(server_start, server_end + 1):
        sub_fill_map[c] = (server_sub_fill, white_font)
    for c in range(basic_start, basic_end + 1):
        sub_fill_map[c] = (basic_sub_fill, white_font)
    for c in range(correction_start, correction_end + 1):
        sub_fill_map[c] = (correction_sub_fill, dark_font)
    for c in range(issue_start, issue_end + 1):
        sub_fill_map[c] = (issue_sub_fill, dark_font)
    for c in range(weightage_start, weightage_end + 1):
        sub_fill_map[c] = (weightage_sub_fill, white_font)

    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        fill, fnt = sub_fill_map.get(col_idx, (server_sub_fill, white_font))
        cell.fill = fill
        cell.font = fnt
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # ── Column widths ───────────────────────────────────────────────────────
    col_widths = (
        [12, 12, 15, 20, 20, 12, 10]      # Server Data (7 cols)
        + [12, 13, 14, 12, 12, 14, 14, 12, 10, 10]    # Basic Info (10 cols: is_Regular, Overall Count, Completed Interviews, Response Rate, Adult Count, Adolescent Count, Gender Ratio, Missing Geocodes, n_splits, Noise)
        + [14, 12, 18, 22, 20, 16, 18]    # Your Correction (7 cols)
        + [14, 12, 12, 15, 14, 18, 20, 14, 22, 11] # Issues (10 cols)
        + [18, 18, 16, 16, 18]             # Weightage Details (5 cols)
    )
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[_col_letter(col_idx)].width = width

    # ── Header row heights ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 45

    # ── Build per-CN stats from df and df_processed ─────────────────────────
    mp_bi1c_values = st.session_state.get('mp_bi1c_values', [])

    def _cn_stats(cn, cn_data):
        """Return dict with Basic Info metrics for a single CN."""
        stats = {
            'is_regular': '',
            'overall_count': 0,
            'completed': 0,
            'response_rate': 0.0,
            'adult_count': 0,
            'adolescent_103': 0,
            'gender_ratio': '',
            'missing_geocodes': 0,
            'n_splits': '',
            'noise': '',
        }
        if cn_data.empty:
            return stats

        stats['overall_count'] = len(cn_data)

        if 'interview_status' in cn_data.columns:
            completed = (cn_data['interview_status'] == 103).sum()
            stats['completed'] = int(completed)
            stats['is_regular'] = 'Yes' if 25 <= completed <= 40 else 'No'
            
            # Calculate response rate
            if stats['overall_count'] > 0:
                stats['response_rate'] = (stats['completed'] / stats['overall_count']) * 100
            else:
                stats['response_rate'] = 0.0

            if 'c2_age' in cn_data.columns:
                adolescent = cn_data[
                    (cn_data['interview_status'] == 103) &
                    (cn_data['c2_age'] > 12) &
                    (cn_data['c2_age'] < 18)
                ]
                stats['adolescent_103'] = int(len(adolescent))
                # Calculate adult count as completed - adolescent
                stats['adult_count'] = stats['completed'] - stats['adolescent_103']

            if 'c2_gender' in cn_data.columns:
                g103 = cn_data[cn_data['interview_status'] == 103]['c2_gender']
                males   = g103.astype(str).str.lower().str.contains('male|m\b|1', regex=True).sum()
                females = g103.astype(str).str.lower().str.contains('female|f\b|2', regex=True).sum()
                total = males + females
                if total > 0:
                    male_pct   = (males / total) * 100
                    female_pct = (females / total) * 100
                    stats['gender_ratio'] = f"{males}:{females} | {male_pct:.0f}%:{female_pct:.0f}%"
                else:
                    stats['gender_ratio'] = 'N/A'

        # Missing geocodes (lat/lon missing)
        lat_col = 'hh_latitude'
        lon_col = 'hh_longitude'
        if lat_col in cn_data.columns and lon_col in cn_data.columns:
            missing = cn_data[[lat_col, lon_col]].isnull().any(axis=1).sum()
            stats['missing_geocodes'] = int(missing)

        # DBSCAN stats — prefer df_processed (has dbscan_cluster), fall back to df
        src = df_processed if df_processed is not None else df
        if 'dbscan_cluster' in src.columns:
            cn_src = src[src['CN'].astype(str) == str(cn)]
            if not cn_src.empty:
                valid_clusters = cn_src[cn_src['dbscan_cluster'] >= 0]['dbscan_cluster']
                stats['n_splits'] = int(valid_clusters.nunique()) if len(valid_clusters) > 0 else 0
                stats['noise']    = int((cn_src['dbscan_cluster'] == -1).sum())

        return stats

    # ── Data validations ────────────────────────────────────────────────────
    # Is Habitation / Is Tribal: Yes / No
    dv_habitation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_habitation.prompt = 'Select Yes or No'
    dv_habitation.promptTitle = 'Is Habitation / Is Tribal'
    ws.add_data_validation(dv_habitation)

    # is_Regular  (read-only info; still allow dropdown so it's visible)
    dv_regular = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_regular)

    # Checkbox: ☐ / ☑  for issue columns
    dv_checkbox = DataValidation(type="list", formula1='"\u2610,\u2611"', allow_blank=True)
    dv_checkbox.prompt      = 'Click to mark'
    dv_checkbox.promptTitle = 'Issue Flag'
    ws.add_data_validation(dv_checkbox)

    # ── Get sorted CNs ──────────────────────────────────────────────────────
    cns = sorted(df['CN'].unique()) if 'CN' in df.columns else []

    # ── Build sort criteria: Rural-Urban, is_Regular, Overall Count ─────────
    cn_sort_data = {}
    for cn in cns:
        cn_data = df[df['CN'] == cn]
        stats = _cn_stats(cn, cn_data)
        
        # Rural-Urban classification
        if not cn_data.empty:
            bi1c_val = str(cn_data['bi1c'].iloc[0]) if 'bi1c' in cn_data.columns and pd.notna(cn_data['bi1c'].iloc[0]) else ''
            is_urban = 'Urban' in bi1c_val
            rural_urban = 'Urban' if is_urban else 'Rural'
        else:
            rural_urban = 'Rural'
        
        # is_Regular: Yes before No
        is_regular = stats['is_regular']  # 'Yes' or 'No'
        
        # Overall count: higher first (use negative for descending sort)
        overall_count = stats['overall_count']
        
        cn_sort_data[cn] = (rural_urban, is_regular, overall_count)
    
    # Sort: Rural first, then by is_Regular (Yes before No), then by overall_count (descending)
    cns_sorted = sorted(cns, key=lambda x: (
        cn_sort_data[x][0] != 'Rural',        # Rural=False (0) comes before Urban=True (1)
        cn_sort_data[x][1] != 'Yes',          # Yes=False (0) comes before No=True (1)
        -cn_sort_data[x][2]                   # Negative for descending order (highest count first)
    ))

    # ── Write data rows ─────────────────────────────────────────────────────
    row_num = 3
    for cn in cns_sorted:
        cn_data = df[df['CN'] == cn]
        stats   = _cn_stats(cn, cn_data)

        # ── Server Data (cols 1-7) ──────────────────────────────────────────
        ws.cell(row=row_num, column=1, value=cn).border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")

        bi1c_val = ''
        for col_name, col_idx in [('bi1a', 2), ('bi1b', 3), ('bi1c', 4), ('bi1c_1', 5)]:
            val = cn_data[col_name].iloc[0] if col_name in cn_data.columns and not cn_data.empty else ''
            if col_name == 'bi1c':
                bi1c_val = str(val) if pd.notna(val) else ''
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        is_urban = 'Urban' in bi1c_val
        ru_cell  = ws.cell(row=row_num, column=6, value='Urban' if is_urban else 'Rural')
        ru_cell.alignment = Alignment(horizontal="center", vertical="center")
        ru_cell.border = thin_border

        if is_urban:
            ctype = 'MP' if bi1c_val in mp_bi1c_values else 'NMP'
        else:
            ctype = 'TRI' if 'TRI-' in bi1c_val else 'Non-TRI'
        type_cell = ws.cell(row=row_num, column=7, value=ctype)
        type_cell.alignment = Alignment(horizontal="center", vertical="center")
        type_cell.border = thin_border

        # ── Basic Info (cols 8-17) — auto-filled ───────────────────────────
        basic_values = [
            stats['is_regular'],
            stats['overall_count'],
            stats['completed'],
            f"{stats['response_rate']:.1f}%",
            stats['adult_count'],
            stats['adolescent_103'],
            stats['gender_ratio'],
            stats['missing_geocodes'],
            stats['n_splits'],
            stats['noise'],
        ]
        for offset, val in enumerate(basic_values):
            col_idx = basic_start + offset
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill      = basic_data_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border
            if col_idx == basic_start:  # is_regular — add dropdown
                dv_regular.add(cell)
            # Colour-code is_Regular cell
            if col_idx == basic_start:
                if val == 'Yes':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="276221", bold=True)
                elif val == 'No':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)

        # ── Your Correction (cols 17-23) — empty editable with VILCODE11 pre-filled ─────────────────
        is_habitation_col = correction_start  # Is Habitation is first correction column
        is_tribal_col = correction_start + 1  # Is Tribal is second correction column
        vilcode11_col = correction_start + 5   # VILCODE11 is sixth correction column
        for offset in range(len(correction_headers)):
            col_idx = correction_start + offset
            # Pre-populate VILCODE11 from vilcode CSV data if available, or checkbox symbol for habitation/tribal
            cell_value = ''
            if col_idx == vilcode11_col and 'VILCODE11' in cn_data.columns:
                # Get VILCODE11 value from the first row (should be same for all rows in cluster)
                vilcode11_val = cn_data['VILCODE11'].iloc[0] if len(cn_data) > 0 else ''
                cell_value = str(vilcode11_val) if pd.notna(vilcode11_val) else ''
            elif col_idx in (is_habitation_col, is_tribal_col):
                cell_value = '☐'  # Unchecked checkbox
            
            cell = ws.cell(row=row_num, column=col_idx, value=cell_value)
            cell.fill      = correction_data_fill
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border    = thin_border
            if col_idx in (is_habitation_col, is_tribal_col):  # Is Habitation / Is Tribal
                dv_habitation.add(cell)

        # ── Issues (cols 24-33) — ☐ checkbox + Notes + Total Issues formula ─
        # Checkbox range covers issue_start .. notes_col-1 (8 cols: Undesirable Split, Noise, Gender Bias, HH Address, Date, LGD, Monitoring, Non-response; excluding Notes & Total)
        checkbox_first_letter = get_column_letter(issue_start)
        checkbox_last_letter  = get_column_letter(notes_col - 1)
        for offset in range(len(issue_headers)):
            col_idx = issue_start + offset
            if col_idx == total_issues_col:        # Total Issues — COUNTIF formula
                formula = (
                    f'=COUNTIF({checkbox_first_letter}{row_num}:'
                    f'{checkbox_last_letter}{row_num},"\u2611")'
                )
                cell = ws.cell(row=row_num, column=col_idx, value=formula)
                cell.fill      = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                cell.font      = Font(bold=True, size=11, color="7F5200")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == notes_col:             # Notes column — free text
                cell = ws.cell(row=row_num, column=col_idx, value='')
                cell.fill      = issue_data_fill
                cell.font      = Font(size=12)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:                                  # Checkbox columns
                cell = ws.cell(row=row_num, column=col_idx, value='\u2610')  # ☐
                cell.fill      = issue_data_fill
                cell.font      = Font(size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                dv_checkbox.add(cell)
            cell.border = thin_border

        # ── Weightage Details (cols 32-36) — empty editable ─────────────────
        for offset in range(len(weightage_headers)):
            col_idx = weightage_start + offset
            cell = ws.cell(row=row_num, column=col_idx, value='')
            cell.fill      = weightage_data_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border

        row_num += 1

    # ── Fallback: empty rows if no CNs ─────────────────────────────────────
    if not cns_sorted:
        checkbox_first_letter = get_column_letter(issue_start)
        checkbox_last_letter  = get_column_letter(notes_col - 1)
        for _ in range(5):
            for col_idx in range(1, len(all_headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx, value='')
                cell.border = thin_border
                if col_idx in (correction_start, correction_start + 1):  # Is Habitation, Is Tribal
                    dv_habitation.add(cell)
                if issue_start <= col_idx < notes_col:
                    cell.value = '\u2610'
                    dv_checkbox.add(cell)
                if col_idx == total_issues_col:
                    formula = (
                        f'=COUNTIF({checkbox_first_letter}{row_num}:'
                        f'{checkbox_last_letter}{row_num},"\u2611")'
                    )
                    cell.value = formula
                    cell.fill  = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                    cell.font  = Font(bold=True, size=11, color="7F5200")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

    # ── Apply Conditional Formatting: Total Issues Column (Green → Yellow → Red) ──────────
    try:
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Color
        
        # Only apply if there are data rows
        if row_num > 3:
            total_issues_col_letter = get_column_letter(total_issues_col)
            color_scale = ColorScaleRule(
                start_type='min', start_color='00B050',   # Green for minimum (0 issues = good)
                mid_type='percentile', mid_value=50, mid_color='FFFF00',  # Yellow for 50th percentile
                end_type='max', end_color='FF0000'        # Red for maximum (8 issues = bad)
            )
            ws.conditional_formatting.add(f'{total_issues_col_letter}3:{total_issues_col_letter}{row_num - 1}', color_scale)
    except Exception as e:
        pass  # Skip if conditional formatting fails

    # ── Freeze header rows ──────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Save ────────────────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ── EUCLIDEAN DISTANCE ───────────────────────────────────────────────────────
def euclidean_distance(lat1, lon1, lat2, lon2):
    km_per_lat = 111.32
    km_per_lon = 111.32 * cos(radians((lat1 + lat2) / 2))
    return sqrt(((lat2 - lat1) * km_per_lat)**2 + ((lon2 - lon1) * km_per_lon)**2)

# ── LOAD AND TRANSFORM DATA (cached) ─────────────────────────────────────────
@st.cache_data(show_spinner="Loading & transforming data…")
def load_and_transform(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Load CSV/Excel and apply all column transformations. Cached on file content."""
    from io import BytesIO
    buf = BytesIO(file_bytes)
    if filename.endswith(".csv"):
        df = pd.read_csv(buf)
    else:
        df = pd.read_excel(buf)
    # Create 'index' column if it doesn't exist (for remark tracking)
    if 'index' not in df.columns:
        df.insert(0, 'index', range(len(df)))
    # Convert interview_status to integer
    df['interview_status'] = df['interview_status'].astype('Int64')
    # Initialize remark column
    df['remark'] = ''
    # Transformations
    df['concat'] = (
        df['bi1a'].astype(str) + '_' +
        df['bi1b'].astype(str) + '_' +
        df['bi1c'].astype(str) + '_' +
        df['bi1c_1'].astype(str)
    )
    df = df[df['interview_status'] != 105]
    uniques = sorted(df['concat'].dropna().unique())
    mapping = {val: f"{i+1:04d}" for i, val in enumerate(uniques)}
    df['code_4digit'] = df['concat'].map(mapping)
    df['CN'] = df['cluster_number'].astype(str).str[:2] + df['code_4digit']
    # Household extraction
    df['HH_CODE'] = (
        df['member_number']
        .astype(str)
        .str.split('.')
        .str[:-1]
        .str.join('.')
    )
    df['HH_NUM'] = (
        df.groupby('CN')['HH_CODE']
          .transform(lambda x: pd.factorize(x)[0] + 1)
    )
    df['HH_ID'] = df['CN'] + '.' + df['HH_NUM'].astype(str).str.zfill(3)
    df['MEM_ID'] = df['HH_ID'] + '.' + df['member_number'].astype(str).str.split('.').str[-1]
    df.drop(columns=['code_4digit', 'HH_CODE', 'HH_NUM'], inplace=True)
    # ── Normalize datetime columns to DD-MM-YYYY HH:MM:SS format ─────────
    for dt_col in ['start_datetime', 'end_datetime']:
        if dt_col in df.columns:
            df[dt_col] = pd.to_datetime(df[dt_col], errors='coerce', dayfirst=True)
            df[dt_col] = df[dt_col].dt.strftime('%d-%m-%Y %H:%M:%S')
    return df

# ── VALIDATE LATITUDE/LONGITUDE DATA ─────────────────────────────────────────
def validate_coordinates(df, lat_col, lon_col):
    """
    Validate latitude/longitude columns for suitability in geographic analysis.
    Returns: df_valid (cleaned), validation_report (dict with issues found)
    """
    validation_report = {
        'total_records': len(df),
        'valid_records': 0,
        'issues': {
            'non_numeric': [],
            'out_of_range': [],
            'missing': [],
            'invalid_precision': []
        },
        'excluded_indices': [],
        'summary': {}
    }
    
    if lat_col not in df.columns or lon_col not in df.columns:
        return df, validation_report

    df_validated = df.copy()
    df_validated['_is_valid_coordinate'] = False
    
    for idx, row in df_validated.iterrows():
        lat_val = row[lat_col]
        lon_val = row[lon_col]
        issues = []
        
        # Check for missing values
        if pd.isna(lat_val) or pd.isna(lon_val):
            issues.append('missing')
            validation_report['issues']['missing'].append(idx)
            continue
        
        # Try to convert to float
        try:
            lat_float = float(lat_val)
            lon_float = float(lon_val)
        except (ValueError, TypeError) as e:
            issues.append(f'non_numeric: {str(lat_val)[:30]}, {str(lon_val)[:30]}')
            validation_report['issues']['non_numeric'].append({
                'index': idx,
                'lat': str(lat_val)[:50],
                'lon': str(lon_val)[:50],
                'error': str(e)[:50]
            })
            continue
        
        # Check for valid latitude range (-90 to 90)
        if lat_float < -90 or lat_float > 90:
            issues.append(f'lat_out_of_range: {lat_float}')
            validation_report['issues']['out_of_range'].append({
                'index': idx,
                'lat': lat_float,
                'lon': lon_float,
                'reason': 'Latitude must be between -90 and 90'
            })
            continue
        
        # Check for valid longitude range (-180 to 180)
        if lon_float < -180 or lon_float > 180:
            issues.append(f'lon_out_of_range: {lon_float}')
            validation_report['issues']['out_of_range'].append({
                'index': idx,
                'lat': lat_float,
                'lon': lon_float,
                'reason': 'Longitude must be between -180 and 180'
            })
            continue
        
        # Check for suspicious precision (e.g., truncated coordinates)
        if abs(lat_float) < 0.001 and abs(lon_float) < 0.001:
            issues.append('near_zero')
        
        # Check for non-finite values
        if not (np.isfinite(lat_float) and np.isfinite(lon_float)):
            issues.append('non_finite')
            continue
        
        # All checks passed - mark as valid
        df_validated.loc[idx, '_is_valid_coordinate'] = True
        validation_report['valid_records'] += 1
    
    # Calculate summary statistics
    validation_report['summary'] = {
        'total_invalid': len(df) - validation_report['valid_records'],
        'invalid_percent': round(100 * (len(df) - validation_report['valid_records']) / len(df), 2),
        'non_numeric_count': len(validation_report['issues']['non_numeric']),
        'out_of_range_count': len(validation_report['issues']['out_of_range']),
        'missing_count': len(validation_report['issues']['missing'])
    }

    # Coerce lat/lon columns to numeric dtype so downstream .mean()/.std() etc.
    # work correctly even when the original CSV loaded the column as object (string)
    # due to a single bad value like '10.9852207w'. Invalid values become NaN.
    df_validated[lat_col] = pd.to_numeric(df_validated[lat_col], errors='coerce')
    df_validated[lon_col] = pd.to_numeric(df_validated[lon_col], errors='coerce')

    return df_validated, validation_report

# ── CENTROID ─────────────────────────────────────────────────────────────────
def calculate_centroid(df, lat_col, lon_col):
    valid = df[[lat_col, lon_col]].dropna()
    if len(valid) == 0:
        return None, None
    return valid[lat_col].mean(), valid[lon_col].mean()


# ── FIND NEAREST CLUSTERS (cached) ───────────────────────────────────────────
@st.cache_data(show_spinner=False)
def _find_nearest_clusters_cached(_centroids_json: str, sel_cn: str, distance_metric: str, top_n: int):
    """Cached NN computation. _centroids_json is a JSON string of {CN: [lat,lon]}."""
    centroids = _orjson.loads(_centroids_json.encode()) if _HAS_ORJSON else json.loads(_centroids_json)
    sel = centroids.get(sel_cn)
    if sel is None:
        return pd.DataFrame()
    sel_lat, sel_lon = sel
    rows = []
    for cn, (c_lat, c_lon) in centroids.items():
        if cn == sel_cn:
            continue
        dist = haversine_distance(sel_lat, sel_lon, c_lat, c_lon) if distance_metric == 'haversine' \
               else euclidean_distance(sel_lat, sel_lon, c_lat, c_lon)
        rows.append({'CN': cn, 'Centroid_Lat': round(c_lat, 6), 'Centroid_Lon': round(c_lon, 6),
                     'Distance_km': round(dist, 2)})
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).sort_values('Distance_km').head(top_n).reset_index(drop=True)

def find_nearest_clusters(df, selected_cn, lat_col, lon_col, distance_metric='haversine', top_n=10):
    """Wrapper that builds centroids dict once, then delegates to cached function."""
    centroids = {}
    for cn, grp in df.groupby('CN'):
        # Filter to only valid coordinates if validation marker exists
        if '_is_valid_coordinate' in grp.columns:
            grp_valid = grp[grp['_is_valid_coordinate'] == True]
        else:
            grp_valid = grp
        
        valid = grp_valid[[lat_col, lon_col]].dropna()
        if len(valid) > 0:
            centroids[cn] = [round(valid[lat_col].mean(), 6), round(valid[lon_col].mean(), 6)]
    result = _find_nearest_clusters_cached(json.dumps(centroids, sort_keys=True), selected_cn, distance_metric, top_n)
    if result.empty:
        return result
    # Add Count column (not part of cached computation since it changes with merges)
    counts = df.groupby('CN').size().to_dict()
    result['Count'] = result['CN'].map(counts).fillna(0).astype(int)
    return result

# --- filter_dataframe utility (from Streamlit blog) ---
from pandas.api.types import (
    is_categorical_dtype,
    is_datetime64_any_dtype,
    is_numeric_dtype,
    is_object_dtype,
)

def filter_dataframe(df: pd.DataFrame, key=None) -> pd.DataFrame:
    """
    Adds a UI on top of a dataframe to let viewers filter columns

    Args:
        df (pd.DataFrame): Original dataframe
        key (str, optional): Unique key for Streamlit widgets

    Returns:
        pd.DataFrame: Filtered dataframe
    """
    modify = st.checkbox("Add filters", key=key)
    if not modify:
        return df

    df = df.copy()

    # Try to convert datetimes into a standard format (datetime, no timezone)
    for col in df.columns:
        if is_object_dtype(df[col]):
            try:
                df[col] = pd.to_datetime(df[col])
            except Exception:
                pass
        if is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)

    modification_container = st.container()
    with modification_container:
        to_filter_columns = st.multiselect("Filter dataframe on", df.columns)

        for column in to_filter_columns:
            left, right = st.columns((1, 20))
            left.write("↳")
            if is_categorical_dtype(df[column]) or df[column].nunique() < 10:
                user_cat_input = right.multiselect(
                    f"Values for {column}",
                    df[column].unique(),
                    default=list(df[column].unique()),
                )
                df = df[df[column].isin(user_cat_input)]
            elif is_numeric_dtype(df[column]):
                _min = float(df[column].min())
                _max = float(df[column].max())
                step = (_max - _min) / 100
                user_num_input = right.slider(
                    f"Values for {column}",
                    min_value=_min,
                    max_value=_max,
                    value=(_min, _max),
                    step=step,
                )
                df = df[df[column].between(*user_num_input)]
            elif is_datetime64_any_dtype(df[column]):
                user_date_input = right.date_input(
                    f"Values for {column}",
                    value=(
                        df[column].min(),
                        df[column].max(),
                    ),
                )
                if len(user_date_input) == 2:
                    user_date_input = tuple(map(pd.to_datetime, user_date_input))
                    start_date, end_date = user_date_input
                    df = df.loc[df[column].between(start_date, end_date)]
            else:
                user_text_input = right.text_input(
                    f"Substring or regex in {column}",
                )
                if user_text_input:
                    df = df[df[column].astype(str).str.contains(user_text_input)]

    return df

# ── GEOCODING FUNCTION FOR MAP SEARCH ────────────────────────────────────────
def search_location(address: str, api_key: str = "6995e79e5b99d534908138osab8a2af"):
    """
    Search for a location using geocode.maps.co API.
    
    Args:
        address: Location name or address
        api_key: geocode.maps.co API key
        
    Returns:
        Tuple (latitude, longitude) or None if not found
    """
    if not address.strip():
        return None
    
    try:
        url = "https://geocode.maps.co/search"
        params = {
            "q": address,
            "api_key": api_key,
            "format": "json"
        }
        
        response = requests.get(url, params=params, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        if isinstance(data, list) and len(data) > 0:
            result = data[0]
            return (float(result['lat']), float(result['lon']), result.get('display_name', address))
        
    except Exception:
        pass
    
    return None

# ── ADD CUSTOM SEARCH CONTROL TO MAP ──────────────────────────────────────────
def add_search_control_to_map(m, api_key: str = "6995e79e5b99d534908138osab8a2af"):
    """
    Add a custom search control to folium map using HTML/CSS/JS.
    Uses geocode.maps.co API for location search.
    """
    # Create custom HTML search control
    search_html = f"""
    <div id="search-control" style="
        position: fixed;
        top: 10px;
        right: 10px;
        width: 300px;
        z-index: 1000;
        background: white;
        padding: 10px;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.2);
        font-family: Arial, sans-serif;
    ">
        <input 
            type="text" 
            id="search-input" 
            placeholder="🔍 Search location..." 
            style="
                width: 100%;
                padding: 8px;
                border: 1px solid #ccc;
                border-radius: 3px;
                font-size: 14px;
                box-sizing: border-box;
            "
        />
        <div id="search-results" style="
            margin-top: 5px;
            max-height: 200px;
            overflow-y: auto;
            border: 1px solid #ddd;
            border-radius: 3px;
            display: none;
        "></div>
    </div>
    
    <script>
        const apiKey = "{api_key}";
        let currentMarker = null;
        
        document.getElementById('search-input').addEventListener('input', async function(e) {{
            const query = e.target.value;
            if (query.length < 2) {{
                document.getElementById('search-results').style.display = 'none';
                return;
            }}
            
            try {{
                const response = await fetch(`https://geocode.maps.co/search?q=${{encodeURIComponent(query)}}&api_key=${{apiKey}}&format=json`);
                const data = await response.json();
                
                const resultsDiv = document.getElementById('search-results');
                resultsDiv.innerHTML = '';
                
                if (data.length > 0) {{
                    resultsDiv.style.display = 'block';
                    data.slice(0, 5).forEach(result => {{
                        const item = document.createElement('div');
                        item.style.cssText = 'padding: 8px; border-bottom: 1px solid #eee; cursor: pointer; hover: background-color: #f0f0f0;';
                        item.textContent = result.display_name;
                        item.onclick = function() {{
                            map.setView([result.lat, result.lon], 13);
                            if (currentMarker) map.removeLayer(currentMarker);
                            currentMarker = L.marker([result.lat, result.lon]).addTo(map).bindPopup(result.display_name);
                            document.getElementById('search-input').value = result.display_name;
                            resultsDiv.style.display = 'none';
                        }};
                        resultsDiv.appendChild(item);
                    }});
                }}
            }} catch(e) {{
                console.error('Search error:', e);
            }}
        }});
    </script>
    """
    
    # Add custom HTML to map
    m.get_root().html.add_child(folium.Element(search_html))
    return m

# ──────────────────────────────────────────────────────────────────────────────
# FRAGMENT: Reconsider Flag/Move Form (for fast mode switching)
# ──────────────────────────────────────────────────────────────────────────────
@st.fragment
def reconsider_flag_move_fragment(selected_cluster, cn_filtered_df, df_processed, lat_col_name):
    """Fragment for quick Flag↔Move mode switching without full app rerun."""
    
    st.markdown("**Reconsider (✅ GOOD & 🚩 Flag / 🚚 Move)**")
    
    reconsider_mode = st.radio(
        "Action mode",
        options=["Flag", "Move"],
        index=0,
        horizontal=True,
        key="reconsider_mode_radio"
    )
    
    if reconsider_mode == "Flag":
        st.caption("Sort by latitude, assign select_id, mark rows: flagged → FLAG, others → GOOD")
        
        with st.form("reconsider_form"):
            feedback_input = st.text_area(
                "Enter select_id to mark as FLAG",
                height=150,
                placeholder="Example:\n1\n3\n5:7\n\nLeave empty to mark\nall rows as GOOD",
                key="feedback_textarea"
            )
            
            reconsider_remarks = st.text_area(
                "Remarks (optional)",
                placeholder="Enter reason for reconsideration...",
                height=80,
                key="reconsider_remarks"
            )
            
            good_clicked = st.form_submit_button("✅ Apply", use_container_width=True)
        
        # Process feedback and GOOD button (includes Reconsider)
        if good_clicked:
            # Get all select_ids in current cluster
            all_select_ids_in_cluster = cn_filtered_df['select_id'].tolist()
            
            # Parse feedback input for FLAG rows
            flag_select_ids = []
            if feedback_input.strip():
                for line in feedback_input.strip().split('\n'):
                    line = line.strip()
                    if ':' in line:
                        # Handle range format like "1:3"
                        try:
                            start, end = line.split(':')
                            start, end = int(start.strip()), int(end.strip())
                            flag_select_ids.extend(range(start, end + 1))
                        except ValueError:
                            st.warning(f"Invalid range format: {line}")
                    elif line.isdigit():
                        flag_select_ids.append(int(line))
            
            # Mark all rows in cluster: FLAG for feedback rows, GOOD for others
            remarks_changes = {}
            for select_id in all_select_ids_in_cluster:
                idx = cn_filtered_df[cn_filtered_df['select_id'] == select_id]['index'].values[0]
                if select_id in flag_select_ids:
                    st.session_state.remarks_dict[idx] = "FLAG"
                    remarks_changes[int(idx)] = "FLAG"
                else:
                    st.session_state.remarks_dict[idx] = "GOOD"
                    remarks_changes[int(idx)] = "GOOD"
            
            # ── Reconsider: sort by latitude & reassign select_id ──
            if lat_col_name and selected_cluster:
                st.session_state.reconsidered_cns.add(str(selected_cluster))
                st.session_state.steps_tracker.append({
                    'S.No': len(st.session_state.steps_tracker) + 1,
                    'Step': 'Reconsider',
                    'Details': f"CN {selected_cluster}: Sorted by latitude, select_id reassigned",
                    'Remark': reconsider_remarks if reconsider_remarks.strip() else '-',
                    '_op_type': 'reconsider',
                    '_op_data': {
                        'cn': str(selected_cluster),
                        'remarks': {str(k): v for k, v in remarks_changes.items()}
                    }
                })
            
            # Update df_processed with remarks
            for idx, remark in st.session_state.remarks_dict.items():
                df_processed.loc[df_processed['index'] == idx, 'remark'] = remark
            
            st.success(f"✅ Marked {len(all_select_ids_in_cluster)} rows:\n{len(flag_select_ids)} FLAG | {len(all_select_ids_in_cluster) - len(flag_select_ids)} GOOD")
            st.rerun(scope="app")
    
    else:  # Move mode
        st.caption("Move selected rows to another CN (HH_ID / MEM_ID reassigned to avoid collision)")
        
        with st.form("move_form"):
            move_select_input = st.text_area(
                "Enter select_id to MOVE",
                height=150,
                placeholder="Example:\n1\n3\n5:7",
                key="move_select_textarea"
            )
            
            # Build target CN list (all CNs except current)
            move_target_cns = sorted(
                [cn for cn in df_processed['CN'].astype(str).unique()
                 if cn != str(selected_cluster)]
            )
            move_target_cn = st.selectbox(
                "Move to CN",
                options=[""] + move_target_cns,
                index=0,
                key="move_target_cn_select"
            )
            
            move_remarks = st.text_area(
                "Remarks (optional)",
                placeholder="Enter reason for move...",
                height=80,
                key="move_remarks"
            )
            
            move_clicked = st.form_submit_button("🚚 Move", use_container_width=True)
        
        if move_clicked:
            # Parse select_ids
            move_select_ids = []
            if move_select_input.strip():
                for line in move_select_input.strip().split('\n'):
                    line = line.strip()
                    if ':' in line:
                        try:
                            start, end = line.split(':')
                            start, end = int(start.strip()), int(end.strip())
                            move_select_ids.extend(range(start, end + 1))
                        except ValueError:
                            st.warning(f"Invalid range format: {line}")
                    elif line.isdigit():
                        move_select_ids.append(int(line))
            
            if not move_select_ids:
                st.error("⚠️ Enter at least one select_id to move.")
            elif not move_target_cn:
                st.error("⚠️ Select a target CN.")
            else:
                # Resolve select_ids → stable index values
                move_indices = []
                for sid in move_select_ids:
                    match = cn_filtered_df[cn_filtered_df['select_id'] == sid]
                    if len(match) > 0:
                        move_indices.append(int(match['index'].values[0]))
                    else:
                        st.warning(f"select_id {sid} not found in current cluster — skipped")
                
                if move_indices:
                    # Store move operation (applied on next rerun)
                    move_op_data = {
                        'indices': move_indices,
                        'source_cn': str(selected_cluster),
                        'target_cn': str(move_target_cn)
                    }
                    st.session_state.move_operations.append(move_op_data)
                    
                    # Mark moved rows with remark showing source CN
                    move_remarks_changes = {}
                    for midx in move_indices:
                        st.session_state.remarks_dict[midx] = f"moved from {selected_cluster}"
                        move_remarks_changes[str(int(midx))] = f"moved from {selected_cluster}"
                    
                    # Also reconsider current cluster (re-sort / re-assign select_id)
                    if lat_col_name and selected_cluster:
                        st.session_state.reconsidered_cns.add(str(selected_cluster))
                    
                    # Check if source CN will be empty after move (all rows moved out)
                    _moved_indices_set = set(move_indices)
                    _move_remaining = [idx for idx in cn_filtered_df['index'].values if int(idx) not in _moved_indices_set]
                    _move_auto_removed_src = None
                    if len(_move_remaining) == 0:
                        _move_auto_removed_src = str(selected_cluster)

                    st.session_state.steps_tracker.append({
                        'S.No': len(st.session_state.steps_tracker) + 1,
                        'Step': 'Move',
                        'Details': f"CN {selected_cluster} → {move_target_cn}: {len(move_indices)} point(s)",
                        'Remark': move_remarks if move_remarks.strip() else '-',
                        '_op_type': 'move',
                        '_op_data': {
                            'move_op': move_op_data,
                            'reconsidered_cn': str(selected_cluster),
                            'remarks': move_remarks_changes,
                            'auto_removed_source_cn': _move_auto_removed_src
                        }
                    })
                    st.success(f"🚚 Moved {len(move_indices)} row(s) from CN {selected_cluster} → {move_target_cn}")
                    st.rerun(scope="app")
                else:
                    st.error("⚠️ No valid select_ids found to move.")

st.set_page_config(page_title="Cluster Validator", layout="wide")

# ── FLUSH PENDING RESTORE (before any widgets are instantiated) ──────────────
# load_progress() stores data in _pending_restore and calls st.rerun().
# On the fresh run, all keys (including widget-bound ones) are safe to set here
# because no widgets have been rendered yet.
if '_pending_restore' in st.session_state:
    _restore = st.session_state.pop('_pending_restore')
    # Restore form state (widget-bound keys are safe because no widgets rendered yet)
    _form = _restore.get('form_state', {})
    for _k, _v in _form.items():
        st.session_state[_k] = _v
    st.session_state.form_submitted = True
    # Restore operations
    _ops = _restore.get('operations', {})
    st.session_state.steps_tracker = _ops.get('steps_tracker', [])
    st.session_state.removed_cns = _ops.get('removed_cns', [])
    st.session_state.merge_operations = [tuple(m) for m in _ops.get('merge_operations', [])]
    st.session_state.break_operations = _ops.get('break_operations', [])
    st.session_state.move_operations = _ops.get('move_operations', [])
    st.session_state.reconsidered_cns = set(_ops.get('reconsidered_cns', []))
    st.session_state.remarks_dict = {int(k): v for k, v in _ops.get('remarks_dict', {}).items()}
    st.session_state._saved_districts_to_remove = _ops.get('districts_to_remove', [])
    # Restore widget-input keys so st.number_input / st.text_input pick them up
    _widget_map = {
        'rural_clusters': 'rural_clusters_input',
        'urban_clusters': 'urban_clusters_input',
        'non_tribal_clusters': 'non_tribal_clusters_input',
        'non_mp_clusters': 'non_mp_clusters_input',
        'mp_cities': 'mp_cities_input',
    }
    for _derived_key, _widget_key in _widget_map.items():
        if _derived_key in _form:
            st.session_state[_widget_key] = _form[_derived_key]
    # Manual entry mode inputs
    if _form.get('manual_entry_mode', False):
        if 'tribal_clusters' in _form:
            st.session_state['manual_tribal_input'] = _form['tribal_clusters']
        if 'million_plus_clusters' in _form:
            st.session_state['manual_million_plus_input'] = _form['million_plus_clusters']
    st.session_state.dbscan_labels = None
    st.session_state.last_dbscan_cluster = None
    st.toast('Progress restored successfully!', icon='✅')

st.title("Cluster Validator")

# Initialize session state for map interactions
if 'last_cluster' not in st.session_state:
    st.session_state.last_cluster = None

# Initialize session state for action features (EARLY - before cluster selection)
if 'removed_cns' not in st.session_state:
    st.session_state.removed_cns = []
if 'merge_operations' not in st.session_state:
    st.session_state.merge_operations = []  # List of (source_cn, target_cn)
if 'reconsidered_cns' not in st.session_state:
    st.session_state.reconsidered_cns = set()
if 'steps_tracker' not in st.session_state:
    st.session_state.steps_tracker = []
if 'map_filter_mode' not in st.session_state:
    st.session_state.map_filter_mode = 'Status 103 Only'
if 'remarks_dict' not in st.session_state:
    st.session_state.remarks_dict = {}
if 'form_submitted' not in st.session_state:
    st.session_state.form_submitted = False
if 'break_operations' not in st.session_state:
    st.session_state.break_operations = []
if 'move_operations' not in st.session_state:
    st.session_state.move_operations = []  # List of {indices: [...], source_cn, target_cn}
if 'dbscan_labels' not in st.session_state:
    st.session_state.dbscan_labels = None  # Store DBSCAN results for persistence across reruns
if '_saved_districts_to_remove' not in st.session_state:
    st.session_state._saved_districts_to_remove = []
if 'geojson_data' not in st.session_state:
    st.session_state.geojson_data = None
if '_geojson_cache_key' not in st.session_state:
    st.session_state._geojson_cache_key = None
if 'geojson_vilcode_index' not in st.session_state:
    st.session_state.geojson_vilcode_index = {}
if 'vilcode11' not in st.session_state:
    st.session_state.vilcode11 = None
if 'vilcode_csv_data' not in st.session_state:
    st.session_state.vilcode_csv_data = None
if 'vilcode11_set' not in st.session_state:
    st.session_state.vilcode11_set = set()
if 'vilcode11_manual_set' not in st.session_state:
    st.session_state.vilcode11_manual_set = set()

# ── CLUSTER CONFIGURATION ────────────────────────────────────────────────────
st.subheader("Planned Cluster Allocation")

# Checkbox for manual entry mode - OUTSIDE FORM for dynamic updates
manual_entry_mode = st.checkbox(
    "Manually enter Tribal and Million Plus reference data",
    key="manual_entry_mode",
    help="Enable manual entry for Tribal and Million Plus clusters"
)


with st.form("cluster_alloc_form"):
    col1, col2, col3, col4, col5 = st.columns(5)

    with col1:
        rural_clusters = st.number_input("Rural", min_value=0, step=1, help="Number of rural clusters", key="rural_clusters_input")

    with col2:
        urban_clusters = st.number_input("Urban", min_value=0, step=1, help="Number of urban clusters", key="urban_clusters_input")

    with col3:
        non_tribal_clusters = st.number_input("Non-Tribal", min_value=0, step=1, help="Number of non-tribal clusters", key="non_tribal_clusters_input")

    with col4:
        non_mp_clusters = st.number_input("Non-MP", min_value=0, step=1, help="Number of non-MP clusters", key="non_mp_clusters_input")

    with col5:
        mp_cities = st.number_input("Number of MP Cities in the state", min_value=0, step=1, help="Number of MP cities in the state", key="mp_cities_input")
    
    # Conditional: Manual entry or calculated values
    if manual_entry_mode:
        st.markdown("**Manual Reference Data Entry**")
        col_tribal, col_mp = st.columns(2)
        
        with col_tribal:
            manual_tribal = st.number_input(
                "Tribal Clusters", 
                min_value=0, 
                step=1, 
                help="Number of tribal clusters (manual entry)",
                key="manual_tribal_input"
            )
        
        with col_mp:
            manual_million_plus = st.number_input(
                "Million Plus Clusters", 
                min_value=0, 
                step=1, 
                help="Number of million plus clusters (manual entry)",
                key="manual_million_plus_input"
            )
        tribal_value = manual_tribal
        mp_value = manual_million_plus
    else:
        # Calculate values
        tribal_value = int(rural_clusters - non_tribal_clusters) if rural_clusters >= non_tribal_clusters else 0
        mp_value = int(urban_clusters - non_mp_clusters) if urban_clusters >= non_mp_clusters else 0
        
        st.markdown("**Calculated Reference Data**")
        col_tribal, col_mp = st.columns(2)
        
        with col_tribal:
            st.metric("Tribal Clusters", tribal_value, help="Calculated as: Rural - Non-Tribal")
        
        with col_mp:
            st.metric("Million Plus Clusters", mp_value, help="Calculated as: Urban - Non-MP")

    st.markdown("**Data Classification Confirmations**")
    confirm_tribal = st.checkbox(
        'I confirm data contains `"TRI-"` keyword for tribal cluster classification',
        key="confirm_tribal"
    )
    confirm_urban = st.checkbox(
        'I confirm data contains `"Urban"` keyword for urban cluster classification',
        key="confirm_urban"
    )
    confirm_mp = st.checkbox(
        'I confirm data contains `"Urban"` & `"Mc"` keywords for MP cities classification or no MP cities in the state ',
        key="confirm_mp"
    )

    submitted = st.form_submit_button("Submit Allocation")

if submitted:
    # persist values in session_state for use elsewhere in the app
    st.session_state.rural_clusters = int(rural_clusters)
    st.session_state.urban_clusters = int(urban_clusters)
    st.session_state.non_tribal_clusters = int(non_tribal_clusters)
    st.session_state.non_mp_clusters = int(non_mp_clusters)
    st.session_state.mp_cities = int(mp_cities)
    
    # Store tribal and million plus values (determined by manual_entry_mode)
    st.session_state.tribal_clusters = tribal_value
    st.session_state.million_plus_clusters = mp_value
    
    st.session_state.form_submitted = True
    # Note: confirm_tribal, confirm_urban, confirm_mp are already stored by Streamlit via key="..."
    
    # Track allocation submission in action tracker
    alloc_detail = (f"Rural={int(rural_clusters)}, Urban={int(urban_clusters)}, "
                    f"Non-Tribal={int(non_tribal_clusters)}, Non-MP={int(non_mp_clusters)}, "
                    f"MP Cities={int(mp_cities)}, Tribal={tribal_value}, MP={mp_value}")
    # Avoid duplicate if re-submitted with same values
    already_alloc = any(s.get('_op_type') == 'allocation' for s in st.session_state.steps_tracker)
    if not already_alloc:
        st.session_state.steps_tracker.append({
            'S.No': len(st.session_state.steps_tracker) + 1,
            'Step': 'Planned Allocation',
            'Details': alloc_detail,
            'Remark': '-',
            '_op_type': 'allocation',
            '_op_data': {
                'rural_clusters': int(rural_clusters),
                'urban_clusters': int(urban_clusters),
                'non_tribal_clusters': int(non_tribal_clusters),
                'non_mp_clusters': int(non_mp_clusters),
                'mp_cities': int(mp_cities),
                'tribal_clusters': tribal_value,
                'million_plus_clusters': mp_value,
            }
        })
    else:
        # Update the existing allocation step
        for s in st.session_state.steps_tracker:
            if s.get('_op_type') == 'allocation':
                s['Details'] = alloc_detail
                s['_op_data'] = {
                    'rural_clusters': int(rural_clusters),
                    'urban_clusters': int(urban_clusters),
                    'non_tribal_clusters': int(non_tribal_clusters),
                    'non_mp_clusters': int(non_mp_clusters),
                    'mp_cities': int(mp_cities),
                    'tribal_clusters': tribal_value,
                    'million_plus_clusters': mp_value,
                }
                break
    st.success("Cluster allocation submitted")

if st.session_state.form_submitted:
    uploaded_file = st.file_uploader("Upload CSV or Excel file", type=["csv", "xlsx"], key="file_uploader_main")
    st.info("👆 Upload a CSV or Excel file to get started.")
    
    # ── GEOJSON UPLOADER & VILCODE11 CSV INPUT ──────────────────────────────
    col_geo1, col_geo2 = st.columns([1, 1])

    with col_geo1:
        geojson_file = st.file_uploader("📍 Upload GeoJSON file (optional)", type=["geojson", "json"], key="geojson_uploader")

    with col_geo2:
        vilcode_csv_file = st.file_uploader(
            "📋 Upload VILCODE11 CSV (CN/bi1a/bi1b/bi1c/bi1c_1/VILCODE11)",
            type=["csv", "xlsx"],
            key="vilcode_csv_uploader",
            help="Upload a file with a VILCODE11 column to filter the GeoJSON layer"
        )

    # ── Helper: Normalize GeoJSON properties to uppercase keys (case-insensitive handling) ─────
    def normalize_geojson_properties(geojson_obj):
        """Recursively normalizes all GeoJSON feature properties to UPPERCASE keys."""
        if isinstance(geojson_obj, dict):
            if geojson_obj.get('type') == 'Feature' and 'properties' in geojson_obj:
                # Normalize properties keys to uppercase
                original_props = geojson_obj['properties']
                normalized_props = {k.upper(): v for k, v in original_props.items()} if original_props else {}
                geojson_obj['properties'] = normalized_props
            elif geojson_obj.get('type') == 'FeatureCollection' and 'features' in geojson_obj:
                # Recursively normalize each feature
                for feature in geojson_obj['features']:
                    normalize_geojson_properties(feature)
        return geojson_obj
    
    # ── Process GeoJSON upload (with caching for fast processing) ────────────
    if geojson_file is not None:
        try:
            # Generate cache key from file name and size
            geojson_file.seek(0, 2)  # Seek to end
            file_size = geojson_file.tell()
            geojson_file.seek(0)  # Reset to start
            geojson_cache_key = f"{geojson_file.name}_{file_size}"
            
            # Check if already cached in session state
            if st.session_state.get('_geojson_cache_key') != geojson_cache_key:
                file_bytes = geojson_file.read()
                # Use msgspec for lightning-fast parsing (2-3x faster than orjson for large files)
                if _HAS_MSGSPEC:
                    geojson_data = msgspec.json.decode(file_bytes)
                else:
                    geojson_data = json.loads(file_bytes.decode('utf-8'))
                # ✅ Normalize all property keys to UPPERCASE (handles mixed-case GeoJSON)
                geojson_data = normalize_geojson_properties(geojson_data)
                st.session_state.geojson_data = geojson_data
                st.session_state._geojson_cache_key = geojson_cache_key
                
                # ✅ Build VILCODE11 index for fast filtering (O(1) lookup instead of O(n) filter)
                st.session_state.geojson_vilcode_index = {}
                for feature in geojson_data.get('features', []):
                    vilcode = str(feature.get('properties', {}).get('VILCODE11', '')).strip().split('.')[0]
                    if vilcode:
                        if vilcode not in st.session_state.geojson_vilcode_index:
                            st.session_state.geojson_vilcode_index[vilcode] = []
                        st.session_state.geojson_vilcode_index[vilcode].append(feature)
                
                st.success(f"✅ GeoJSON loaded with {len(st.session_state.geojson_vilcode_index)} unique village codes indexed (msgspec fast parser)")
            else:
                st.success("✅ GeoJSON loaded from cache")
        except (msgspec.DecodeError if _HAS_MSGSPEC else json.JSONDecodeError, ValueError, UnicodeDecodeError):
            st.error("❌ Invalid GeoJSON file. Please upload a valid GeoJSON file.")
            st.session_state.geojson_data = None
            st.session_state._geojson_cache_key = None
    else:
        st.session_state.geojson_data = None

    # ── Process VILCODE11 CSV upload ──────────────────────────────────────────
    if vilcode_csv_file is not None:
        try:
            vilcode_csv_file.seek(0)
            if vilcode_csv_file.name.endswith('.xlsx'):
                _vdf = pd.read_excel(vilcode_csv_file)
            else:
                _vdf = pd.read_csv(vilcode_csv_file)

            if 'VILCODE11' not in _vdf.columns:
                st.error("❌ Uploaded file must contain a 'VILCODE11' column.")
                st.session_state.vilcode_csv_data = None
                st.session_state.vilcode11_set = set()
            else:
                # Convert to strings and drop blanks/NaN — handles int/float columns
                _codes = (
                    _vdf['VILCODE11']
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .str.split('.').str[0]   # remove .0 from float-read ints
                    .replace('', pd.NA)
                    .dropna()
                    .unique()
                    .tolist()
                )
                st.session_state.vilcode_csv_data = _vdf
                st.session_state.vilcode11_set = set(_codes)
                st.success(f"✅ VILCODE11 CSV loaded — {len(_codes)} unique codes from {len(_vdf)} rows")
        except Exception as e:
            st.error(f"❌ Error reading VILCODE11 file: {e}")
            st.session_state.vilcode_csv_data = None
            st.session_state.vilcode11_set = set()
    else:
        st.session_state.vilcode_csv_data = None
        st.session_state.vilcode11_set = set()

    # ── Manual VILCODE11 multi-feeder ─────────────────────────────────────────
    with st.form("vilcode11_manual_form", border=True):
        st.caption("🔢 **Manual VILCODE11 overlay** — adds a separate layer on the main map")
        _manual_raw = st.text_area(
            "VILCODE11 codes",
            placeholder="One code per line, or comma-separated, e.g.\n403542\n403453\n403457",
            height=110,
            key="vilcode11_manual_input",
            help="These codes are overlaid as a separate orange layer on the main cluster map, independent of the CSV."
        )
        _manual_submit = st.form_submit_button("📌 Apply to Map", use_container_width=True)

    if _manual_submit:
        import re as _re
        if _manual_raw.strip():
            _manual_codes = set(_re.split(r'[,\n\r\s]+', _manual_raw.strip()))
            _manual_codes = {c.strip().split('.')[0] for c in _manual_codes if c.strip()}
            st.session_state.vilcode11_manual_set = _manual_codes
            st.success(f"✅ {len(_manual_codes)} VILCODE11 code(s) queued for main map overlay")
        else:
            st.session_state.vilcode11_manual_set = set()
            st.info("ℹ️ Manual overlay cleared.")

    # ── RESUME PROGRESS (shown below file uploader when file is uploaded) ─────
    if uploaded_file is not None:
        _current_file_md5 = calculate_md5_checksum(uploaded_file)
        if _current_file_md5:
            matching_saves = get_saved_progress_files(file_md5=_current_file_md5)
            if matching_saves:
                with st.expander(f"📂 Resume Saved Progress ({len(matching_saves)} save(s) found for this file)", expanded=True):
                    st.success(f"✅ Found {len(matching_saves)} saved progress file(s) matching MD5: `{_current_file_md5[:12]}...`")
                    save_options = {
                        f"{s['saved_at']} — {s['steps_count']} step(s) — {s['file_name']}": s['path']
                        for s in matching_saves
                    }
                    selected_save = st.selectbox(
                        "Select save to resume",
                        options=[""] + list(save_options.keys()),
                        index=0,
                        key="resume_save_select"
                    )
                    if selected_save and st.button("▶️ Resume Selected Progress", use_container_width=True, key="resume_btn"):
                        load_progress(save_options[selected_save])
                        st.rerun()
    
    # ── SAVE PROGRESS BUTTON ──────────────────────────────────────────────────
    if uploaded_file is not None and st.session_state.get('steps_tracker'):
        _save_file_md5 = calculate_md5_checksum(uploaded_file)
        if _save_file_md5:
            if st.button("💾 Save Progress", use_container_width=True, key="save_progress_btn"):
                filepath, filename = save_progress(_save_file_md5, uploaded_file.name)
                st.success(f"✅ Progress saved: **{filename}**")
else:
    st.info("👆 Please submit the Planned Cluster Allocation form to proceed")
    uploaded_file = None

if uploaded_file is not None:

    # ── Load data (cached on file content) ───────────────────────────────────
    _file_bytes = uploaded_file.getvalue()
    df = load_and_transform(_file_bytes, uploaded_file.name).copy()

    # Apply stored remarks from session state to df
    for idx, remark in st.session_state.remarks_dict.items():
        df.loc[df['index'] == idx, 'remark'] = remark

    # ── Detect lat/lon column names once (used in both panels & NN analysis) ──
    # Always look for hh_latitude and hh_longitude first
    lat_col_name = 'hh_latitude' if 'hh_latitude' in df.columns else None
    lon_col_name = 'hh_longitude' if 'hh_longitude' in df.columns else None
    
    # ── VALIDATE LATITUDE/LONGITUDE DATA ──────────────────────────────────────
    if lat_col_name and lon_col_name:
        df, coord_validation = validate_coordinates(df, lat_col_name, lon_col_name)
        
        # Display validation report
        if coord_validation['summary']['total_invalid'] > 0:
            st.error("⚠️ **Geographic Data Quality Issue Detected**")
            
            # Display metrics in 4-column layout
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Valid Records", coord_validation['valid_records'], 
                         f"of {coord_validation['total_records']}")
            with col2:
                st.metric("Invalid Records", coord_validation['summary']['total_invalid'],
                         f"{coord_validation['summary']['invalid_percent']}%")
            with col3:
                st.metric("Non-Numeric", coord_validation['summary']['non_numeric_count'])
            with col4:
                st.metric("Out of Range", coord_validation['summary']['out_of_range_count'])
            
            st.write("**Issue Details:**")
            
            # Non-numeric values
            if coord_validation['issues']['non_numeric']:
                with st.expander("🔤 Non-Numeric Coordinates", expanded=False):
                    non_num_df = pd.DataFrame(coord_validation['issues']['non_numeric'][:10])
                    st.dataframe(non_num_df, use_container_width=True, hide_index=True)
                    if len(coord_validation['issues']['non_numeric']) > 10:
                        st.caption(f"... and {len(coord_validation['issues']['non_numeric']) - 10} more records")
            
            # Out of range values
            if coord_validation['issues']['out_of_range']:
                with st.expander("📍 Out of Geographic Range", expanded=False):
                    oor_df = pd.DataFrame(coord_validation['issues']['out_of_range'][:10])
                    st.dataframe(oor_df, use_container_width=True, hide_index=True)
                    if len(coord_validation['issues']['out_of_range']) > 10:
                        st.caption(f"... and {len(coord_validation['issues']['out_of_range']) - 10} more records")
            
            # Missing values
            if coord_validation['issues']['missing']:
                with st.expander(f"⚫ Missing Coordinates ({len(coord_validation['issues']['missing'])} records)", expanded=False):
                    st.caption("Records with empty latitude or longitude values (skipped in geographic analysis)")
            
            # Download excluded records button (integrated into report)
            if '_is_valid_coordinate' in df.columns and df['_is_valid_coordinate'].isin([False]).any():
                excluded_records = df[df['_is_valid_coordinate'] == False].copy()
                excluded_records = excluded_records[[
                    'CN', 'MEM_ID', 'bi1a', 'bi1b', 'bi1c', 'bi1c_1', 'hh_latitude', 'hh_longitude'
                ]].reset_index(drop=True)
                excluded_records.insert(0, 'Row#', range(1, len(excluded_records) + 1))
                
                from io import BytesIO
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    excluded_records.to_excel(writer, sheet_name='Excluded Points', index=False)
                    workbook = writer.book
                    worksheet = writer.sheets['Excluded Points']
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                output.seek(0)
                
                st.markdown("---")
                st.download_button(
                    label="📥 Download Excluded Records (XLSX)",
                    data=output.getvalue(),
                    file_name=f"excluded_coordinates_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_excluded_coords_integrated"
                )
                st.caption(f"Columns: Row#, CN, MEM_ID, State, District, GP/IV, Village/Block, Latitude, Longitude")
            
            st.warning("**Note:** These records are excluded from DBSCAN clustering and geographic analysis.")
        else:
            st.success(f"✅ All {coord_validation['valid_records']} records have valid geographic coordinates!")

    #FUNCTION: GENERATE CLUSTER IMAGES (Manual session state caching)
    # ════════════════════════════════════════════════════════════════════════
    # COMMENTED OUT - Image generation disabled
    # def generate_cluster_images(_df, _lat_col, _lon_col, file_checksum, _cluster_classifications=None):
    #     """Generate cluster images - uses manual session state caching"""
    #     import os
    #     import shutil
    #     from pathlib import Path
    #     
    #     # CLEAR all files in images directory at the start
    #     images_dir = Path("cluster_images")
    #     if images_dir.exists():
    #         try:
    #             # Close all matplotlib figures to release file handles
    #             plt.close('all')
    #             # Recursively delete all files and subdirectories
    #             shutil.rmtree(images_dir, ignore_errors=True)
    #         except (OSError, PermissionError) as e:
    #             pass  # If cleanup fails, continue anyway
    #     
    #     # Create fresh directory with subdirectories
    #     images_dir.mkdir(parents=True, exist_ok=True)
    #     regular_dir = images_dir / "regular_clusters"
    #     irregular_dir = images_dir / "irregular_clusters"
    #     regular_dir.mkdir(parents=True, exist_ok=True)
    #     irregular_dir.mkdir(parents=True, exist_ok=True)
    #     
    #     unique_clusters = sorted(_df['CN'].unique())
    #     
    #     for cluster_id in unique_clusters:
    #         try:
    #             cluster_data = _df[_df['CN'].astype(str) == cluster_id].copy()
    #             n_observations = len(cluster_data)
    #             
    #             # Count status 103 for classification
    #             status_103_count = (cluster_data['interview_status'] == 103).sum()
    #             
    #             # Determine if regular or irregular
    #             if status_103_count < 25 or status_103_count > 40:
    #                 classification = 'irregular'
    #                 save_dir = irregular_dir
    #             else:
    #                 classification = 'regular'
    #                 save_dir = regular_dir
    #             
    #             # Create figure with 2 subplots side by side
    #             fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))
    #             
    #             # ── Graph 1: Scatter Plot ────────────────────────────────
    #             cluster_data_copy = cluster_data.copy()
    #             cluster_data_copy['lat_r'] = cluster_data_copy[_lat_col].round(3)
    #             cluster_data_copy['lon_r'] = cluster_data_copy[_lon_col].round(3)
    #             
    #             counts = (
    #                 cluster_data_copy.groupby(['lat_r', 'lon_r'])
    #                 .size()
    #                 .reset_index(name='n')
    #             )
    #             
    #             ax1.scatter(counts['lon_r'], counts['lat_r'], s=counts['n']*20, alpha=0.6, color='steelblue')
    #             ax1.set_title(f'Cluster {cluster_id} - Location Count\n({n_observations} observations)', fontsize=12, fontweight='bold')
    #             ax1.set_xlabel('Longitude')
    #             ax1.set_ylabel('Latitude')
    #             ax1.grid(True, alpha=0.3)
    #             
    #             # ── Graph 2: Hexbin Density ──────────────────────────────
    #             if len(cluster_data) > 0:
    #                 hb = ax2.hexbin(
    #                     cluster_data[_lon_col],
    #                     cluster_data[_lat_col],
    #                     gridsize=35,
    #                     cmap='viridis',
    #                     mincnt=1
    #                 )
    #                 ax2.set_title(f'Cluster {cluster_id} - Density/Overlap\n({n_observations} observations)', fontsize=12, fontweight='bold')
    #                 ax2.set_xlabel('Longitude')
    #                 ax2.set_ylabel('Latitude')
    #                 cbar = plt.colorbar(hb, ax=ax2, label='Household count')
    #             
    #             # ── Add checksum text at the bottom ──────────────────────────
    #             fig.text(0.5, 0.02, f'MD5: {file_checksum} | Classification: {classification.upper()}', ha='center', fontsize=8, style='italic', color='gray')
    #             
    #             plt.tight_layout(rect=[0, 0.03, 1, 1])
    #             
    #             # Save image to appropriate subdirectory
    #             filename = save_dir / f"{cluster_id}.png"
    #             plt.savefig(filename, dpi=100, bbox_inches='tight')
    #             plt.close(fig)
    #             
    #         except Exception as e:
    #             st.warning(f"⚠️ Error generating image for cluster {cluster_id}: {e}")
    #             plt.close('all')
    #     
    #     return len(unique_clusters)
    
    # # Calculate file checksum once for caching key
    # file_checksum = calculate_md5_checksum(uploaded_file)
    # 
    # # Initialize session state for file checksum tracking
    # if 'file_checksum' not in st.session_state:
    #     st.session_state.file_checksum = None
    # 
    # # Initialize generation ID for background thread tracking
    # if 'generation_id' not in st.session_state:
    #     st.session_state.generation_id = 0
    # 
    # # Check if file has changed (different checksum = new file, even if same name)
    # file_changed = (st.session_state.file_checksum != file_checksum)
    # 
    # # CLEAR SESSION STATE when file changes (new file = fresh state)
    # if file_changed:
    #     # Keep only non-file-specific keys
    #     keys_to_keep = ['file_checksum', 'remarks_dict', 'generation_id']
    #     keys_to_delete = [k for k in st.session_state.keys() if k not in keys_to_keep]
    #     for key in keys_to_delete:
    #         del st.session_state[key]
    #     # Update checksum immediately
    #     st.session_state.file_checksum = file_checksum
    #     # Increment generation ID to terminate old background thread
    #     st.session_state.generation_id += 1
    # 
    # # Generate images for all clusters (only if file checksum is different)
    # if lat_col_name and lon_col_name and file_checksum and file_changed:
    #     # Capture current generation ID for this thread
    #     current_gen_id = st.session_state.generation_id
    #     
    #     # Run image generation in background thread (doesn't block Streamlit)
    #     def generate_images_background():
    #         """Generate images in background without blocking UI"""
    #         try:
    #             num_images = generate_cluster_images(df, lat_col_name, lon_col_name, file_checksum)
    #             # Only update session state if this is still the current generation
    #             try:
    #                 if st.session_state.generation_id == current_gen_id:
    #                     st.session_state['image_generation_complete'] = True
    #                     st.session_state['num_images_generated'] = num_images
    #             except (AttributeError, KeyError):
    #                 # generation_id may not exist in thread context, assume stale
    #                 pass
    #         except Exception as e:
    #             # Safe generation ID check for error handling
    #             try:
    #                 if st.session_state.generation_id == current_gen_id:
    #                     st.session_state['image_generation_error'] = str(e)
    #             except (AttributeError, KeyError):
    #                 # generation_id may not exist in thread context, silently fail
    #                 pass
    #     
    #     # Start background thread (daemon=True means it closes with app)
    #     bg_thread = threading.Thread(target=generate_images_background, daemon=True)
    #     bg_thread.start()
    #     st.info("🔄 Generating cluster density images in background...")
    # elif file_checksum and not file_changed:
    #     st.success(f"✅ Using cached cluster density images")
    # elif not file_checksum:
    #     st.warning("⚠️ Could not calculate file checksum for image generation")

    # ════════════════════════════════════════════════════════════════════════
    # MAIN LAYOUT  ── left panel (60%) | right panel (40%)
    # ════════════════════════════════════════════════════════════════════════
    left, right = st.columns([6, 4])

    # ── LEFT PANEL ───────────────────────────────────────────────────────────
    with left:

        # Cache dropdown options in session_state; invalidate when file changes
        _current_file_id = uploaded_file.file_id if hasattr(uploaded_file, 'file_id') else uploaded_file.name
        if st.session_state.get('_dropdown_file_id') != _current_file_id:
            st.session_state._dropdown_file_id = _current_file_id
            st.session_state.states_list = ["All"] + sorted(df['bi1a'].astype(str).unique().tolist())
            st.session_state.all_districts_list = ["All"] + sorted(df['bi1b'].astype(str).unique().tolist())
            st.session_state.cluster_nums_list = sorted(df['CN'].astype(str).unique().tolist())

        # Row 1: State / District dropdowns
        col_state, col_district = st.columns(2)
        with col_state:
            selected_state = st.selectbox("State", st.session_state.states_list)
        # Filter districts based on selected state
        if selected_state != "All":
            filtered_districts = ["All"] + sorted(df[df['bi1a'].astype(str) == selected_state]['bi1b'].astype(str).unique().tolist())
        else:
            filtered_districts = st.session_state.all_districts_list
        with col_district:
            selected_district = st.selectbox("District", filtered_districts)

        # Row 2: Cluster Number dropdown (filtered by state and district)
        # First, get base clusters based on state/district filters from df
        if selected_state != "All" and selected_district != "All":
            # Both state and district selected
            base_cluster_nums = sorted(df[(df['bi1a'].astype(str) == selected_state) & (df['bi1b'].astype(str) == selected_district)]['CN'].astype(str).unique().tolist())
        elif selected_state != "All":
            # Only state selected
            base_cluster_nums = sorted(df[df['bi1a'].astype(str) == selected_state]['CN'].astype(str).unique().tolist())
        elif selected_district != "All":
            # Only district selected
            base_cluster_nums = sorted(df[df['bi1b'].astype(str) == selected_district]['CN'].astype(str).unique().tolist())
        else:
            # All selected
            base_cluster_nums = st.session_state.cluster_nums_list.copy()
        
        # Add newly created CNs from break operations
        # Only add break CNs whose source CN is in the current state/district filter
        for bop in st.session_state.break_operations:
            if bop['source_cn'] in base_cluster_nums and bop['new_cn'] not in base_cluster_nums:
                base_cluster_nums.append(bop['new_cn'])
        base_cluster_nums = sorted(base_cluster_nums)
        
        # Filter out removed clusters and source clusters from merges
        removed_cns = set(st.session_state.removed_cns)
        merged_source_cns = set(str(source) for source, _ in st.session_state.merge_operations)
        
        # Collect source CNs auto-emptied by break/move (all rows reassigned)
        _auto_emptied_cns = set()
        for _step in st.session_state.steps_tracker:
            _ae = _step.get('_op_data', {}).get('auto_removed_source_cn')
            if _ae:
                _auto_emptied_cns.add(_ae)
        
        # Keep only clusters that are not removed, merged away, or auto-emptied
        cluster_nums = [cn for cn in base_cluster_nums if cn not in removed_cns and cn not in merged_source_cns and cn not in _auto_emptied_cns]
        
        selected_cluster = st.selectbox("CN", cluster_nums) if cluster_nums else None
        
        # Initialize session state for DBSCAN auto-apply tracking
        if 'last_dbscan_cluster' not in st.session_state:
            st.session_state.last_dbscan_cluster = None
        if 'last_dbscan_eps' not in st.session_state:
            st.session_state.last_dbscan_eps = None
        if 'last_dbscan_neighbors' not in st.session_state:
            st.session_state.last_dbscan_neighbors = None

        # ── DBSCAN controls ──────────────────────────────────────────────────
        # Initialize session state for default values (before widget creation)
        if 'eps_default_val' not in st.session_state:
            eps_val_default = "3.0"
            min_neighbors_default = "10"
            try:
                with open("setting.txt", "r") as f:
                    lines = f.readlines()
                    for line in lines:
                        if line.startswith("EPS="):
                            eps_val_default = line.split("=", 1)[1].strip()
                        elif line.startswith("MIN_NEIGHBORS="):
                            min_neighbors_default = line.split("=", 1)[1].strip()
            except Exception as e:
                pass
            st.session_state.eps_default_val = eps_val_default
            st.session_state.min_neighbors_default_val = min_neighbors_default
        
        # Initialize widget values using the widget keys
        if 'eps_input' not in st.session_state:
            st.session_state.eps_input = st.session_state.eps_default_val
        if 'min_neighbors_input' not in st.session_state:
            st.session_state.min_neighbors_input = st.session_state.min_neighbors_default_val
        
        # Define callback for Apply Default button (executes before widgets on next rerun)
        def apply_default_callback():
            st.session_state.eps_input = st.session_state.eps_default_val
            st.session_state.min_neighbors_input = st.session_state.min_neighbors_default_val
            st.session_state.last_dbscan_cluster = None
            st.session_state.last_dbscan_eps = None
            st.session_state.last_dbscan_neighbors = None
        
        with st.expander("DBSCAN Settings", expanded=True):
            #st.info("⚠️ **EPS is now in KILOMETERS (geographically accurate using Haversine metric)**")
            d1, d2 = st.columns(2)
            with d1:
                eps_val = st.text_input("EPS (km)", key="eps_input", help="Distance in kilometers (e.g., 2.0 for ~2km radius)")
            with d2:
                min_neighbors = st.text_input("MIN Neighbors", key="min_neighbors_input")
            
            # Button row
            btn_col1, btn_col2, btn_col3 = st.columns(3)
            with btn_col1:
                st.button("🔧 Apply Default", on_click=apply_default_callback, use_container_width=True, help="Apply default DBSCAN parameters", key="apply_default_btn")
            with btn_col2:
                apply_dbscan = st.button("⚙️ Apply DBSCAN", use_container_width=True, help="Apply custom DBSCAN parameters", key="apply_dbscan_btn")
        
        # Auto-apply DBSCAN when cluster changes or parameters change
        if selected_cluster is not None:
            cluster_changed = str(selected_cluster) != st.session_state.last_dbscan_cluster
            eps_changed = eps_val != st.session_state.last_dbscan_eps
            neighbors_changed = min_neighbors != st.session_state.last_dbscan_neighbors
            
            # Auto-run DBSCAN if cluster changed (priority) or if parameters changed
            if cluster_changed:
                apply_dbscan = True  # Force DBSCAN to run when cluster is selected
            else:
                apply_dbscan = apply_dbscan or eps_changed or neighbors_changed
            
            # ADDITIONAL: Force DBSCAN if no labels in session state (first run scenario)
            if not apply_dbscan and st.session_state.dbscan_labels is None:
                apply_dbscan = True
        else:
            apply_dbscan = False

        # ── Filter dataset ───────────────────────────────────────────────────
        filtered_df = df.copy()
        if selected_state    != "All":
            filtered_df = filtered_df[filtered_df['bi1a'].astype(str) == selected_state]
        if selected_district != "All":
            filtered_df = filtered_df[filtered_df['bi1b'].astype(str) == selected_district]

        # Restore persisted DBSCAN labels if DBSCAN didn't run this rerun
        if 'dbscan_cluster' not in filtered_df.columns and st.session_state.dbscan_labels is not None:
            filtered_df['dbscan_cluster'] = -1
            common_idx = filtered_df.index.intersection(st.session_state.dbscan_labels.index)
            filtered_df.loc[common_idx, 'dbscan_cluster'] = st.session_state.dbscan_labels.loc[common_idx]

        # ── Handle Missing Values ─────────────────────────────────────────────
        # NOTE: Missing rows are NOT dropped from filtered_df or df_processed.
        # DBSCAN and detect_split() handle missing coords internally via .dropna()
        # on lat/lon columns only. All records (including those with missing values)
        # are preserved in the dataset and complete data displays.

        # ── DBSCAN clustering PER CN (HAVERSINE METRIC - GEOGRAPHICALLY ACCURATE) ───
        # Label encoding:
        #   >= 0 : Actual DBSCAN cluster ID
        #   -1   : Noise (valid coordinates but not assigned to any cluster)
        #   -2   : Excluded (invalid or missing geocodes - not included in DBSCAN)
        #dbscan_label = "DBSCAN - SUBGROUPS = N/A"
        if apply_dbscan:
            try:
                from sklearn.cluster import DBSCAN
                
                # Initialize dbscan_cluster column for entire filtered_df
                filtered_df['dbscan_cluster'] = -1
                
                if 'hh_latitude' in filtered_df.columns and 'hh_longitude' in filtered_df.columns:
                    # ── Validate and convert parameters ──────────────────────────────
                    try:
                        eps_val_clean = eps_val.strip() if isinstance(eps_val, str) else str(eps_val)
                        if not eps_val_clean:
                            raise ValueError("EPS value is empty. Using default value of 3.0")
                        eps_km = float(eps_val_clean)
                        if eps_km <= 0:
                            raise ValueError(f"EPS must be positive, got {eps_km}")
                    except ValueError as ve:
                        st.error(f"Invalid EPS value: {ve}")
                        st.stop()
                    
                    try:
                        min_neighbors_clean = min_neighbors.strip() if isinstance(min_neighbors, str) else str(min_neighbors)
                        if not min_neighbors_clean:
                            raise ValueError("MIN Neighbors value is empty. Using default value of 10")
                        min_samples = int(min_neighbors_clean)
                        if min_samples < 1:
                            raise ValueError(f"MIN Neighbors must be at least 1, got {min_samples}")
                    except ValueError as ve:
                        st.error(f"Invalid MIN Neighbors value: {ve}")
                        st.stop()
                    
                    # Convert EPS from km to radians (Earth radius = 6371 km)
                    eps_radians = eps_km / 6371.0
                    
                    # Validate eps_radians
                    if not np.isfinite(eps_radians) or eps_radians <= 0:
                        st.error(f"Invalid EPS calculation: calculated eps_radians = {eps_radians}")
                        st.stop()
                    
                    # Get unique CNs in filtered data
                    unique_cns = filtered_df['CN'].unique()
                    total_subgroups = 0
                    cn_subgroup_details = []
                    
                    # Apply DBSCAN per CN
                    for cn in unique_cns:
                        cn_data = filtered_df[filtered_df['CN'] == cn]
                        
                        # Filter to only valid coordinates (if validation was done)
                        if '_is_valid_coordinate' in cn_data.columns:
                            cn_data_valid = cn_data[cn_data['_is_valid_coordinate'] == True]
                        else:
                            cn_data_valid = cn_data
                        
                        coords = cn_data_valid[['hh_latitude', 'hh_longitude']].dropna().to_numpy(dtype=np.float64)
                        valid_idx = cn_data_valid[['hh_latitude', 'hh_longitude']].dropna().index
                        
                        if len(coords) >= 2:
                            # Validate coordinates
                            if not np.all(np.isfinite(coords)):
                                st.warning(f"CN {cn}: Skipping due to invalid coordinates (NaN/Inf values)")
                                continue
                            
                            # Convert to radians for haversine metric
                            coords_radians = np.radians(coords)
                            
                            # Use haversine metric with ball_tree algorithm
                            db = DBSCAN(eps=eps_radians, min_samples=min_samples, 
                                       algorithm='ball_tree', metric='haversine')
                            labels = db.fit_predict(coords_radians)
                            
                            # Assign labels back to filtered_df for this CN
                            filtered_df.loc[valid_idx, 'dbscan_cluster'] = labels
                            
                            # Count sub-clusters (exclude noise -1)
                            n_subgroups = len(set(labels)) - (1 if -1 in labels else 0)
                            total_subgroups += n_subgroups
                            cn_subgroup_details.append({
                                'CN': cn,
                                'n_subgroups': n_subgroups,
                                'total_points': len(coords),
                                'noise_points': (labels == -1).sum()
                            })
                    
                    # Persist DBSCAN labels in session_state for survival across reruns
                    st.session_state.dbscan_labels = filtered_df['dbscan_cluster'].copy()
                    
                    # Mark records with invalid coordinates as EXCLUDED (-2)
                    # This distinguishes them from actual noise points (-1)
                    if '_is_valid_coordinate' in filtered_df.columns:
                        invalid_mask = filtered_df['_is_valid_coordinate'] == False
                        filtered_df.loc[invalid_mask, 'dbscan_cluster'] = -2
                        # Update session state with the corrected labels
                        st.session_state.dbscan_labels = filtered_df['dbscan_cluster'].copy()
                    
                    # Update tracking to prevent re-applying
                    st.session_state.last_dbscan_cluster = str(selected_cluster)
                    st.session_state.last_dbscan_eps = eps_val
                    st.session_state.last_dbscan_neighbors = min_neighbors
                    
                    # Show status message with per-CN details
                    status_msg = "✓ DBSCAN Auto-Applied" if cluster_changed else "✓ DBSCAN Applied"
                    st.success(f"{status_msg}: Total {total_subgroups} sub-clusters detected across {len(unique_cns)} CN(s) (EPS={eps_km}km, MIN_NEIGHBORS={min_samples})")
                    
                    # Show per-CN details
                    # if cn_subgroup_details:
                    #     with st.expander("📊 DBSCAN Details by CN", expanded=False):
                    #         details_df = pd.DataFrame(cn_subgroup_details)
                    #         st.dataframe(details_df, use_container_width=True, hide_index=True)
                else:
                    st.error("Latitude/Longitude columns not found in dataset")
            except ImportError:
                st.error("DBSCAN - scikit-learn not installed")
            except Exception as e:
                st.error(f"DBSCAN - Error: {str(e)}")
        
        # ── DOWNLOAD EXCLUDED RECORDS (Invalid Coordinates) ──────────────────
        # NOTE: Download button has been moved to the validation report section above

        #st.caption(dbscan_label)

        # ── DBSCAN SPLIT DETECTION (HAVERSINE METRIC) ────────────────────────
        from sklearn.cluster import DBSCAN
        
        # Use user-provided DBSCAN parameters for split detection
        try:
            eps_val_clean = eps_val.strip() if isinstance(eps_val, str) else str(eps_val)
            dbscan_eps_km = float(eps_val_clean) if eps_val_clean else float(st.session_state.eps_default_val)
        except (ValueError, AttributeError):
            dbscan_eps_km = float(st.session_state.eps_default_val)
        
        try:
            min_neighbors_clean = min_neighbors.strip() if isinstance(min_neighbors, str) else str(min_neighbors)
            dbscan_min_samples = int(min_neighbors_clean) if min_neighbors_clean else int(st.session_state.min_neighbors_default_val)
        except (ValueError, AttributeError):
            dbscan_min_samples = int(st.session_state.min_neighbors_default_val)
        
        dbscan_eps_radians = dbscan_eps_km / 6371.0  # Convert km to radians
        
        def detect_split(group):
            """Count number of DBSCAN sub-clusters (excluding noise points labeled -1)"""
            try:
                # Filter to only valid coordinates if validation marker exists
                if '_is_valid_coordinate' in group.columns:
                    group_valid = group[group['_is_valid_coordinate'] == True]
                else:
                    group_valid = group
                
                coords = group_valid[['hh_latitude','hh_longitude']].dropna().to_numpy(dtype=np.float64)
                if len(coords) < 2:
                    return 0
                # Validate coordinates
                if not np.all(np.isfinite(coords)):
                    return 0
                coords_radians = np.radians(coords)
                db = DBSCAN(eps=dbscan_eps_radians, min_samples=dbscan_min_samples, 
                           algorithm='ball_tree', metric='haversine').fit(coords_radians)
                # Count clusters: all unique labels minus noise label (-1)
                unique_labels = set(db.labels_)
                if -1 in unique_labels:
                    # Subtract 1 to exclude noise from cluster count
                    return max(0, len(unique_labels) - 1)
                else:
                    # All points are assigned to clusters (no noise)
                    return len(unique_labels)
            except Exception:
                return 0
        
        # ── CLUSTER DATASET (will be calculated after df_processed is built) ───
        # Placeholder - actual cluster summary will be displayed after district removal
        # and other operations are applied to df_processed

        # Get form input values (default to 0 if not set)
        form_rural = st.session_state.get('rural_clusters', 0)
        form_urban = st.session_state.get('urban_clusters', 0)
        
        form_non_mp = st.session_state.get('non_mp_clusters', 0)
        form_non_tribal = st.session_state.get('non_tribal_clusters', 0)
        
        # ── ALLOCATIONS: CALCULATED AFTER DISTRICT REMOVAL ──────────────────────
        # These tables will be displayed after district removal filter below

        # ── INITIALIZE MP SELECTION IN SESSION STATE ─────────────────────────
        if 'mp_bi1c_values' not in st.session_state:
            st.session_state.mp_bi1c_values = []
        if 'mp_clusters_selected' not in st.session_state:
            st.session_state.mp_clusters_selected = []

        # ── SELECT MILLION PLUS CLUSTERS BY BI1C VALUES ──────────────────────
        st.markdown("**Select Million Plus Clusters from the Data**")
        st.caption("Choose which bi1c values indicate MP clusters (all clusters with those values will be tagged as MP)")
        
        # Get unique bi1c values that contain both Urban and Mc
        all_bi1c = df['bi1c'].astype(str).unique()
        unique_bi1c = sorted([val for val in all_bi1c if 'Urban' in val and 'Mc' in val])
        
        with st.form("select_mp_form"):
            mp_bi1c_values = st.multiselect(
                "MP bi1c Values",
                options=unique_bi1c,
                default=st.session_state.mp_bi1c_values,
                help="Select the exact bi1c values that indicate MP clusters",
                key="mp_bi1c_multiselect"
            )
            mp_submit = st.form_submit_button("✅ Confirm MP Values")
        
        if mp_submit:
            st.session_state.mp_bi1c_values = mp_bi1c_values
            # Find all clusters with selected bi1c values
            mp_clusters_list = sorted(df[df['bi1c'].astype(str).isin(mp_bi1c_values)]['CN'].astype(str).unique().tolist())
            st.session_state.mp_clusters_selected = mp_clusters_list
            
            # Track MP selection in action tracker
            mp_detail = f"MP bi1c values: {', '.join(mp_bi1c_values)}" if mp_bi1c_values else "Cleared MP selection"
            # Replace existing MP selection step or add new
            existing_mp_idx = next(
                (i for i, s in enumerate(st.session_state.steps_tracker) if s.get('_op_type') == 'select_mp'),
                None
            )
            mp_step = {
                'S.No': (st.session_state.steps_tracker[existing_mp_idx]['S.No']
                         if existing_mp_idx is not None
                         else len(st.session_state.steps_tracker) + 1),
                'Step': 'Select MP Clusters',
                'Details': mp_detail,
                'Remark': f"{len(mp_clusters_list)} cluster(s) marked",
                '_op_type': 'select_mp',
                '_op_data': {
                    'mp_bi1c_values': list(mp_bi1c_values),
                    'mp_clusters_selected': mp_clusters_list
                }
            }
            if existing_mp_idx is not None:
                st.session_state.steps_tracker[existing_mp_idx] = mp_step
            else:
                st.session_state.steps_tracker.append(mp_step)
            
            st.success(f"✅ {len(mp_clusters_list)} cluster(s) marked as MP")

        # ── REMOVE DISTRICTS: Keep only MP from selected districts ──────────────
        st.markdown("**Remove Districts**")
        st.caption("Select districts to exclude: only MP clusters will be retained from those districts")
        
        available_districts = sorted(df['bi1b'].unique().astype(str))
        
        with st.form("remove_districts_form"):
            districts_to_remove = st.multiselect(
                "Select districts to remove (non-MP clusters)",
                options=available_districts,
                default=st.session_state._saved_districts_to_remove,
                key="districts_to_remove"
            )
            remove_submit = st.form_submit_button("✅ Apply District Removal")
        
        # Build processed dataframe based on district removal
        df_processed = df.copy()
        
        # Copy DBSCAN cluster assignments from filtered_df to df_processed if available
        if 'dbscan_cluster' in filtered_df.columns:
            # Match rows by index and copy the column
            df_processed = df_processed.assign(dbscan_cluster=-1)  # Initialize all as noise
            # Find common indices and copy values
            common_indices = df_processed.index.intersection(filtered_df.index)
            if len(common_indices) > 0:
                df_processed.loc[common_indices, 'dbscan_cluster'] = filtered_df.loc[common_indices, 'dbscan_cluster']
        
        # FALLBACK: Ensure dbscan_cluster exists (from session state) even if not in filtered_df
        if 'dbscan_cluster' not in df_processed.columns and st.session_state.dbscan_labels is not None:
            df_processed = df_processed.assign(dbscan_cluster=-1)  # Initialize all as noise
            common_indices = df_processed.index.intersection(st.session_state.dbscan_labels.index)
            if len(common_indices) > 0:
                df_processed.loc[common_indices, 'dbscan_cluster'] = st.session_state.dbscan_labels.loc[common_indices]
        
        # Track step only on form submit
        if remove_submit:
            if districts_to_remove:
                st.session_state._saved_districts_to_remove = sorted(districts_to_remove)
                # Only add to tracker if this exact set hasn't been tracked yet
                removal_detail = f"Removed non-MP from: {', '.join(sorted(districts_to_remove))}"
                already_tracked = any(
                    s.get('Step') == 'Remove Districts' and s.get('Details') == removal_detail 
                    for s in st.session_state.steps_tracker
                )
                if not already_tracked:
                    # Remove any older remove_districts entry (we replace it)
                    st.session_state.steps_tracker = [
                        s for s in st.session_state.steps_tracker if s.get('_op_type') != 'remove_districts'
                    ]
                    st.session_state.steps_tracker.append({
                        'S.No': len(st.session_state.steps_tracker) + 1,
                        'Step': 'Remove Districts',
                        'Details': removal_detail,
                        'Remark': '-',
                        '_op_type': 'remove_districts',
                        '_op_data': {'districts': sorted(districts_to_remove)}
                    })
            else:
                # Empty submission = clear the filter
                st.session_state._saved_districts_to_remove = []
                st.session_state.steps_tracker = [
                    s for s in st.session_state.steps_tracker if s.get('_op_type') != 'remove_districts'
                ]
                # Re-number S.No
                for _i, _s in enumerate(st.session_state.steps_tracker):
                    _s['S.No'] = _i + 1
            st.rerun()  # Rerun to apply immediately
        
        # ALWAYS apply district removal filter from saved state (single source of truth)
        if st.session_state._saved_districts_to_remove:
            _eff_dists = st.session_state._saved_districts_to_remove
            is_mp_cluster = df_processed['CN'].astype(str).isin(st.session_state.mp_clusters_selected)
            is_remove_dist = df_processed['bi1b'].astype(str).isin([str(d) for d in _eff_dists])
            
            records_before = len(df_processed)
            
            # Keep rows where: (district to remove AND MP cluster) OR (district NOT to remove)
            df_processed = df_processed[
                (is_remove_dist & is_mp_cluster) | ~is_remove_dist
            ]
            
            records_after = len(df_processed)
            st.success(f"✅ District filter active: {len(_eff_dists)} district(s) filtered — {records_after} records ({records_before - records_after} removed)")
        else:
            st.info("No districts selected for removal — using full dataset")

        # ── APPLY STORED CN REMOVALS ─────────────────────────────────────────
        if st.session_state.removed_cns:
            df_processed = df_processed[~df_processed['CN'].astype(str).isin(st.session_state.removed_cns)]

        # ── APPLY STORED MERGE OPERATIONS ────────────────────────────────────
        # Add original_CN and original_concat columns if not already present
        # original_concat captures bi1a_bi1b_bi1c_bi1c_1 BEFORE any merge overwrites
        if 'original_CN' not in df_processed.columns:
            df_processed['original_CN'] = df_processed['CN']
        if 'original_concat' not in df_processed.columns:
            df_processed['original_concat'] = df_processed['concat']
        
        for source_cn, target_cn in st.session_state.merge_operations:
            mask = df_processed['CN'].astype(str) == str(source_cn)
            if mask.any():
                # Get target cluster's bi1a/bi1b/bi1c/bi1c_1 values from first matching row
                target_mask = df_processed['CN'].astype(str) == str(target_cn)
                if target_mask.any():
                    target_row = df_processed.loc[target_mask].iloc[0]
                    # Transfer geographic/classification fields from target to source rows
                    for field in ['bi1a', 'bi1b', 'bi1c', 'bi1c_1']:
                        if field in df_processed.columns:
                            df_processed.loc[mask, field] = target_row[field]
                    # Update concat to match new bi1a_bi1b_bi1c_bi1c_1
                    df_processed.loc[mask, 'concat'] = (
                        str(target_row['bi1a']) + '_' +
                        str(target_row['bi1b']) + '_' +
                        str(target_row['bi1c']) + '_' +
                        str(target_row['bi1c_1'])
                    )
                
                # ── REASSIGN HH_ID AND MEM_ID IN CONTINUING SEQUENCE ──────────
                # Reconstruct HH_CODE for source rows to identify unique households
                src_rows_idx = df_processed[mask].index
                src_rows_copy = df_processed.loc[src_rows_idx].copy()
                src_rows_copy['HH_CODE'] = (
                    src_rows_copy['member_number']
                    .astype(str)
                    .str.split('.')
                    .str[:-1]
                    .str.join('.')
                )
                
                # Get unique HH_CODEs in source (preserves order of first appearance)
                unique_hh_codes_src = src_rows_copy['HH_CODE'].drop_duplicates().values
                
                # Find max HH_NUM in target cluster (extract HHH from HH_ID)
                tgt_rows = df_processed[df_processed['CN'].astype(str) == str(target_cn)]
                if len(tgt_rows) > 0:
                    tgt_hh_nums = tgt_rows['HH_ID'].astype(str).str.split('.').str[-1]
                    max_hh_num = int(tgt_hh_nums.astype(int).max()) if len(tgt_hh_nums) > 0 else 0
                else:
                    max_hh_num = 0
                
                # Create mapping: HH_CODE -> new HH_NUM (continuing sequence)
                hh_code_to_new_num = {}
                for i, hh_code in enumerate(unique_hh_codes_src, 1):
                    hh_code_to_new_num[hh_code] = str(max_hh_num + i).zfill(3)
                
                # Apply new HH_ID and MEM_ID to all source rows
                tgt_cn_str = str(target_cn)
                for idx in src_rows_idx:
                    hh_code = '.'.join(str(df_processed.loc[idx, 'member_number']).split('.')[:-1])
                    new_hh_num = hh_code_to_new_num.get(hh_code)
                    if new_hh_num:
                        new_hh_id = tgt_cn_str + '.' + new_hh_num
                        member_part = str(df_processed.loc[idx, 'member_number']).split('.')[-1]
                        df_processed.loc[idx, 'HH_ID'] = new_hh_id
                        df_processed.loc[idx, 'MEM_ID'] = new_hh_id + '.' + member_part
                
                # Append merge chain: A→B, then if B→C, A becomes A→B→C
                # Also update original B rows that haven't been part of a prior merge
                # (their original_CN still equals their CN, i.e. 'B')
                target_native_mask = (
                    mask &
                    (df_processed['original_CN'].astype(str) == str(source_cn))
                )
                already_chained_mask = mask & ~target_native_mask
                # Rows already carrying a chain (e.g. A→B) — extend it
                df_processed.loc[already_chained_mask, 'original_CN'] = (
                    df_processed.loc[already_chained_mask, 'original_CN'].astype(str) + '→' + str(target_cn)
                )
                # Native rows of source_cn — start a new chain
                df_processed.loc[target_native_mask, 'original_CN'] = (
                    str(source_cn) + '→' + str(target_cn)
                )
                df_processed.loc[mask, 'CN'] = str(target_cn)

        # ── APPLY STORED BREAK OPERATIONS ────────────────────────────────────
        for break_op in st.session_state.break_operations:
            src_cn = break_op['source_cn']
            # Handle new format: dbscan_ids (list) instead of dbscan_cluster_id (single)
            if 'dbscan_ids' in break_op:
                dbscan_ids = break_op['dbscan_ids']  # Multiple DBSCAN IDs → ONE CN
            else:
                dbscan_ids = [break_op['dbscan_cluster_id']]  # Old format: single ID

            new_cn = break_op['new_cn']

            if 'dbscan_cluster' not in df_processed.columns:
                continue

            # Find rows matching ANY of the selected DBSCAN cluster IDs
            break_mask = (
                (df_processed['CN'].astype(str) == str(src_cn)) &
                (df_processed['dbscan_cluster'].isin(dbscan_ids))
            )
            if not break_mask.any():
                continue

            # Track original CN (same chain style as merge: original→new)
            df_processed.loc[break_mask, 'original_CN'] = df_processed.loc[break_mask, 'original_CN'].astype(str) + '→' + new_cn

            # Override bi1a/bi1b/bi1c/bi1c_1 if user provided values
            for field in ['bi1a', 'bi1b', 'bi1c', 'bi1c_1']:
                val = break_op.get(field)
                if val and field in df_processed.columns:
                    df_processed.loc[break_mask, field] = val

            # Override lat/lon if user provided custom coordinates
            if break_op.get('lat') and 'hh_latitude' in df_processed.columns:
                try:
                    df_processed.loc[break_mask, 'hh_latitude'] = float(break_op['lat'])
                except ValueError:
                    pass
            if break_op.get('lon') and 'hh_longitude' in df_processed.columns:
                try:
                    df_processed.loc[break_mask, 'hh_longitude'] = float(break_op['lon'])
                except ValueError:
                    pass

            # Update concat field to reflect new classification
            if all(f in df_processed.columns for f in ['bi1a', 'bi1b', 'bi1c', 'bi1c_1']):
                df_processed.loc[break_mask, 'concat'] = (
                    df_processed.loc[break_mask, 'bi1a'].astype(str) + '_' +
                    df_processed.loc[break_mask, 'bi1b'].astype(str) + '_' +
                    df_processed.loc[break_mask, 'bi1c'].astype(str) + '_' +
                    df_processed.loc[break_mask, 'bi1c_1'].astype(str)
                )

            # Assign new CN
            df_processed.loc[break_mask, 'CN'] = new_cn

            # Reassign HH_ID and MEM_ID for the new CN (start from 001)
            new_cn_rows = df_processed[df_processed['CN'] == new_cn].copy()
            new_cn_rows['_HH_CODE'] = (
                new_cn_rows['member_number']
                .astype(str)
                .str.split('.')
                .str[:-1]
                .str.join('.')
            )
            unique_hh_codes = new_cn_rows['_HH_CODE'].drop_duplicates().values
            hh_code_map = {code: str(i).zfill(3) for i, code in enumerate(unique_hh_codes, 1)}

            for idx in new_cn_rows.index:
                hh_code = '.'.join(str(df_processed.loc[idx, 'member_number']).split('.')[:-1])
                new_hh_num = hh_code_map.get(hh_code, '001')
                new_hh_id = new_cn + '.' + new_hh_num
                member_part = str(df_processed.loc[idx, 'member_number']).split('.')[-1]
                df_processed.loc[idx, 'HH_ID'] = new_hh_id
                df_processed.loc[idx, 'MEM_ID'] = new_hh_id + '.' + member_part

        # ── APPLY STORED MOVE OPERATIONS ─────────────────────────────────────
        for move_op in st.session_state.move_operations:
            move_indices = move_op['indices']       # list of 'index' col values
            move_target_cn = str(move_op['target_cn'])

            # Build mask from stable index column
            move_mask = df_processed['index'].isin(move_indices)
            if not move_mask.any():
                continue

            # Copy target cluster's classification from first existing row
            tgt_mask_m = df_processed['CN'].astype(str) == move_target_cn
            if tgt_mask_m.any():
                tgt_row_m = df_processed.loc[tgt_mask_m].iloc[0]
                for fld in ['bi1a', 'bi1b', 'bi1c', 'bi1c_1']:
                    if fld in df_processed.columns:
                        df_processed.loc[move_mask, fld] = tgt_row_m[fld]
                df_processed.loc[move_mask, 'concat'] = (
                    str(tgt_row_m['bi1a']) + '_' +
                    str(tgt_row_m['bi1b']) + '_' +
                    str(tgt_row_m['bi1c']) + '_' +
                    str(tgt_row_m['bi1c_1'])
                )

            # ── Reassign HH_ID / MEM_ID avoiding collision ──────────────
            src_rows_m = df_processed.loc[move_mask].copy()
            src_rows_m['_HH_CODE'] = (
                src_rows_m['member_number']
                .astype(str).str.split('.').str[:-1].str.join('.')
            )
            unique_hh_m = src_rows_m['_HH_CODE'].drop_duplicates().values

            tgt_rows_m = df_processed[df_processed['CN'].astype(str) == move_target_cn]
            if len(tgt_rows_m) > 0:
                tgt_hh_nums_m = tgt_rows_m['HH_ID'].astype(str).str.split('.').str[-1]
                max_hh_m = int(tgt_hh_nums_m.astype(int).max())
            else:
                max_hh_m = 0

            hh_map_m = {}
            for i, hc in enumerate(unique_hh_m, 1):
                hh_map_m[hc] = str(max_hh_m + i).zfill(3)

            for idx in df_processed.loc[move_mask].index:
                hc = '.'.join(str(df_processed.loc[idx, 'member_number']).split('.')[:-1])
                new_hh_num = hh_map_m.get(hc)
                if new_hh_num:
                    new_hh_id = move_target_cn + '.' + new_hh_num
                    mem_part = str(df_processed.loc[idx, 'member_number']).split('.')[-1]
                    df_processed.loc[idx, 'HH_ID'] = new_hh_id
                    df_processed.loc[idx, 'MEM_ID'] = new_hh_id + '.' + mem_part

            # Track original_CN chain
            source_cn_m = move_op['source_cn']
            native_m = move_mask & (df_processed['original_CN'].astype(str) == str(source_cn_m))
            chained_m = move_mask & ~native_m
            df_processed.loc[chained_m, 'original_CN'] = (
                df_processed.loc[chained_m, 'original_CN'].astype(str) + '→' + move_target_cn
            )
            df_processed.loc[native_m, 'original_CN'] = str(source_cn_m) + '→' + move_target_cn

            # Assign new CN
            df_processed.loc[move_mask, 'CN'] = move_target_cn

        # ── APPLY RECONSIDER (Sort by latitude & assign select_id) ───────────
        if st.session_state.reconsidered_cns:
            for rcn in st.session_state.reconsidered_cns:
                rcn_mask = df_processed['CN'].astype(str) == str(rcn)
                if rcn_mask.any() and lat_col_name:
                    rcn_indices = df_processed.loc[rcn_mask].sort_values(lat_col_name).index
                    df_processed.loc[rcn_indices, 'select_id'] = range(len(rcn_indices))

        # ── CLUSTER SUMMARY (from df_processed - CURRENT STATE) ───────────────
        st.subheader("Clusters Summary")
        
        # Build cluster_summary from df_processed
        cluster_summary_current = (
            df_processed.groupby('CN')
            .agg(
                overall_count=('MEM_ID', 'count'),
                status_103_count=('interview_status', lambda x: (x == 103).sum())
            )
            .reset_index()
        )
        
        # ── Add adolescent count ──
        adolescent_by_cluster_current = df_processed[
            (df_processed['c2_age'] > 12) & 
            (df_processed['c2_age'] < 18) & 
            (df_processed['interview_status'] == 103)
        ].groupby('CN').size().reset_index(name='adolescent_count')
        cluster_summary_current = cluster_summary_current.merge(adolescent_by_cluster_current, on='CN', how='left')
        cluster_summary_current['adolescent_count'] = cluster_summary_current['adolescent_count'].fillna(0).astype(int)
        
        # ── Add CPC count ──
        cpc_by_cluster_current = df_processed[
            df_processed['c2_age'] < 13
        ].groupby('CN').size().reset_index(name='cpc_count')
        cluster_summary_current = cluster_summary_current.merge(cpc_by_cluster_current, on='CN', how='left')
        cluster_summary_current['cpc_count'] = cluster_summary_current['cpc_count'].fillna(0).astype(int)
        
        # ── Add missing geocodes count (lat/lon only) ──
        _geo_cols = [c for c in ['hh_latitude', 'hh_longitude'] if c in df_processed.columns]
        if _geo_cols:
            missing_by_cluster_current = df_processed.groupby('CN').apply(
                lambda group: group[_geo_cols].isnull().any(axis=1).sum()
            ).reset_index()
        else:
            missing_by_cluster_current = df_processed.groupby('CN').apply(lambda _: 0).reset_index()
        missing_by_cluster_current.columns = ['CN', 'missing_geocodes']
        cluster_summary_current = cluster_summary_current.merge(missing_by_cluster_current, on='CN', how='left')
        cluster_summary_current['missing_geocodes'] = cluster_summary_current['missing_geocodes'].fillna(0).astype(int)
        
        # ── Add n_splits column (count DBSCAN sub-clusters within each CN) ──
        # Use pre-computed dbscan_cluster assignments (from filtered_df) for consistency with charts
        if 'dbscan_cluster' in df_processed.columns:
            def count_clusters_from_dbscan(group):
                """Count unique DBSCAN clusters (excluding noise -1 and invalid -2) from pre-computed column"""
                unique_clusters = set(group['dbscan_cluster'].dropna().unique())
                # Remove noise (-1) and invalid/excluded geocodes (-2) from count
                unique_clusters.discard(-1)
                unique_clusters.discard(-2)
                return len(unique_clusters)
            
            try:
                split_check_current = df_processed.groupby('CN').apply(count_clusters_from_dbscan).reset_index(name='n_splits')
            except Exception:
                split_check_current = pd.DataFrame(columns=['CN', 'n_splits'])
        else:
            # Fallback: recalculate DBSCAN if pre-computed column missing
            try:
                split_check_current = df_processed.groupby('CN').apply(detect_split).reset_index()
                split_check_current.columns = ['CN', 'n_splits']
            except Exception:
                split_check_current = pd.DataFrame(columns=['CN', 'n_splits'])
        
        cluster_summary_current = cluster_summary_current.merge(split_check_current, on='CN', how='left')
        cluster_summary_current['n_splits'] = cluster_summary_current['n_splits'].fillna(0).astype(int)
        
        # ── Add Noise column (count DBSCAN noise points: label == -1) ──
        if 'dbscan_cluster' in df_processed.columns:
            def count_noise_points(group):
                """Count points labeled as noise (-1) in DBSCAN results"""
                return (group['dbscan_cluster'] == -1).sum()
            
            noise_check_current = df_processed.groupby('CN').apply(count_noise_points).reset_index(name='Noise')
        else:
            # Fallback if dbscan_cluster column doesn't exist
            noise_check_current = pd.DataFrame({'CN': df_processed['CN'].unique(), 'Noise': 0})
        
        cluster_summary_current = cluster_summary_current.merge(noise_check_current, on='CN', how='left')
        cluster_summary_current['Noise'] = cluster_summary_current['Noise'].fillna(0).astype(int)
        
        # ── Add TYPE, District, bi1c, bi1c_1 columns (vectorized) ─────────
        # Single groupby().first() replaces 4× per-CN DataFrame scans
        _cn_meta = df_processed.groupby('CN')[['bi1b', 'bi1c', 'bi1c_1']].first().reset_index()
        cluster_summary_current = cluster_summary_current.merge(_cn_meta, on='CN', how='left')
        cluster_summary_current['bi1c'] = cluster_summary_current['bi1c'].fillna('Unknown')
        cluster_summary_current['bi1c_1'] = cluster_summary_current['bi1c_1'].fillna('Unknown')
        cluster_summary_current['District'] = cluster_summary_current['bi1b'].fillna('Unknown')
        # Derive TYPE from bi1c using selected MP values
        def _classify_type(v):
            s = str(v)
            if 'Urban' in s:
                return 'MP' if s in st.session_state.mp_bi1c_values else 'NMP'
            return 'TRI' if 'TRI-' in s else 'Non-TRI'
        cluster_summary_current['TYPE'] = cluster_summary_current['bi1c'].apply(_classify_type)
        
        # ── Classify as Regular or Irregular ──
        def classify_cluster(count):
            return 'Irregular' if (count < 25 or count > 40) else 'Regular'
        
        cluster_summary_current['classification'] = cluster_summary_current['status_103_count'].apply(classify_cluster)
        
        # ── Split Regular and Irregular ──
        regular_clusters_current = cluster_summary_current[cluster_summary_current['classification'] == 'Regular'].drop(columns=['classification', 'bi1b'])
        irregular_clusters_current = cluster_summary_current[cluster_summary_current['classification'] == 'Irregular'].drop(columns=['classification', 'bi1b'])
        
        # Display Regular Clusters
        st.subheader("😊👌 Regular Clusters")
        st.caption("Clusters with Status_103 count between 25 and 40 (inclusive)")
        if not regular_clusters_current.empty:
            regular_display = filter_dataframe(regular_clusters_current, key="regular_clusters_filter")
            st.dataframe(regular_display, use_container_width=True, height=400)
        else:
            st.info("No regular clusters found.")
        
        # Display Irregular Clusters
        st.subheader("😒👎 Irregular Clusters")
        st.caption("Clusters with Status_103 count less than 25 or greater than 40")
        if not irregular_clusters_current.empty:
            irregular_display = filter_dataframe(irregular_clusters_current, key="irregular_clusters_filter")
            st.dataframe(irregular_display, use_container_width=True, height=400)
        else:
            st.info("No irregular clusters found.")

        # ── DISPLAY ALLOCATIONS FROM df_processed (ALWAYS CURRENT) ──────────────
        # Calculate from df_processed (whether filtered or full)
        is_urban_current = df_processed['bi1c'].astype(str).str.contains("Urban", case=False, na=False)
        is_tribal_current = df_processed['bi1c'].astype(str).str.contains("TRI-", case=False, na=False)
        is_mp_current = df_processed['bi1c'].astype(str).isin(st.session_state.mp_bi1c_values)
        
        rural_count_current = (~is_urban_current).sum()
        urban_count_current = is_urban_current.sum()
        tribal_count_current = (~is_urban_current & is_tribal_current).sum()
        non_tribal_count_current = (~is_urban_current & ~is_tribal_current).sum()
        mp_count_current = (is_urban_current & is_mp_current).sum()
        nmp_count_current = (is_urban_current & ~is_mp_current).sum()
        
        # Count unique clusters per category
        rural_clusters_count_current = len(df_processed[~is_urban_current]['CN'].unique())
        urban_clusters_count_current = len(df_processed[is_urban_current]['CN'].unique())
        tribal_clusters_count_current = len(df_processed[~is_urban_current & is_tribal_current]['CN'].unique())
        non_tribal_clusters_count_current = len(df_processed[~is_urban_current & ~is_tribal_current]['CN'].unique())
        mp_clusters_count_current = len(df_processed[is_urban_current & is_mp_current]['CN'].unique())
        nmp_clusters_count_current = len(df_processed[is_urban_current & ~is_mp_current]['CN'].unique())
        
        # TABLE 1: RURAL / URBAN breakdown  
        allocation_summary_1_current = pd.DataFrame({
            'Category': ['Rural', 'Urban'],
            'Reference Allocation': [
                st.session_state.get('rural_clusters', 0),
                st.session_state.get('urban_clusters', 0)
            ],
            'Actual Allocation': [rural_clusters_count_current, urban_clusters_count_current],
            '#Observations': [rural_count_current, urban_count_current]
        })
        
        # TABLE 2: NMP, MP, Tribal, Non-Tribal breakdown
        # Get user inputs - use stored values that were either manually entered or calculated
        user_rural = st.session_state.get('rural_clusters', 0)
        user_urban = st.session_state.get('urban_clusters', 0)
        user_non_mp = st.session_state.get('non_mp_clusters', 0)
        user_non_tribal = st.session_state.get('non_tribal_clusters', 0)
        user_mp = st.session_state.get('million_plus_clusters', 0)
        user_tribal = st.session_state.get('tribal_clusters', 0)
        
        allocation_summary_2_current = pd.DataFrame({
            'Category': ['NMP', 'MP', 'Tribal', 'Non-Tribal'],
            'Reference Allocation': [
                user_non_mp,
                user_mp,
                user_tribal,
                user_non_tribal
            ],
            'Actual Allocation': [nmp_clusters_count_current, mp_clusters_count_current, tribal_clusters_count_current, non_tribal_clusters_count_current],
            '#Observations': [nmp_count_current, mp_count_current, tribal_count_current, non_tribal_count_current]
        })
        
        # ── Response Rate Metrics (2x2 Layout) ────────────────────────────────
        # Calculate response rates first
        status_103_count = (df_processed['interview_status'] == 103).sum()
        total_records = len(df_processed)
        response_rate_103 = (status_103_count / total_records * 100) if total_records > 0 else 0
        
        status_102_103_count = df_processed['interview_status'].isin([102, 103]).sum()
        response_rate_102_103 = (status_102_103_count / total_records * 100) if total_records > 0 else 0
        
        # Create 2x2 grid for metrics
        metric_row1 = st.columns(2)
        with metric_row1[0]:
            st.metric("Total Records", len(df_processed))
        with metric_row1[1]:
            st.metric("Total Clusters", len(df_processed['CN'].unique()))
        
        metric_row2 = st.columns(2)
        with metric_row2[0]:
            st.metric("Response Rate (103 only)", f"{response_rate_103:.2f}%")
        with metric_row2[1]:
            st.metric("Response Rate (102+103)", f"{response_rate_102_103:.2f}%")
        
        # Display Table 1
        st.markdown("**Rural / Urban Split**")
        st.dataframe(allocation_summary_1_current, use_container_width=True, hide_index=True)
        
        # Display Table 2
        st.markdown("**Further Classification**")
        st.dataframe(allocation_summary_2_current, use_container_width=True, hide_index=True)

        # ── CROSSTAB: BREAKDOWN BY BI1B (DISTRICT) ────────────────────────────
        st.markdown("**Breakdown by District**")
        
        # Build crosstab from df_processed
        crosstab_data_current = []
        for district in sorted(df_processed['bi1b'].unique()):
            dist_df = df_processed[df_processed['bi1b'] == district]
            is_urban_dist = dist_df['bi1c'].astype(str).str.contains("Urban", case=False, na=False)
            is_tribal_dist = dist_df['bi1c'].astype(str).str.contains("TRI-", case=False, na=False)
            is_mp_dist = dist_df['bi1c'].astype(str).isin(st.session_state.mp_bi1c_values)
            
            crosstab_data_current.append({
                'District': district,
                '🔴Rural': len(dist_df[~is_urban_dist]['CN'].unique()),
                '🔴Urban': len(dist_df[is_urban_dist]['CN'].unique()),
                '🟢Tribal': len(dist_df[~is_urban_dist & is_tribal_dist]['CN'].unique()),
                '🟢Non-Tribal': len(dist_df[~is_urban_dist & ~is_tribal_dist]['CN'].unique()),
                '🟢MP': len(dist_df[is_urban_dist & is_mp_dist]['CN'].unique()),
                '🟢NMP': len(dist_df[is_urban_dist & ~is_mp_dist]['CN'].unique()),
                '#TOTAL': len(dist_df['CN'].unique())
            })
        
        crosstab_df_current = pd.DataFrame(crosstab_data_current)
        
        # Add TOTAL row with unique cluster counts across all districts
        totals_row_current = pd.DataFrame({
            'District': ['#TOTAL'],
            '🔴Rural': [len(df_processed[~is_urban_current]['CN'].unique())],
            '🔴Urban': [len(df_processed[is_urban_current]['CN'].unique())],
            '🟢Tribal': [len(df_processed[~is_urban_current & is_tribal_current]['CN'].unique())],
            '🟢Non-Tribal': [len(df_processed[~is_urban_current & ~is_tribal_current]['CN'].unique())],
            '🟢MP': [len(df_processed[is_urban_current & is_mp_current]['CN'].unique())],
            '🟢NMP': [len(df_processed[is_urban_current & ~is_mp_current]['CN'].unique())],
            '#TOTAL': [len(df_processed['CN'].unique())]
        })
        crosstab_df_current = pd.concat([crosstab_df_current, totals_row_current], ignore_index=True)
        st.dataframe(crosstab_df_current, use_container_width=True, hide_index=True)

        # Check if clusters are available for display
        if selected_cluster is None:
            st.warning("⚠️ All available clusters have been deleted or merged. Select a different state/district or undo your changes.")
            st.stop()

        # ── CLUSTER CLASSIFICATION ────────────────────────────────────────────
        # Get selected cluster's classification
        selected_cluster_data = df_processed[df_processed['CN'].astype(str) == selected_cluster]
        if len(selected_cluster_data) > 0:
            cluster_bi1c = selected_cluster_data['bi1c'].iloc[0]
            is_urban_sel = 'Urban' in str(cluster_bi1c)
            is_tribal_sel = 'TRI-' in str(cluster_bi1c)
            is_mp_sel = str(cluster_bi1c) in st.session_state.mp_bi1c_values
            
            # Determine primary category
            primary = "Urban" if is_urban_sel else "Rural"
            
            # Determine secondary category
            if is_urban_sel:
                secondary = "MP" if is_mp_sel else "NMP"
            else:
                secondary = "Tribal" if is_tribal_sel else "Non-Tribal"
            
            cluster_type = f"{primary} - {secondary}"
            
            # Display classification with two columns
            col_class1, col_class2 = st.columns(2)
            with col_class1:
                st.metric("🏘️ Rural/Urban", primary)
            with col_class2:
                st.metric("📍 Classification", secondary)
            
            st.caption(
                f"Cluster {selected_cluster} Type: **{cluster_type}** | "
                f"**bi1a:** {selected_cluster_data['bi1a'].iloc[0] if 'bi1a' in selected_cluster_data.columns else ''} | "
                f"**bi1b:** {selected_cluster_data['bi1b'].iloc[0] if 'bi1b' in selected_cluster_data.columns else ''} | "
                f"**bi1c:** {selected_cluster_data['bi1c'].iloc[0] if 'bi1c' in selected_cluster_data.columns else ''} | "
                f"**bi1c_1:** {selected_cluster_data['bi1c_1'].iloc[0] if 'bi1c_1' in selected_cluster_data.columns else ''}"
            )

        # ── Interview Timing Calendar Heatmap (per DBSCAN cluster) ─────────
        _heatmap_data = df_processed[df_processed['CN'].astype(str) == selected_cluster].copy()
        _has_start = 'start_datetime' in _heatmap_data.columns
        _has_end   = 'end_datetime'   in _heatmap_data.columns
        if _has_start or _has_end:
            @st.fragment
            def render_heatmaps():
                from plotly_calheatmap import calheatmap
                import plotly.graph_objects as go
                _dt_options = []
                if _has_start:
                    _dt_options.append('start_datetime')
                if _has_end:
                    _dt_options.append('end_datetime')
                
                # ── Gender Distribution Pie Chart ────────────────────────────
                if 'c2_gender' in _heatmap_data.columns:
                    st.markdown("### 👥 Gender Distribution")
                    
                    # Fragment for gender chart to update independently
                    @st.fragment
                    def render_gender_chart():
                        # Get DBSCAN cluster IDs if available
                        if 'dbscan_cluster' in _heatmap_data.columns:
                            _dbscan_ids = sorted(_heatmap_data[_heatmap_data['dbscan_cluster'] >= 0]['dbscan_cluster'].unique())
                            # Add "Overall" option at the beginning
                            _dbscan_options = ['Overall'] + [f'Sub-Cluster {int(cid)}' for cid in _dbscan_ids]
                            if (_heatmap_data['dbscan_cluster'] == -1).any():
                                _dbscan_options.append('Noise / Outliers')
                            
                            _col_dbscan, _col_status = st.columns(2)
                            with _col_dbscan:
                                _selected_gender_dbscan = st.selectbox(
                                    "DBSCAN Cluster:",
                                    options=_dbscan_options,
                                    key="gender_dbscan_filter",
                                    label_visibility="collapsed"
                                )
                            with _col_status:
                                _gender_status_filter = st.radio(
                                    "Status:",
                                    options=['Status 103 Only', 'Overall'],
                                    horizontal=True,
                                    key="gender_status_filter",
                                    label_visibility="collapsed"
                                )
                        else:
                            _selected_gender_dbscan = 'Overall'
                            _gender_status_filter = 'Overall'
                        
                        # Filter data based on DBSCAN selection
                        if _selected_gender_dbscan == 'Overall':
                            _gender_data = _heatmap_data.copy()
                        elif _selected_gender_dbscan == 'Noise / Outliers':
                            _gender_data = _heatmap_data[_heatmap_data['dbscan_cluster'] == -1].copy()
                        else:
                            # Extract cluster ID from "Sub-Cluster X"
                            _cid = int(_selected_gender_dbscan.split()[-1])
                            _gender_data = _heatmap_data[_heatmap_data['dbscan_cluster'] == _cid].copy()
                        
                        # Apply status filter
                        if _gender_status_filter == 'Status 103 Only':
                            _gender_data = _gender_data[_gender_data.get('interview_status', pd.Series()) == 103].copy()
                        
                        if len(_gender_data) > 0:
                            _gender_counts = _gender_data['c2_gender'].value_counts().sort_index()
                            _gender_total = _gender_counts.sum()
                            _gender_pct = (_gender_counts / _gender_total * 100).round(1)
                            
                            # Create custom text with count for pie slices
                            _custom_text = [f"{count}" for count in _gender_counts.values]
                            
                            # Create hover text with count and percentage
                            _hover_text = [f"<b>{label}</b><br>Count: {count}<br>Percentage: {pct}%" 
                                           for label, count, pct in zip(_gender_counts.index, _gender_counts.values, _gender_pct.values)]
                            
                            fig_gender = go.Figure(data=[go.Pie(
                                labels=_gender_counts.index.astype(str),
                                values=_gender_counts.values,
                                text=_custom_text,
                                hovertemplate='%{customdata}<extra></extra>',
                                customdata=_hover_text,
                                textposition='inside',
                                textinfo='text+percent',
                                marker=dict(colors=['#4A90E2', '#FF6B9D', '#7CB342', '#F5A623', '#9B59B6', '#E67E22'])
                            )])
                            
                            _title_parts = [_selected_gender_dbscan]
                            if _gender_status_filter != 'Overall':
                                _title_parts.append(_gender_status_filter)
                            _title_str = ' • '.join(_title_parts)
                            
                            fig_gender.update_layout(
                                title=f"<b>Gender Distribution - {_title_str} (n={_gender_total})</b>",
                                height=400,
                                showlegend=True,
                                font=dict(size=12)
                            )
                            st.plotly_chart(fig_gender, use_container_width=True, key="gender_pie_chart")
                        else:
                            st.info(f"No data available for the selected filters")
                    
                    render_gender_chart()
                
                
                st.markdown("### 📅 Interview Timing Heatmap")
                # Place DateTime, Filter, and Toggle controls
                _col_dt, _col_filter, _col_toggle = st.columns([2, 2, 1.2])
                with _col_dt:
                    _selected_dt = st.radio(
                        "DateTime column", _dt_options,
                        horizontal=True, key="heatmap_dt_radio"
                    )
                with _col_filter:
                    # Add status filter for heatmaps (matching map filter pattern)
                    _heatmap_status_filter = st.radio(
                        "Filter by Interview Status:",
                        options=['Status 103 Only', 'Overall (All Interview Status)'],
                        horizontal=True, key="heatmap_status_radio"
                    )
                with _col_toggle:
                    st.write("Show Charts:")
                    _show_heatmaps = st.checkbox(
                        "Generate Heatmaps",
                        value=True, key="show_heatmaps_toggle",
                        label_visibility="collapsed"
                    )
                
                if not _show_heatmaps:
                    st.info("Heatmap charts are disabled. Toggle to enable.")
                    return
                
                # Apply status filter to heatmap data
                _hm_data = _heatmap_data.copy()
                if _heatmap_status_filter == 'Status 103 Only':
                    _hm_data = _hm_data[_hm_data.get('interview_status', pd.Series()) == 103].copy()
                
                # Color palette matching DBSCAN scatter plot
                _hm_colors_palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                      '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
                # Map each color to a plotly-compatible sequential colorscale [[0, white], [1, color]]
                def _color_to_scale(hex_color):
                    # Convert hex to rgba for the mid-point (plotly doesn't support 8-digit hex)
                    r, g, b = int(hex_color[1:3], 16), int(hex_color[3:5], 16), int(hex_color[5:7], 16)
                    mid = f'rgba({r},{g},{b},0.5)'
                    return [[0, '#f5f5f5'], [0.5, mid], [1, hex_color]]

                _dt_series = pd.to_datetime(_hm_data[_selected_dt], errors='coerce', dayfirst=True)
                _hm_data['_parsed_dt'] = _dt_series

                if 'dbscan_cluster' in _hm_data.columns:
                    _dbscan_ids = sorted(_hm_data[_hm_data['dbscan_cluster'] >= 0]['dbscan_cluster'].unique())
                    # Include noise (-1) at the end
                    if (_hm_data['dbscan_cluster'] == -1).any():
                        _dbscan_ids_all = list(_dbscan_ids) + [-1]
                    else:
                        _dbscan_ids_all = list(_dbscan_ids)
                else:
                    _dbscan_ids_all = []

                if len(_dbscan_ids_all) > 0:
                    for _didx, _db_id in enumerate(_dbscan_ids_all):
                        _sub = _hm_data[_hm_data['dbscan_cluster'] == _db_id]
                        _valid = _sub['_parsed_dt'].dropna()
                        if len(_valid) == 0:
                            continue
                        _cal_df = pd.DataFrame({'date': _valid.dt.date, 'count': 1})
                        _cal_df['date'] = pd.to_datetime(_cal_df['date'])
                        _cal_agg = _cal_df.groupby('date', as_index=False)['count'].sum()
                        if _db_id == -1:
                            _cs = [[0, '#f5f5f5'], [0.5, 'rgba(214,39,40,0.5)'], [1, '#d62728']]
                            _lbl = 'Noise / Outliers'
                        else:
                            _c = _hm_colors_palette[_didx % len(_hm_colors_palette)]
                            _cs = _color_to_scale(_c)
                            _lbl = f'Sub-Cluster {int(_db_id)}'
                        fig_cal = calheatmap(
                            data=_cal_agg,
                            x='date',
                            y='count',
                            title=f'{_lbl} — {_selected_dt} — CN {selected_cluster}',
                            colorscale=_cs,
                            gap=2,
                            month_lines=True,
                            years_title=True,
                            showscale=False,
                            annotations=True,
                            total_height=200,
                            name=_lbl,
                        )
                        st.plotly_chart(fig_cal, use_container_width=True)
                else:
                    # Fallback: single heatmap for entire cluster
                    _valid = _dt_series.dropna()
                    if len(_valid) > 0:
                        _cal_df = pd.DataFrame({'date': _valid.dt.date, 'count': 1})
                        _cal_df['date'] = pd.to_datetime(_cal_df['date'])
                        _cal_agg = _cal_df.groupby('date', as_index=False)['count'].sum()
                        fig_cal = calheatmap(
                            data=_cal_agg, x='date', y='count',
                            title=f'{_selected_dt} — CN {selected_cluster}',
                            colorscale='YlOrRd', gap=2, month_lines=True,
                            years_title=True, showscale=True, annotations=True,
                            total_height=200, name='Interviews',
                        )
                        st.plotly_chart(fig_cal, use_container_width=True)
                    else:
                        st.info(f"No valid datetime data in {_selected_dt}.")
            
            render_heatmaps()

        # ── DATASET preview ───────────────────────────────────────────────────
        st.subheader("Dataset")
        cn_filtered_df = df_processed[df_processed['CN'].astype(str) == selected_cluster].copy()
        
        # Always sort by hh_latitude in ascending order before creating select_id
        if lat_col_name and lat_col_name in cn_filtered_df.columns:
            cn_filtered_df = cn_filtered_df.sort_values(lat_col_name, ascending=True).reset_index(drop=True)
        
        # ── Missing Values Analysis for Selected Cluster ──────────────────────
        # missing_values = cn_filtered_df.isnull().sum()
        # total_missing = missing_values.sum()
        # total_records = len(cn_filtered_df)
        
        # if total_missing > 0:
        #     st.warning(f"🔔 Missing Values in Cluster {selected_cluster}: {total_missing} total missing value(s)")
            
        #     # Show missing values by column
        #     missing_by_col = missing_values[missing_values > 0].sort_values(ascending=False)
        #     missing_df = pd.DataFrame({
        #         'Column': missing_by_col.index,
        #         'Missing Count': missing_by_col.values,
        #         'Percentage': (missing_by_col.values / total_records * 100).round(2)
        #     })
        #     st.dataframe(missing_df, use_container_width=True, hide_index=True)
        # else:
        #     st.success(f"✅ Cluster {selected_cluster}: No missing values")
        
        # Add select ID column starting from 0
        # Remove select_id if it already exists to avoid duplication
        if 'select_id' in cn_filtered_df.columns:
            cn_filtered_df = cn_filtered_df.drop(columns=['select_id'])
        cn_filtered_df.insert(0, 'select_id', range(len(cn_filtered_df)))
        
        # Add DBSCAN Cluster column (DBSCAN sub-cluster if available, else CN)
        if 'dbscan_cluster' in cn_filtered_df.columns:
            cn_filtered_df.insert(1, 'DBSCAN Cluster', cn_filtered_df['dbscan_cluster'].astype(str))
        else:
            cn_filtered_df.insert(1, 'DBSCAN Cluster', cn_filtered_df['CN'].astype(str))
        
        # Reorder columns as specified
        col_order = ['select_id', 'DBSCAN Cluster', 'CN', 'original_CN', 'original_concat', 'HH_ID', 'MEM_ID', 'bi1b', 'bi1c', 'bi1c_1', 'hh_latitude', 'hh_longitude', 'hh_address', 'start_datetime', 'end_datetime', 'created_by', 'c2_name','c2_age', 'c2_gender', 'cluster_number', 'household_number', 'member_number', 'interview_status', 'concat', 'index']
        available_cols = [col for col in col_order if col in cn_filtered_df.columns]
        cn_filtered_df_reordered = cn_filtered_df[available_cols]
        cn_filtered_df_display = filter_dataframe(cn_filtered_df_reordered, key="main_df_filter")
        st.dataframe(cn_filtered_df_display, use_container_width=True)

        # ── Duplicates in Name ────────────────────────────────────────────
        st.subheader("Duplicates in Name of Individuals (c2_name)")
        
        # Find rows with duplicate c2_name
        duplicates_mask = cn_filtered_df['c2_name'].duplicated(keep=False)
        if duplicates_mask.any():
            duplicates_df = cn_filtered_df[duplicates_mask].copy()
            # Select columns to display
            dup_cols = ['c2_name', 'c2_age', 'hh_address', 'start_datetime', 'end_datetime']
            available_dup_cols = [col for col in dup_cols if col in duplicates_df.columns]
            duplicates_display = duplicates_df[available_dup_cols].sort_values('c2_name')
            
            st.dataframe(duplicates_display, use_container_width=True, hide_index=True)
            st.caption(f"Found {len(duplicates_display)} duplicate record(s) for {len(duplicates_df['c2_name'].unique())} unique names in this cluster")
        else:
            st.info("✅ No duplicates found in name for this cluster")

        # ── Complete Dataset Display ──────────────────────────────────────────
        st.subheader("Complete Dataset (All Records)")
        
        # Ensure remarks are up-to-date from session state before displaying/downloading
        for idx, remark in st.session_state.remarks_dict.items():
            df_processed.loc[df_processed['index'] == idx, 'remark'] = remark
        
        # Add classification columns to df_processed
        df_with_id = df_processed.copy()
        
        # Determine Rural/Urban
        is_urban_all = df_with_id['bi1c'].astype(str).str.contains("Urban", case=False, na=False)
        df_with_id.insert(0, 'Rural/Urban', is_urban_all.apply(lambda x: 'Urban' if x else 'Rural'))
        
        # Determine Classification (Tribal/Non-Tribal or MP/NMP)
        is_tribal_all = df_with_id['bi1c'].astype(str).str.contains("TRI-", case=False, na=False)
        is_mp_all = df_with_id['bi1c'].astype(str).isin(st.session_state.mp_bi1c_values)
        
        def get_classification(row):
            if row['Rural/Urban'] == 'Urban':
                return 'MP' if is_mp_all[row.name] else 'NMP'
            else:
                return 'Tribal' if is_tribal_all[row.name] else 'Non-Tribal'
        
        df_with_id.insert(1, 'Classification', df_with_id.apply(get_classification, axis=1))
        
        _MAX_DISPLAY_ROWS = 500
        if len(df_with_id) > _MAX_DISPLAY_ROWS:
            st.caption(f"Showing first {_MAX_DISPLAY_ROWS} of {len(df_with_id)} rows (download for full data)")
            st.dataframe(df_with_id.head(_MAX_DISPLAY_ROWS), use_container_width=True, height=400)
        else:
            st.dataframe(df_with_id, use_container_width=True, height=400)
        
        # Download complete dataset
        csv_complete = df_with_id.to_csv(index=False).encode('utf-8')
        
        # Store bytes in session state so on_click callback can access them
        st.session_state['_csv_complete_bytes'] = csv_complete
        
        def on_download_complete_dataset():
            import os, tempfile
            data = st.session_state.get('_csv_complete_bytes', b'')
            if data:
                tmp_path = os.path.join(tempfile.gettempdir(), 'cluster_validator_output.csv')
                with open(tmp_path, 'wb') as f:
                    f.write(data)
                st.session_state['output_checksum'] = hashlib.md5(data).hexdigest()
                st.session_state['output_temp_path'] = tmp_path
                # Enable PDF download when complete dataset is downloaded
                st.session_state['pdf_downloaded'] = True
        
        st.download_button(
            "⬇ Download Complete Dataset",
            csv_complete,
            "complete_data_with_remarks.csv",
            "text/csv",
            use_container_width=True,
            on_click=on_download_complete_dataset
        )
        
        # ── PDF REPORT DOWNLOAD (shown immediately below Complete Dataset button) ───────────
        output_md5_val = st.session_state.get('output_checksum', None)
        output_tmp_path = st.session_state.get('output_temp_path', None)
        pdf_downloaded_flag = st.session_state.get('pdf_downloaded', False)
        
        if pdf_downloaded_flag:
            # PDF can only be downloaded after complete dataset is downloaded
            try:
                # Get state name for PDF title and filename
                parent_md5_val = str(df['Parent MD5'].dropna().iloc[0]) if 'Parent MD5' in df.columns and not df['Parent MD5'].dropna().empty else "(empty)"
                uploaded_md5_val = calculate_md5_checksum(uploaded_file)
                state_name = df['bi1a'].iloc[0] if len(df) > 0 else 'Unknown'
                # Calculate response rates for PDF using COMPLETE dataset (df, not df_processed)
                status_103_count_pdf = (df['interview_status'] == 103).sum()
                total_records_pdf = len(df)
                response_rate_103_pdf = (status_103_count_pdf / total_records_pdf * 100) if total_records_pdf > 0 else 0
                
                status_102_103_count_pdf = df['interview_status'].isin([102, 103]).sum()
                response_rate_102_103_pdf = (status_102_103_count_pdf / total_records_pdf * 100) if total_records_pdf > 0 else 0
                
                pdf_bytes = generate_pdf_report(
                    parent_md5=parent_md5_val,
                    uploaded_md5=uploaded_md5_val,
                    output_md5=output_md5_val,
                    state_name=state_name,
                    confirm_tribal=st.session_state.get('confirm_tribal', False),
                    confirm_urban=st.session_state.get('confirm_urban', False),
                    confirm_mp=st.session_state.get('confirm_mp', False),
                    mp_cities_count=st.session_state.get('mp_cities', 0),
                    eps_val=st.session_state.get('eps_input', st.session_state.get('eps_default_val', '3.0')),
                    min_neighbors_val=st.session_state.get('min_neighbors_input', st.session_state.get('min_neighbors_default_val', '10')),
                    allocation_summary_1=allocation_summary_1_current,
                    allocation_summary_2=allocation_summary_2_current,
                    crosstab_df=crosstab_df_current,
                    regular_clusters=regular_clusters_current,
                    irregular_clusters=irregular_clusters_current,
                    steps_tracker=st.session_state.get('steps_tracker', []),
                    removed_cns=st.session_state.get('removed_cns', []),
                    reconsidered_cns=st.session_state.get('reconsidered_cns', set()),
                    merge_operations=st.session_state.get('merge_operations', []),
                    break_operations=st.session_state.get('break_operations', []),
                    move_operations=st.session_state.get('move_operations', []),
                    response_rate_103=response_rate_103_pdf,
                    response_rate_102_103=response_rate_102_103_pdf,
                    total_clusters=len(df['CN'].unique()),
                    total_records=len(df),
                    cluster_summary=cluster_summary_current,
                    mp_bi1c_values=st.session_state.get('mp_bi1c_values', [])
                )
                
                def on_pdf_download():
                    # Reset flag after PDF download so button reappears
                    st.session_state['pdf_downloaded'] = False
                
                st.download_button(
                    "📥 Download PDF Report",
                    pdf_bytes,
                    f"cluster_report_{state_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    "application/pdf",
                    use_container_width=True,
                    on_click=on_pdf_download
                )
            except Exception as e:
                st.error(f"Error generating PDF: {e}")
        else:
            st.info("ℹ️ Click **⬇ Download Complete Dataset** above first to enable PDF report download")

        # ── NEAREST NEIGHBOUR ANALYSIS ───────────────────────────────────────
        
        st.subheader("📍 Nearest Neighbour Analysis - Completed Interview (103)")

        # Place controls OUTSIDE fragment so they update state and trigger map rerun
        if lat_col_name and lon_col_name:
            nn_c1, nn_c2 = st.columns(2)
            with nn_c1:
                nn_metric = st.selectbox(
                    "Distance Metric",
                    ["Haversine", "Euclidean"],
                    index=0,
                    key="nn_metric_left",
                    help="Haversine: great-circle distance (recommended). Euclidean: flat-earth approximation."
                )
            with nn_c2:
                nn_top_n = st.slider(
                    "Top N Clusters",
                    min_value=1, max_value=10, value=3, step=1,
                    key="nn_topn_left"
                )

        @st.fragment
        def render_nearest_neighbour():
            if lat_col_name and lon_col_name:
                # Get values from session state (updated by controls above)
                nn_metric = st.session_state.get('nn_metric_left', 'Haversine')
                nn_top_n = st.session_state.get('nn_topn_left', 3)

                nn_result = find_nearest_clusters(
                    df_processed, selected_cluster, lat_col_name, lon_col_name,
                    distance_metric=nn_metric.lower(), top_n=nn_top_n
                )

                if not nn_result.empty:
                    # Build combined household-level records for selected + nearest clusters
                    all_cns = [selected_cluster] + nn_result['CN'].tolist()
                    nn_hh_cols = ['CN', 'HH_ID', 'MEM_ID']
                    if 'hh_address' in df_processed.columns:
                        nn_hh_cols.append('hh_address')
                    if lat_col_name:
                        nn_hh_cols.append(lat_col_name)
                    if lon_col_name:
                        nn_hh_cols.append(lon_col_name)
                    available_nn_cols = [c for c in nn_hh_cols if c in df_processed.columns]
                    nn_hh_df = df_processed[df_processed['CN'].isin(all_cns)][available_nn_cols].copy()
                    # Tag source cluster
                    nn_hh_df.insert(0, 'type', nn_hh_df['CN'].apply(
                        lambda x: '★ Selected' if str(x) == str(selected_cluster) else 'Neighbour'
                    ))
                    st.write(
                        f"**Selected cluster `{selected_cluster}` + {nn_top_n} nearest clusters** "
                        f"— {len(nn_hh_df)} records ({nn_metric} distance):"
                    )
                    st.dataframe(nn_hh_df, use_container_width=True, hide_index=True, height=400)
                else:
                    st.info("No nearby clusters found.")
            else:
                st.warning("Latitude / longitude columns not detected — cannot run nearest neighbour analysis.")
        
        if lat_col_name and lon_col_name:
            render_nearest_neighbour()
        else:
            st.warning("Latitude / longitude columns not detected — cannot run nearest neighbour analysis.")

        st.markdown("---")

        # ── REASSIGN (MERGE) ─────────────────────────────────────────────────
        st.markdown("**🔀 Reassign (Merge)**")
        st.caption("Merge current cluster into another CN")
        with st.form("reassign_cn_form"):
            target_cns = sorted([cn for cn in df_processed['CN'].astype(str).unique() if cn != str(selected_cluster)])
            target_cn = st.selectbox(
                f"Merge CN {selected_cluster} into:",
                options=target_cns if target_cns else ["No other CN available"],
                key="reassign_target_cn"
            )
            merge_remarks = st.text_area(
                "Remarks (optional)",
                placeholder="Enter reason for merge...",
                height=80,
                key="merge_remarks"
            )
            reassign_submit = st.form_submit_button("✅ Merge", use_container_width=True)

        if reassign_submit and target_cn and target_cn != "No other CN available":
            st.session_state.merge_operations.append((str(selected_cluster), str(target_cn)))
            st.session_state.steps_tracker.append({
                'S.No': len(st.session_state.steps_tracker) + 1,
                'Step': 'Reassign (Merge)',
                'Details': f"CN {selected_cluster} → {target_cn}",
                'Remark': merge_remarks if merge_remarks.strip() else '-',
                '_op_type': 'merge',
                '_op_data': {'source': str(selected_cluster), 'target': str(target_cn)}
            })
            st.success(f"✅ Merged CN {selected_cluster} → {target_cn}")
            st.rerun()

    # ── RIGHT PANEL ───────────────────────────────────────────────────────────
    with right:
        # Display steps tracker at the top if available
        if 'steps_tracker' in st.session_state and st.session_state.steps_tracker:
            st.subheader("📋 Processing Steps")
            steps_df_right = pd.DataFrame(st.session_state.steps_tracker)
            # Show S.No, Step, Details, and Remark columns
            display_cols = []
            if 'S.No' in steps_df_right.columns:
                display_cols.append('S.No')
            display_cols.append('Step')
            display_cols.append('Details')
            if 'Remark' in steps_df_right.columns:
                display_cols.append('Remark')
            steps_df_right_display = steps_df_right[display_cols]
            st.dataframe(steps_df_right_display, use_container_width=True, hide_index=True)
            
            # ── UNDO BY S.No DROPDOWN ────────────────────────────────────────
            if 'S.No' in steps_df_right.columns:
                available_snos = sorted(steps_df_right['S.No'].dropna().astype(int).tolist())
                if available_snos:
                    undo_sno = st.selectbox(
                        "⏪ Undo from S.No (all steps from this number onward will be undone)",
                        options=[""] + [str(s) for s in available_snos],
                        index=0,
                        key="undo_sno_select"
                    )
                    if undo_sno and st.button("⏪ Undo", use_container_width=True, type="secondary", key="undo_btn"):
                        sno_val = int(undo_sno)
                        steps_to_undo = len([s for s in st.session_state.steps_tracker if s.get('S.No', 0) >= sno_val])
                        undo_to_sno(sno_val)
                        st.success(f"✅ Undone {steps_to_undo} step(s) from S.No {sno_val} onward")
                        st.rerun()
            st.markdown("---")
        st.markdown("📍 EPS Reference (Haversine Metric)")
        
        try:
            eps_numeric = float(eps_val) if eps_val else float(st.session_state.eps_default_val)
            
            # EPS is now in KILOMETERS (Haversine metric)
            # Haversine calculates great-circle distance on Earth's surface
            # Accurate for any latitude (not limited to India)
            
            col_dist1, col_dist2 = st.columns(2)
            
            with col_dist1:
                st.metric(
                    "🌍 Search Radius",
                    f"{eps_numeric:.2f} km",
                    "Haversine (Great-circle distance)"
                )
            
            with col_dist2:
                earth_radius_km = 6371
                eps_radians = eps_numeric / earth_radius_km
                eps_degrees = np.degrees(eps_radians)
                st.metric(
                    "📐 Equivalent Degrees",
                    f"≈ {eps_degrees:.4f}°",
                    f"{eps_numeric:.2f} km ÷ 6371 km"
                )
            
            # Recommendation based on EPS value
            st.markdown("**Usage Guide:**")
            if eps_numeric <= 2:
                st.success(f"✅ **Hyper-local** (≤2 km): Single village/neighborhood clustering")
            elif eps_numeric <= 5:
                st.info(f"✅ **Local** (2-5 km): Cluster within panchayat/block")
            elif eps_numeric <= 10:
                st.info(f"✅ **Sub-district** (5-10 km): Multiple blocks within district")
            elif eps_numeric <= 30:
                st.warning(f"⚠️ **District-wide** (10-30 km): Multi-block clustering")
            else:
                st.warning(f"❌ **Too Large** (>{eps_numeric:.1f} km): Consider reducing EPS value")
            
            # Show technical details in expander
            with st.expander("🔬 Technical Details", expanded=False):
                st.markdown("**Haversine Metric:**")
                st.code(f"""
# Earth radius: 6371 km
eps_km = {eps_numeric}
eps_radians = {eps_radians:.6f}

# Clustering uses: metric='haversine', algorithm='ball_tree'
# Distance formula: Great-circle distance on sphere
                """, language="python")
                st.markdown("**Reference (India):**")
                st.markdown(f"""
- **Latitude:** 1° ≈ 111.32 km (constant everywhere)
- **Longitude:** 1° ≈ 103.2 km (at 22°N latitude)
- **Your EPS ({eps_numeric} km):** Searches in all directions equally
- **Algorithm:** scikit-learn DBSCAN with haversine + ball_tree
                """)
                    
        except ValueError:
            st.error("Invalid EPS value - enter a number")

        # ── CLUSTER DISTRIBUTION HISTOGRAMS (Overall vs Status 103) ─────────
        st.markdown("### Cluster Distribution Analysis")
        
        try:
            # Create side-by-side histograms
            hist_col1, hist_col2 = st.columns(2)
            
            with hist_col1:
                fig, ax = plt.subplots(figsize=(6, 4))
                ax.hist(cluster_summary_current['overall_count'], bins=20, color='#3498db', edgecolor='black', alpha=0.7)
                ax.set_title('Overall Distribution', fontsize=12, fontweight='bold')
                ax.set_xlabel('Overall Count')
                ax.set_ylabel('Frequency')
                ax.grid(True, alpha=0.3, linestyle='--')
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
            
            with hist_col2:
                fig, ax = plt.subplots(figsize=(6, 4))
                ax.hist(cluster_summary_current['status_103_count'], bins=20, color='#2ecc71', edgecolor='black', alpha=0.7)
                ax.set_title('Status 103 (Completed) Distribution', fontsize=12, fontweight='bold')
                ax.set_xlabel('Status 103 Count')
                ax.set_ylabel('Frequency')
                ax.grid(True, alpha=0.3, linestyle='--')
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
        except Exception as e:
            st.error(f"Error creating distribution histograms: {e}")


        st.subheader("Location Density")
        # Graph 1 & Graph 2 side by side
        g1, g2 = st.columns(2)

        # Get data for selected cluster (lat/lon col names already detected above)
        cluster_data = df_processed[df_processed['CN'].astype(str) == selected_cluster].copy()
        
        # SAFETY: Ensure dbscan_cluster column exists (fallback if not transferred properly)
        if 'dbscan_cluster' not in cluster_data.columns:
            cluster_data['dbscan_cluster'] = -1  # Initialize all as noise

        with g1:
            
            if lat_col_name and lon_col_name:
                try:
                    # Convert to numeric and round coordinates to 3 decimal places
                    cluster_data_copy = cluster_data.copy()
                    cluster_data_copy[lat_col_name] = pd.to_numeric(cluster_data_copy[lat_col_name], errors='coerce')
                    cluster_data_copy[lon_col_name] = pd.to_numeric(cluster_data_copy[lon_col_name], errors='coerce')
                    cluster_data_copy['lat_r'] = cluster_data_copy[lat_col_name].round(3)
                    cluster_data_copy['lon_r'] = cluster_data_copy[lon_col_name].round(3)
                    
                    # Drop rows with NaN coordinates
                    cluster_data_copy = cluster_data_copy.dropna(subset=['lat_r', 'lon_r'])
                    
                    if len(cluster_data_copy) > 0:
                        # Count by rounded coordinates
                        counts = (
                            cluster_data_copy.groupby(['lat_r', 'lon_r'])
                            .size()
                            .reset_index(name='n')
                        )
                        
                        fig, ax = plt.subplots(figsize=(5, 5))
                        ax.scatter(counts['lon_r'], counts['lat_r'], s=counts['n']*20, alpha=0.6)
                        ax.set_title(f'Cluster {selected_cluster}')
                        ax.set_xlabel('Longitude')
                        ax.set_ylabel('Latitude')
                        st.pyplot(fig, use_container_width=True)
                    else:
                        st.info("No valid location data available for this cluster.")
                except Exception as e:
                    st.error(f"Error creating scatter plot: {e}")
            else:
                st.info("No location data available.")

        with g2:
            if lat_col_name and lon_col_name and len(cluster_data) > 0:
                try:
                    # Convert coordinates to numeric
                    cluster_data_hex = cluster_data.copy()
                    cluster_data_hex[lat_col_name] = pd.to_numeric(cluster_data_hex[lat_col_name], errors='coerce')
                    cluster_data_hex[lon_col_name] = pd.to_numeric(cluster_data_hex[lon_col_name], errors='coerce')
                    
                    # Drop rows with NaN coordinates
                    cluster_data_hex = cluster_data_hex.dropna(subset=[lat_col_name, lon_col_name])
                    
                    if len(cluster_data_hex) > 0:
                        fig, ax = plt.subplots(figsize=(6, 6))
                        hb = ax.hexbin(
                            cluster_data_hex[lon_col_name],
                            cluster_data_hex[lat_col_name],
                            gridsize=35,
                            cmap='viridis',
                            mincnt=1
                        )
                        ax.set_title(f'Cluster {selected_cluster} — Density / Overlap')
                        ax.set_xlabel('Longitude')
                        ax.set_ylabel('Latitude')
                        plt.colorbar(hb, ax=ax, label='Household count (overlap)')
                        st.pyplot(fig, use_container_width=True)
                    else:
                        st.info("No valid location data available for this cluster.")
                except Exception as e:
                    st.error(f"Error creating hexbin plot: {e}")
            else:
                st.info("No location data available.")


        # ── DBSCAN Predictions Visualization (Full data vs Clustered) ────────
        st.markdown("### 🔍 DBSCAN Sub-Cluster Detection")
        
        if lat_col_name and lon_col_name and 'dbscan_cluster' in cluster_data.columns:
            try:
                # Create attractive figure
                fig, ax = plt.subplots(figsize=(12, 8), facecolor='white')
                ax.set_facecolor('#f8f9fa')
                
                # Full dataset points (all points with lat/lon)
                full_data = cluster_data[[lat_col_name, lon_col_name]].dropna()
                
                # Plot full dataset background (very light)
                ax.scatter(full_data[lon_col_name], full_data[lat_col_name], 
                          s=80, alpha=0.15, color='#888888', marker='.',
                          label=f"All observations ({len(full_data)})", zorder=1)
                
                # Enhanced color palette with vibrant colors
                colors_palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                 '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
                                 '#1f77b4', '#ff7f0e', '#2ca02c']  # Repeat for more clusters
                
                # Get unique cluster IDs (excluding -1 noise and -2 excluded)
                unique_clusters = sorted(cluster_data[cluster_data['dbscan_cluster'] >= 0]['dbscan_cluster'].unique())
                
                # Plot each cluster with a different color
                for idx, cluster_id in enumerate(unique_clusters):
                    cluster_pts = cluster_data[cluster_data['dbscan_cluster'] == cluster_id][[lat_col_name, lon_col_name]].dropna()
                    color = colors_palette[idx % len(colors_palette)]
                    ax.scatter(cluster_pts[lon_col_name], cluster_pts[lat_col_name],
                              s=120, alpha=0.35, color=color, marker='o',
                              edgecolors='none', linewidth=0,
                              label=f"Sub-Cluster {int(cluster_id)} ({len(cluster_pts)} pts)", 
                              zorder=2)
                
                # Plot EXCLUDED points (invalid/missing geocodes) with red X
                excluded_data = cluster_data[cluster_data['dbscan_cluster'] == -2][[lat_col_name, lon_col_name]].dropna()
                if len(excluded_data) > 0:
                    ax.scatter(excluded_data[lon_col_name], excluded_data[lat_col_name],
                              s=120, alpha=0.5, color='#d62728', marker='X', linewidth=2,
                              edgecolors='#8B0000', 
                              label=f"EXCLUDED (Invalid Geocodes) ({len(excluded_data)} pts)", zorder=4)
                
                # Plot NOISE points (valid coords but not in cluster) with different style
                noise_data = cluster_data[cluster_data['dbscan_cluster'] == -1][[lat_col_name, lon_col_name]].dropna()
                if len(noise_data) > 0:
                    ax.scatter(noise_data[lon_col_name], noise_data[lat_col_name],
                              s=100, alpha=0.35, color='#ff7f0e', marker='^', linewidth=1,
                              edgecolors='none', 
                              label=f"Noise/Outliers ({len(noise_data)} pts)", zorder=3)
                
                # Enhanced styling
                ax.set_title(f'DBSCAN Sub-Cluster Detection - Main Cluster {selected_cluster}\nEPS={float(eps_val):.2f}km | MIN_NEIGHBORS={int(min_neighbors)}', 
                            fontsize=13, fontweight='bold', pad=20, color='#1a1a1a')
                ax.set_xlabel('Longitude', fontsize=11, fontweight='bold')
                ax.set_ylabel('Latitude', fontsize=11, fontweight='bold')
                
                # Improved grid
                ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, color='gray')
                ax.set_axisbelow(True)
                
                # Enhanced legend (upper right)
                legend = ax.legend(loc='upper right', framealpha=0.98, fontsize=9.5, 
                                  edgecolor='gray', fancybox=True, shadow=True)
                legend.get_frame().set_linewidth(1)
                
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
            except Exception as e:
                st.error(f"Error creating DBSCAN prediction plot: {e}")
        else:
            st.info("ℹ️ DBSCAN sub-clusters will appear here after clustering is applied.")

        # ── DBSCAN Predictions for Interview Status = 103 ────────────────────
        
        if lat_col_name and lon_col_name and 'dbscan_cluster' in cluster_data.columns:
            try:
                # Filter for interview_status = 103
                status_103_data = cluster_data[cluster_data['interview_status'] == 103].copy()
                
                if len(status_103_data) > 0:
                    # Create attractive figure
                    fig, ax = plt.subplots(figsize=(12, 8), facecolor='white')
                    ax.set_facecolor('#f8f9fa')
                    
                    # Full dataset points (all status 103 points with lat/lon)
                    full_data_103 = status_103_data[[lat_col_name, lon_col_name]].dropna()
                    
                    # Plot full dataset background (very light)
                    ax.scatter(full_data_103[lon_col_name], full_data_103[lat_col_name], 
                              s=80, alpha=0.15, color='#888888', marker='.',
                              label=f"All observations ({len(full_data_103)})", zorder=1)
                    
                    # Enhanced color palette with vibrant colors
                    colors_palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                     '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
                                     '#1f77b4', '#ff7f0e', '#2ca02c']
                    
                    # Get unique cluster IDs (excluding -1 which is noise)
                    unique_clusters_103 = sorted(status_103_data[status_103_data['dbscan_cluster'] >= 0]['dbscan_cluster'].unique())
                    
                    # Plot each cluster with a different color
                    for idx, cluster_id in enumerate(unique_clusters_103):
                        cluster_pts_103 = status_103_data[status_103_data['dbscan_cluster'] == cluster_id][[lat_col_name, lon_col_name]].dropna()
                        color = colors_palette[idx % len(colors_palette)]
                        ax.scatter(cluster_pts_103[lon_col_name], cluster_pts_103[lat_col_name],
                                  s=120, alpha=0.35, color=color, marker='o',
                                  edgecolors='none', linewidth=0,
                                  label=f"Sub-Cluster {int(cluster_id)} ({len(cluster_pts_103)} pts)", 
                                  zorder=2)
                    
                    # Plot noise points with distinctive style
                    noise_data_103 = status_103_data[status_103_data['dbscan_cluster'] == -1][[lat_col_name, lon_col_name]].dropna()
                    if len(noise_data_103) > 0:
                        ax.scatter(noise_data_103[lon_col_name], noise_data_103[lat_col_name],
                                  s=100, alpha=0.35, color='#d62728', marker='X', linewidth=1,
                                  edgecolors='none', 
                                  label=f"Noise/Outliers ({len(noise_data_103)} pts)", zorder=3)
                    
                    # Enhanced styling
                    ax.set_title(f'DBSCAN Sub-Cluster Detection - Main Cluster {selected_cluster} (Status 103 Only)\nEPS={float(eps_val):.2f}km | MIN_NEIGHBORS={int(min_neighbors)}', 
                                fontsize=13, fontweight='bold', pad=20, color='#1a1a1a')
                    ax.set_xlabel('Longitude', fontsize=11, fontweight='bold')
                    ax.set_ylabel('Latitude', fontsize=11, fontweight='bold')
                    
                    # Improved grid
                    ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, color='gray')
                    ax.set_axisbelow(True)
                    
                    # Enhanced legend (upper right)
                    legend = ax.legend(loc='upper right', framealpha=0.98, fontsize=9.5, 
                                      edgecolor='gray', fancybox=True, shadow=True)
                    legend.get_frame().set_linewidth(1)
                    
                    plt.tight_layout()
                    st.pyplot(fig, use_container_width=True)
                else:
                    st.info("ℹ️ No observations with interview_status = 103 in this cluster.")
            except Exception as e:
                st.error(f"Error creating DBSCAN status 103 plot: {e}")
        else:
            st.info("ℹ️ DBSCAN clustering not yet applied.")

        # ── Folium Map (only for selected cluster) ──────────────────────────
        st.subheader("Map")
        
        
        @st.fragment
        @st.fragment
        def render_map():
            # Toggle to switch between Status 103 Only and Overall (all statuses)
            col_map_filter2, col_map_filter3 = st.columns([2.5, 1.5])
            with col_map_filter2:
                map_filter = st.radio(
                    "Show Individuals:",
                    options=['Status 103 Only', 'Overall (All Interview Status)'],
                    index=0 if st.session_state.map_filter_mode == 'Status 103 Only' else 1,
                    horizontal=True,
                    key='map_filter_radio'
                )
                st.session_state.map_filter_mode = map_filter
            with col_map_filter3:
                show_circles = st.checkbox(
                    "Show Circles",
                    value=True,
                    key='map_show_circles'
                )
            
            # Force map rebuild by clearing cache when cluster changes
            if 'current_cluster_map' not in st.session_state or st.session_state.current_cluster_map != selected_cluster:
                # Clear all map-related session states
                for key in list(st.session_state.keys()):
                    if key.startswith('folium_map_'):
                        del st.session_state[key]
                st.session_state.current_cluster_map = selected_cluster
            
            if lat_col_name and lon_col_name:
                # Filter for interview_status based on toggle
                if st.session_state.map_filter_mode == 'Status 103 Only':
                    map_data = cluster_data[cluster_data['interview_status'] == 103][[lat_col_name, lon_col_name]].dropna()
                else:
                    map_data = cluster_data[[lat_col_name, lon_col_name]].dropna().copy()

                if not map_data.empty:
                    # Convert to numeric (handles both string and numeric data)
                    map_data[lat_col_name] = pd.to_numeric(map_data[lat_col_name], errors='coerce')
                    map_data[lon_col_name] = pd.to_numeric(map_data[lon_col_name], errors='coerce')
                    map_data = map_data.dropna()
                    
                    if not map_data.empty:
                        center = [map_data[lat_col_name].mean(), map_data[lon_col_name].mean()]
                    m = folium.Map(
                        location=center,
                        zoom_start=7,
                        tiles=None,
                        control_scale=True
                    )
                    
                    # Add tile layers
                    folium.TileLayer(
                        tiles='cartodb positron',
                        attr='CartoDB Positron',
                        name='CartoDB Positron',
                        overlay=False,
                        control=True
                    ).add_to(m)
                    
                    folium.TileLayer(
                        tiles='https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}',
                        attr='Google Satellite',
                        name='Google Satellite',
                        overlay=False,
                        control=True
                    ).add_to(m)
                    
                    folium.TileLayer(
                        tiles='https://mt1.google.com/vt/lyrs=m&x={x}&y={y}&z={z}',
                        attr='Google Maps',
                        name='Google Maps',
                        overlay=False,
                        control=True
                    ).add_to(m)
                    
                    # Search bar (Nominatim/OSM geocoder built into folium)
                    Geocoder(
                        collapsed=False,
                        position='topright',
                        add_marker=True,
                        zoom=13
                    ).add_to(m)
                    
                    # Measure tool
                    MeasureControl(
                        position='topleft',
                        primary_length_unit='kilometers',
                        secondary_length_unit='meters',
                        primary_area_unit='sqkilometers',
                        secondary_area_unit='sqmeters'
                    ).add_to(m)
                    
                    # Markers with clustering
                    marker_cluster = MarkerCluster().add_to(m)
                    
                    # Filter households based on toggle
                    if st.session_state.map_filter_mode == 'Status 103 Only':
                        cluster_households = cluster_data[cluster_data['interview_status'] == 103]
                    else:
                        cluster_households = cluster_data
                    
                    # Function to generate random but deterministic color from HH_ID
                    def generate_color_from_hh_id(hh_id):
                        """Generate a random but deterministic color based on HH_ID hash"""
                        import hashlib
                        import colorsys
                        
                        hash_val = int(hashlib.md5(str(hh_id).encode()).hexdigest(), 16)
                        hue = (hash_val % 360) / 360.0
                        saturation = 0.6 + (hash_val // 360 % 40) / 100.0  # 0.6-1.0 for vibrant
                        value = 0.8 + (hash_val // 14400 % 20) / 100.0  # 0.8-1.0 for bright
                        r, g, b = colorsys.hsv_to_rgb(hue, saturation, value)
                        return '#{:02x}{:02x}{:02x}'.format(int(r*255), int(g*255), int(b*255))
                    
                    # Create color mapping for HH_IDs
                    hh_colors = {}
                    unique_hh_ids = cluster_households['HH_ID'].unique()
                    for hh_id in unique_hh_ids:
                        hh_colors[hh_id] = generate_color_from_hh_id(hh_id)
                    
                    for _, r in cluster_households.iterrows():
                        if pd.notna(r[lat_col_name]) and pd.notna(r[lon_col_name]):
                            # Build tooltip with HH_ID, MEM_ID, coordinates, and address
                            hh_id_val = r.get('HH_ID', '')
                            mem_id_val = r.get('MEM_ID', '')
                            lat_val = r[lat_col_name] if lat_col_name else 'N/A'
                            lon_val = r[lon_col_name] if lon_col_name else 'N/A'
                            hh_address = r.get('hh_address', 'N/A') if 'hh_address' in r else 'N/A'
                            tooltip_text = f"HH: {hh_id_val}, MEM: {mem_id_val}\nLat: {lat_val}, Lon: {lon_val}\nAddress: {hh_address}"
                            
                            # Get color based on HH_ID
                            marker_color = hh_colors.get(hh_id_val, 'blue')
                            
                            folium.CircleMarker(
                                location=[r[lat_col_name], r[lon_col_name]],
                                popup=folium.Popup(f"HH: {r['household_number']}<br>HH_ID: {hh_id_val}<br>MEM_ID: {mem_id_val}<br>LATT,LONG: {lat_val}, {lon_val}<br>Address: {hh_address}", max_width=300),
                                tooltip=tooltip_text,
                                radius=5,
                                color=marker_color,
                                fill=True,
                                fillColor=marker_color,
                                fillOpacity=0.7,
                                weight=2
                            ).add_to(marker_cluster)
                    
                    # ── Add DBSCAN Cluster Center Circles ────────────────────────
                    # Show circles for each DBSCAN sub-cluster with radius = EPS km
                    if 'dbscan_cluster' in cluster_data.columns and show_circles:
                        _circle_data = cluster_households  # respects map_filter_mode
                        dbscan_clusters = _circle_data[_circle_data['dbscan_cluster'] >= 0]['dbscan_cluster'].unique()
                        
                        if len(dbscan_clusters) > 0:
                            colors_palette = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                           '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
                                           '#1f77b4', '#ff7f0e', '#2ca02c']
                            
                            for idx, dbscan_id in enumerate(sorted(dbscan_clusters)):
                                dbscan_points = _circle_data[
                                    (_circle_data['dbscan_cluster'] == dbscan_id) & 
                                    (pd.notna(_circle_data[lat_col_name])) & 
                                    (pd.notna(_circle_data[lon_col_name]))
                                ]
                                
                                if not dbscan_points.empty:
                                    # Convert to numeric before calculating mean
                                    lat_vals = pd.to_numeric(dbscan_points[lat_col_name], errors='coerce')
                                    lon_vals = pd.to_numeric(dbscan_points[lon_col_name], errors='coerce')
                                    center_lat = lat_vals.mean()
                                    center_lon = lon_vals.mean()
                                    
                                    try:
                                        eps_km = float(eps_val) if eps_val else 3.0
                                    except (ValueError, TypeError):
                                        eps_km = 3.0
                                    
                                    circle_color = colors_palette[idx % len(colors_palette)]
                                    
                                    folium.Circle(
                                        location=[center_lat, center_lon],
                                        radius=eps_km * 1000,
                                        popup=folium.Popup(
                                            f"<b>DBSCAN Cluster {int(dbscan_id)}</b><br>"
                                            f"Center: ({center_lat:.4f}, {center_lon:.4f})<br>"
                                            f"Radius: {eps_km} km<br>"
                                            f"Points in cluster: {len(dbscan_points)}",
                                            max_width=250
                                        ),
                                        tooltip=f"DBSCAN #{int(dbscan_id)} (EPS={eps_km}km)",
                                        color=circle_color,
                                        fill=True,
                                        fillColor=circle_color,
                                        fillOpacity=0.15,
                                        weight=2.5
                                    ).add_to(m)
                    
                    # ── Add GeoJSON Layer (filtered to selected CN only) ──────
                    if st.session_state.geojson_data:
                        try:
                            geojson_layer_data = st.session_state.geojson_data
                            features = geojson_layer_data.get('features', [])

                            _vdf = st.session_state.get('vilcode_csv_data')
                            if _vdf is not None and 'CN' in _vdf.columns and 'VILCODE11' in _vdf.columns:
                                # Get VILCODE11 codes belonging to the currently selected CN only
                                _cn_rows = _vdf[_vdf['CN'].astype(str).str.strip() == str(selected_cluster)]
                                _cn_vilcodes = set(
                                    _cn_rows['VILCODE11']
                                    .dropna()
                                    .astype(str)
                                    .str.strip()
                                    .str.split('.').str[0]
                                    .replace('', pd.NA)
                                    .dropna()
                                    .tolist()
                                )
                                if _cn_vilcodes:
                                    _vc_idx = st.session_state.get('geojson_vilcode_index', {})
                                    if _vc_idx:
                                        filtered_features = [f for vc in _cn_vilcodes for f in _vc_idx.get(vc, [])]
                                    else:
                                        filtered_features = [
                                            f for f in features
                                            if str(f.get('properties', {}).get('VILCODE11', '')).strip().split('.')[0]
                                            in _cn_vilcodes
                                        ]
                                    if filtered_features:
                                        geojson_layer_data = {'type': 'FeatureCollection', 'features': filtered_features}
                                        title_suffix = f" — CN {selected_cluster} ({len(filtered_features)} village{'s' if len(filtered_features) != 1 else ''})"
                                    else:
                                        geojson_layer_data = None
                                        st.warning(f"⚠️ No GeoJSON features found for CN {selected_cluster} (VILCODE11: {sorted(_cn_vilcodes)})")
                                        title_suffix = ""
                                else:
                                    # CN exists in CSV but all VILCODE11 values are blank
                                    geojson_layer_data = None
                                    title_suffix = ""
                            else:
                                # No CSV uploaded — show full GeoJSON without filtering
                                title_suffix = " (unfiltered)"

                            if geojson_layer_data:
                                folium.GeoJson(
                                    geojson_layer_data,
                                    name=f"📍 GeoJSON{title_suffix}",
                                    style_function=lambda x: {
                                        'color': '#1f77b4',
                                        'weight': 2,
                                        'opacity': 0.7,
                                        'fillColor': '#1f77b4',
                                        'fillOpacity': 0.2
                                    },
                                    popup=folium.GeoJsonPopup(fields=['VILCODE11'], labels=True)
                                ).add_to(m)

                            # ── Manual VILCODE11 overlay ────────────────────────────────
                            _manual_set = st.session_state.get('vilcode11_manual_set', set())
                            if _manual_set:
                                _vc_idx_m = st.session_state.get('geojson_vilcode_index', {})
                                if _vc_idx_m:
                                    _manual_features = [f for vc in _manual_set for f in _vc_idx_m.get(vc, [])]
                                else:
                                    _manual_features = [
                                        f for f in features
                                        if str(f.get('properties', {}).get('VILCODE11', '')).strip().split('.')[0]
                                        in _manual_set
                                    ]
                                if _manual_features:
                                    folium.GeoJson(
                                        {'type': 'FeatureCollection', 'features': _manual_features},
                                        name=f"🔢 Manual VILCODE11 ({len(_manual_features)} village{'s' if len(_manual_features) != 1 else ''})",
                                        style_function=lambda x: {
                                            'color': '#e85c2b',
                                            'weight': 2.5,
                                            'opacity': 0.9,
                                            'fillColor': '#e85c2b',
                                            'fillOpacity': 0.25
                                        },
                                        popup=folium.GeoJsonPopup(fields=['VILCODE11'], labels=True)
                                    ).add_to(m)
                                else:
                                    st.warning(f"⚠️ Manual codes not found in GeoJSON: {sorted(_manual_set)}")
                        except Exception as e:
                            st.warning(f"⚠️ Error adding GeoJSON layer: {e}")
                    
                    # Add layer control
                    Fullscreen(position='topleft', title='Full Screen', title_cancel='Exit Full Screen').add_to(m)
                    folium.LayerControl().add_to(m)
                    
                    # Display the map
                    folium_static(m, width=None, height=450)
                else:
                    st.info("No location data available for selected cluster.")
            else:
                st.info("No latitude / longitude columns detected.\n\nExpected columns containing **'lat'** and **'lon'** / **'lng'**.")
        
        render_map()

        st.html("<br>")
        
        # ── Add Other Cluster Plot Fragment ────────────────────────────────────
        @st.fragment
        def render_additional_cluster_plot():
            st.markdown("### 📍 Add Other Clusters")
            
            # Get all available clusters from the dataset
            all_clusters = sorted(df_processed['CN'].astype(str).unique())
            
            with st.form("additional_cluster_plot_form", border=True):
                col_form1, col_form2 = st.columns([3, 1])
                
                with col_form1:
                    selected_clusters_additional = st.multiselect(
                        "Select Clusters to Plot",
                        options=all_clusters,
                        default=[],
                        key="additional_clusters_select"
                    )
                
                with col_form2:
                    filter_mode_additional = st.radio(
                        "Status",
                        options=['Status 103 Only', 'Overall'],
                        index=0,
                        horizontal=True,
                        key='additional_cluster_filter_mode'
                    )
                
                submit_button = st.form_submit_button("✅ Plot Selected Clusters (all if none selected)")
            
            if submit_button:
                # If nothing selected, fall back to all clusters
                _plot_clusters = selected_clusters_additional if selected_clusters_additional else all_clusters
                if not selected_clusters_additional:
                    st.info(f"ℹ️ No clusters selected — plotting all {len(all_clusters)} clusters.")

                # Filter data for selected clusters
                selected_data = df_processed[df_processed['CN'].astype(str).isin(_plot_clusters)]
                
                # Apply status 103 filter if selected
                if filter_mode_additional == 'Status 103 Only':
                    selected_data = selected_data[selected_data['interview_status'] == 103]
                
                # Check if lat/lon columns exist
                if lat_col_name and lon_col_name:
                    map_data = selected_data[[lat_col_name, lon_col_name]].dropna().copy()
                    
                    # Convert to numeric (handles both string and numeric data)
                    map_data[lat_col_name] = pd.to_numeric(map_data[lat_col_name], errors='coerce')
                    map_data[lon_col_name] = pd.to_numeric(map_data[lon_col_name], errors='coerce')
                    # Drop rows that couldn't be converted
                    map_data = map_data.dropna()
                    
                    if not map_data.empty:
                        # Create base map centered on selected clusters
                        center = [map_data[lat_col_name].mean(), map_data[lon_col_name].mean()]
                        additional_map = folium.Map(
                            location=center,
                            zoom_start=7,
                            tiles=None,
                            control_scale=True
                        )
                        
                      
                        folium.TileLayer(
                            tiles='https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}',
                            attr='Google Satellite',
                            name='Google Satellite',
                            overlay=False,
                            control=True
                        ).add_to(additional_map)
                        
                        folium.TileLayer(
                            tiles='https://mt1.google.com/vt/lyrs=m&x={x}&y={y}&z={z}',
                            attr='Google Maps',
                            name='Google Maps',
                            overlay=False,
                            control=True,
                            show=True
                        ).add_to(additional_map)

                               # Add tile layers
                        folium.TileLayer(
                            tiles='cartodb positron',
                            attr='CartoDB Positron',
                            name='CartoDB Positron',
                            overlay=False,
                            control=True
                        ).add_to(additional_map)
                   
                        
                        # Add Geocoder
                        Geocoder(
                            collapsed=False,
                            position='topright',
                            add_marker=True,
                            zoom=13
                        ).add_to(additional_map)
                        
                        # Add Measure Control
                        MeasureControl(
                            position='topleft',
                            primary_length_unit='kilometers',
                            secondary_length_unit='meters',
                            primary_area_unit='sqkilometers',
                            secondary_area_unit='sqmeters'
                        ).add_to(additional_map)
                        
                        # Add marker cluster
                        marker_cluster_additional = MarkerCluster().add_to(additional_map)
                        
                        # Add markers for all selected data points
                        for _, r in selected_data.iterrows():
                            if pd.notna(r[lat_col_name]) and pd.notna(r[lon_col_name]):
                                try:
                                    lat_val = float(r[lat_col_name])
                                    lon_val = float(r[lon_col_name])
                                    hh_id_val = r.get('HH_ID', '')
                                    mem_id_val = r.get('MEM_ID', '')
                                    cluster_num = r.get('CN', '')
                                    
                                    folium.Marker(
                                        location=[lat_val, lon_val],
                                        popup=folium.Popup(
                                            f"Cluster: {cluster_num}<br>HH_ID: {hh_id_val}<br>MEM_ID: {mem_id_val}",
                                            max_width=250
                                        ),
                                        tooltip=f"CN: {cluster_num} | HH: {hh_id_val}"
                                    ).add_to(marker_cluster_additional)
                                except (ValueError, TypeError):
                                    # Skip rows with invalid lat/lon values
                                    pass
                        
                        # Add circles for each cluster if DBSCAN data is available
                        if 'dbscan_cluster' in selected_data.columns:
                            # Define color palette for clusters
                            cluster_colors = {
                                cluster: colors_palette[idx % len(colors_palette)]
                                for idx, cluster in enumerate(sorted(_plot_clusters))
                            }
                            colors_palette_additional = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
                                                         '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
                            cluster_colors = {
                                cluster: colors_palette_additional[idx % len(colors_palette_additional)]
                                for idx, cluster in enumerate(sorted(_plot_clusters))
                            }
                            
                            # For each selected cluster, draw circles around DBSCAN clusters
                            for cluster_num in sorted(_plot_clusters):
                                cluster_subset = selected_data[selected_data['CN'].astype(str) == cluster_num]
                                dbscan_clusters = cluster_subset[cluster_subset['dbscan_cluster'] >= 0]['dbscan_cluster'].unique()
                                
                                for dbscan_id in sorted(dbscan_clusters):
                                    dbscan_points = cluster_subset[
                                        (cluster_subset['dbscan_cluster'] == dbscan_id) &
                                        (pd.notna(cluster_subset[lat_col_name])) &
                                        (pd.notna(cluster_subset[lon_col_name]))
                                    ]
                                    
                                    if not dbscan_points.empty:
                                        # Convert to numeric before calculating mean
                                        lat_vals = pd.to_numeric(dbscan_points[lat_col_name], errors='coerce')
                                        lon_vals = pd.to_numeric(dbscan_points[lon_col_name], errors='coerce')
                                        center_lat = lat_vals.mean()
                                        center_lon = lon_vals.mean()
                                        
                                        try:
                                            eps_km = float(eps_val) if eps_val else 3.0
                                        except (ValueError, TypeError):
                                            eps_km = 3.0
                                        
                                        circle_color = cluster_colors.get(cluster_num, '#1f77b4')
                                        
                                        folium.Circle(
                                            location=[center_lat, center_lon],
                                            radius=eps_km * 1000,
                                            popup=folium.Popup(
                                                f"<b>Cluster {cluster_num} - DBSCAN {int(dbscan_id)}</b><br>"
                                                f"Center: ({center_lat:.4f}, {center_lon:.4f})<br>"
                                                f"Radius: {eps_km} km<br>"
                                                f"Points: {len(dbscan_points)}",
                                                max_width=250
                                            ),
                                            tooltip=f"CN:{cluster_num} DBS#{int(dbscan_id)}",
                                            color=circle_color,
                                            fill=True,
                                            fillColor=circle_color,
                                            fillOpacity=0.15,
                                            weight=2.5
                                        ).add_to(additional_map)
                        
                        # ── Add GeoJSON Layer (filtered to selected CNs only) ───────────
                        if st.session_state.geojson_data:
                            try:
                                geojson_layer_data = st.session_state.geojson_data
                                features = geojson_layer_data.get('features', [])

                                _vdf = st.session_state.get('vilcode_csv_data')
                                if _vdf is not None and 'CN' in _vdf.columns and 'VILCODE11' in _vdf.columns:
                                    # Collect VILCODE11 codes for ALL selected CNs
                                    _cn_rows = _vdf[_vdf['CN'].astype(str).str.strip().isin(
                                        [str(c) for c in _plot_clusters]
                                    )]
                                    _cn_vilcodes = set(
                                        _cn_rows['VILCODE11']
                                        .dropna()
                                        .astype(str)
                                        .str.strip()
                                        .str.split('.').str[0]
                                        .replace('', pd.NA)
                                        .dropna()
                                        .tolist()
                                    )
                                    if _cn_vilcodes:
                                        _vc_idx = st.session_state.get('geojson_vilcode_index', {})
                                        if _vc_idx:
                                            filtered_features = [f for vc in _cn_vilcodes for f in _vc_idx.get(vc, [])]
                                        else:
                                            filtered_features = [
                                                f for f in features
                                                if str(f.get('properties', {}).get('VILCODE11', '')).strip().split('.')[0]
                                                in _cn_vilcodes
                                            ]
                                        if filtered_features:
                                            geojson_layer_data = {'type': 'FeatureCollection', 'features': filtered_features}
                                            title_suffix = f" ({len(filtered_features)} village{'s' if len(filtered_features) != 1 else ''})"
                                        else:
                                            geojson_layer_data = None
                                            st.warning(f"⚠️ No GeoJSON features found for selected CNs.")
                                            title_suffix = ""
                                    else:
                                        geojson_layer_data = None
                                        title_suffix = ""
                                else:
                                    title_suffix = " (unfiltered)"

                                if geojson_layer_data:
                                    folium.GeoJson(
                                        geojson_layer_data,
                                        name=f"📍 GeoJSON{title_suffix}",
                                        style_function=lambda x: {
                                            'color': '#1f77b4',
                                            'weight': 2,
                                            'opacity': 0.7,
                                            'fillColor': '#1f77b4',
                                            'fillOpacity': 0.2
                                        },
                                        popup=folium.GeoJsonPopup(fields=['VILCODE11'], labels=True)
                                    ).add_to(additional_map)
                            except Exception as e:
                                st.warning(f"⚠️ Error adding GeoJSON layer: {e}")
                        
                        # Add layer control
                        Fullscreen(position='topleft', title='Full Screen', title_cancel='Exit Full Screen').add_to(additional_map)
                        folium.LayerControl().add_to(additional_map)
                        
                        # Display the map
                        st.markdown(f"**Displaying {len(_plot_clusters)} cluster(s) | Status: {filter_mode_additional}**")
                        folium_static(additional_map, width=None, height=500)
                    else:
                        st.warning("No data points found for the selected clusters with the chosen filter.")
                else:
                    st.error("Latitude/Longitude columns not found in data.")
        
        render_additional_cluster_plot()

        st.html("<br>")
        # ── Cluster Address Dataset ────────────────────────────────────────────
        st.subheader("Households Address")
        if 'hh_address' in df_processed.columns:
            # Determine columns to include
            cols_to_select = ['HH_ID', 'MEM_ID', 'hh_address']
            if 'dbscan_cluster' in df_processed.columns:
                cols_to_select.insert(2, 'dbscan_cluster')
            
            address_data = df_processed[df_processed['CN'].astype(str) == selected_cluster][cols_to_select].drop_duplicates(subset=['HH_ID'])
            if not address_data.empty:
                st.dataframe(address_data, use_container_width=True, height=420)
            else:
                st.info("No address data available for selected cluster.")
        else:
            st.info("No hh_address column found in dataset.")

        
        # ── Actions & Feedback ────────────────────────────────────────────────
        st.subheader("Actions & Feedback")
        
        # ── REMOVE CN ────────────────────────────────────────────────────────
        st.markdown("**🗑️ Remove CN**")
        st.caption("Remove selected clusters from dataset")
        with st.form("remove_cn_form"):
            all_cns = sorted(df_processed['CN'].astype(str).unique().tolist())
            cns_to_remove = st.multiselect(
                "Select CN to remove",
                options=all_cns,
                default=[],
                key="cns_to_remove_select"
            )
            remove_remarks = st.text_area(
                "Remarks (optional)",
                placeholder="Enter reason for removal...",
                height=80,
                key="remove_cn_remarks"
            )
            remove_cn_submit = st.form_submit_button("✅ Remove CN", use_container_width=True)
        
        if remove_cn_submit and cns_to_remove:
            st.session_state.removed_cns.extend(cns_to_remove)
            st.session_state.steps_tracker.append({
                'S.No': len(st.session_state.steps_tracker) + 1,
                'Step': 'Remove CN',
                'Details': f"Removed CN: {', '.join(cns_to_remove)}",
                'Remark': remove_remarks if remove_remarks.strip() else '-',
                '_op_type': 'remove_cn',
                '_op_data': {'cns': list(cns_to_remove)}
            })
            st.success(f"✅ Removed {len(cns_to_remove)} CN(s): {', '.join(cns_to_remove)}")
            st.rerun(scope="app")

        st.markdown("---")
        
        # ── RECONSIDER FLAG / MOVE (FRAGMENT) ────────────────────────────────
        reconsider_flag_move_fragment(selected_cluster, cn_filtered_df, df_processed, lat_col_name)
        
        st.markdown("---")




        # ── BREAK CN ─────────────────────────────────────────────────────────
        st.subheader("✂️ Break Cluster")
        st.caption("Merge selected DBSCAN sub-clusters into one new CN with unified geocodes & classification")

        cluster_data_for_break = df_processed[df_processed['CN'].astype(str) == selected_cluster].copy()
        if 'dbscan_cluster' in cluster_data_for_break.columns:
            available_dbscan_ids = sorted(cluster_data_for_break['dbscan_cluster'].dropna().unique().tolist())
            # Remove noise (-1) from choices — noise stays with original CN
            available_dbscan_ids_display = [int(x) for x in available_dbscan_ids if int(x) >= 0]

            if len(available_dbscan_ids_display) > 0:
                with st.form("break_cn_form"):
                    st.markdown(f"**Current CN: `{selected_cluster}`** — DBSCAN sub-clusters: {available_dbscan_ids_display}")
                    break_selected = st.multiselect(
                        "Select DBSCAN sub-clusters to merge into new CN",
                        options=available_dbscan_ids_display,
                        default=[],
                        key="break_dbscan_select"
                    )
                    st.markdown("**Optional overrides** (leave empty to keep existing values from selected obs)")
                    brk_c1, brk_c2 = st.columns(2)
                    with brk_c1:
                        break_bi1a = st.text_input("bi1a", value="", key="break_bi1a")
                        break_bi1c = st.text_input("bi1c", value="", key="break_bi1c")
                    with brk_c2:
                        break_bi1b = st.text_input("bi1b", value="", key="break_bi1b")
                        break_bi1c_1 = st.text_input("bi1c_1", value="", key="break_bi1c_1")
                    brk_g1, brk_g2 = st.columns(2)
                    with brk_g1:
                        break_lat = st.text_input("Custom hh_latitude (all obs)", value="", key="break_lat")
                    with brk_g2:
                        break_lon = st.text_input("Custom hh_longitude (all obs)", value="", key="break_lon")

                    break_remarks = st.text_area(
                        "Remarks (optional)",
                        placeholder="Enter reason for breaking cluster...",
                        height=80,
                        key="break_remarks"
                    )
                    break_submit = st.form_submit_button("✂️ Create New CN", use_container_width=True)

                if break_submit and break_selected:
                    # Generate ONE new CN for all selected DBSCAN clusters
                    existing_cns = sorted(df_processed['CN'].astype(str).unique().tolist())
                    max_cn_num = 0
                    for cn in existing_cns:
                        try:
                            num_part = int(cn[2:])  # skip first 2 chars (cluster_number prefix)
                            if num_part > max_cn_num:
                                max_cn_num = num_part
                        except (ValueError, IndexError):
                            pass
                    cn_prefix = str(selected_cluster)[:2]  # keep same 2-char prefix
                    new_cn = cn_prefix + f"{max_cn_num + 1:04d}"

                    # Store as single break operation with list of DBSCAN IDs
                    break_op_data = {
                        'source_cn': str(selected_cluster),
                        'dbscan_ids': [int(x) for x in break_selected],  # LIST of IDs → ONE CN
                        'new_cn': new_cn,
                        'bi1a': break_bi1a.strip() if break_bi1a.strip() else None,
                        'bi1b': break_bi1b.strip() if break_bi1b.strip() else None,
                        'bi1c': break_bi1c.strip() if break_bi1c.strip() else None,
                        'bi1c_1': break_bi1c_1.strip() if break_bi1c_1.strip() else None,
                        'lat': break_lat.strip() if break_lat.strip() else None,
                        'lon': break_lon.strip() if break_lon.strip() else None,
                    }
                    st.session_state.break_operations.append(break_op_data)

                    # Check if source CN will be empty after break (all DBSCAN clusters selected)
                    _break_remaining = cluster_data_for_break[
                        ~cluster_data_for_break['dbscan_cluster'].isin([int(x) for x in break_selected])
                    ]
                    _break_auto_removed_src = None
                    if len(_break_remaining) == 0:
                        _break_auto_removed_src = str(selected_cluster)

                    st.session_state.steps_tracker.append({
                        'S.No': len(st.session_state.steps_tracker) + 1,
                        'Step': 'Break CN',
                        'Details': f"CN {selected_cluster}: DBSCAN {', '.join(map(str, break_selected))} → New CN: {new_cn}",
                        'Remark': break_remarks if break_remarks.strip() else '-',
                        '_op_type': 'break',
                        '_op_data': {
                            'break_op': break_op_data,
                            'auto_removed_source_cn': _break_auto_removed_src
                        }
                    })
                    st.success(f"✅ Created new CN {new_cn} from DBSCAN {break_selected}")
                    st.rerun(scope="app")
            else:
                st.info("ℹ️ No DBSCAN sub-clusters detected.")
        else:
            st.info("ℹ️ Apply DBSCAN clustering first to enable Break Cluster.")

        st.markdown("---")
        
        # ── NEAREST NEIGHBOURS — MAP ─────────────────────────────────────────
        
        st.subheader("📍 Nearest Neighbours — Map")

        @st.fragment
        def render_nearest_neighbour_map():
            if lat_col_name and lon_col_name:
                # Reuse parameters set in the left panel (no duplicate controls)
                _nn_metric = st.session_state.get('nn_metric_left', 'Haversine')
                _nn_top_n  = st.session_state.get('nn_topn_left', 3)

                nn_viz = find_nearest_clusters(
                    df_processed, selected_cluster, lat_col_name, lon_col_name,
                    distance_metric=_nn_metric.lower(), top_n=_nn_top_n
                )

                if not nn_viz.empty:
                    # Distinct folium-compatible colors for up to 10 neighbour clusters
                    _palette = [
                        'blue', 'green', 'purple', 'orange', 'darkred',
                        'darkblue', 'darkgreen', 'cadetblue', 'darkpurple', 'pink'
                    ]

                    sel_lat, sel_lon = calculate_centroid(
                        df_processed[df_processed['CN'] == selected_cluster], lat_col_name, lon_col_name
                    )

                    nn_map = folium.Map(
                        location=[sel_lat, sel_lon],
                        zoom_start=9,
                        tiles=None,
                        control_scale=True
                    )

                    # Add explicit tile layers with proper names
                    folium.TileLayer(
                        tiles='cartodb positron',
                        attr='CartoDB Positron',
                        name='CartoDB Positron',
                        overlay=False,
                        control=True
                    ).add_to(nn_map)
                    
                    folium.TileLayer(
                        tiles='https://mt1.google.com/vt/lyrs=s&x={x}&y={y}&z={z}',
                        attr='Google Satellite',
                        name='Google Satellite',
                        overlay=False,
                        control=True
                    ).add_to(nn_map)

                    folium.TileLayer(
                        tiles='https://mt1.google.com/vt/lyrs=m&x={x}&y={y}&z={z}',
                        attr='Google Maps',
                        name='Google Maps',
                        overlay=False,
                        control=True,
                        show=True
                    ).add_to(nn_map)

                    # ── Selected cluster → RED ────────────────────────────────
                    sel_fg = folium.FeatureGroup(name=f"★ {selected_cluster} (Selected)").add_to(nn_map)
                    
                    # Filter selected cluster based on toggle
                    if st.session_state.map_filter_mode == 'Status 103 Only':
                        selected_cluster_data = df_processed[(df_processed['CN'].astype(str) == selected_cluster) & 
                                                             (df_processed['interview_status'] == 103)]
                    else:
                        selected_cluster_data = df_processed[df_processed['CN'].astype(str) == selected_cluster]
                    
                    for _, r in selected_cluster_data.iterrows():
                        if pd.notna(r[lat_col_name]) and pd.notna(r[lon_col_name]):
                            hh_id_val = r.get('HH_ID', '')
                            mem_id_val = r.get('MEM_ID', '')
                            folium.CircleMarker(
                                location=[r[lat_col_name], r[lon_col_name]],
                                radius=6, color='red', fill=True,
                                fill_color='red', fill_opacity=0.85,
                                popup=folium.Popup(
                                    f"<b>{selected_cluster}</b><br>HH_ID: {hh_id_val}<br>MEM_ID: {mem_id_val}", max_width=200
                                ),
                                tooltip=f"HH: {hh_id_val}, MEM: {mem_id_val}"
                            ).add_to(sel_fg)
                    # Centroid label for selected cluster
                    folium.Marker(
                        location=[sel_lat, sel_lon],
                        icon=folium.DivIcon(
                            html=f'<div style="font-size:11px;font-weight:bold;color:red;'
                                 f'background:white;padding:2px 5px;border:2px solid red;'
                                 f'border-radius:4px;white-space:nowrap;">★ {selected_cluster}</div>'
                        ),
                        tooltip=f"Centroid of {selected_cluster}"
                    ).add_to(sel_fg)

                    # ── Nearest clusters — each a different color ─────────────────
                    for i, (_, row) in enumerate(nn_viz.iterrows()):
                        cn     = row['CN']
                        color  = _palette[i % len(_palette)]
                        cn_fg  = folium.FeatureGroup(
                            name=f"{cn}  ({row['Distance_km']} km)"
                        ).add_to(nn_map)

                        # Filter neighbor cluster based on toggle
                        if st.session_state.map_filter_mode == 'Status 103 Only':
                            neighbor_data = df_processed[(df_processed['CN'] == cn) & 
                                                         (df_processed['interview_status'] == 103)]
                        else:
                            neighbor_data = df_processed[df_processed['CN'] == cn]
                        
                        for _, r in neighbor_data.iterrows():
                            if pd.notna(r[lat_col_name]) and pd.notna(r[lon_col_name]):
                                hh_id_val = r.get('HH_ID', '')
                                mem_id_val = r.get('MEM_ID', '')
                                folium.CircleMarker(
                                    location=[r[lat_col_name], r[lon_col_name]],
                                    radius=5, color=color, fill=True,
                                    fill_color=color, fill_opacity=0.75,
                                    popup=folium.Popup(
                                        f"<b>{cn}</b><br>HH_ID: {hh_id_val}<br>MEM_ID: {mem_id_val}"
                                        f"<br>Dist: {row['Distance_km']} km", max_width=200
                                    ),
                                    tooltip=f"HH: {hh_id_val}, MEM: {mem_id_val}"
                                ).add_to(cn_fg)

                        # Centroid label for each neighbour
                        folium.Marker(
                            location=[row['Centroid_Lat'], row['Centroid_Lon']],
                            icon=folium.DivIcon(
                                html=f'<div style="font-size:10px;font-weight:bold;color:{color};'
                                     f'background:white;padding:1px 4px;border:1px solid {color};'
                                     f'border-radius:3px;white-space:nowrap;">'
                                     f'{cn} · {row["Distance_km"]} km</div>'
                            ),
                            tooltip=f"Centroid: {cn} | {row['Distance_km']} km"
                        ).add_to(cn_fg)

                    # ── Add GeoJSON Layer (filtered to selected CN only) ───────────
                    if st.session_state.geojson_data:
                        try:
                            geojson_layer_data = st.session_state.geojson_data
                            features = geojson_layer_data.get('features', [])

                            _vdf = st.session_state.get('vilcode_csv_data')
                            if _vdf is not None and 'CN' in _vdf.columns and 'VILCODE11' in _vdf.columns:
                                _cn_rows = _vdf[_vdf['CN'].astype(str).str.strip() == str(selected_cluster)]
                                _cn_vilcodes = set(
                                    _cn_rows['VILCODE11']
                                    .dropna()
                                    .astype(str)
                                    .str.strip()
                                    .str.split('.').str[0]
                                    .replace('', pd.NA)
                                    .dropna()
                                    .tolist()
                                )
                                if _cn_vilcodes:
                                    _vc_idx = st.session_state.get('geojson_vilcode_index', {})
                                    if _vc_idx:
                                        filtered_features = [f for vc in _cn_vilcodes for f in _vc_idx.get(vc, [])]
                                    else:
                                        filtered_features = [
                                            f for f in features
                                            if str(f.get('properties', {}).get('VILCODE11', '')).strip().split('.')[0]
                                            in _cn_vilcodes
                                        ]
                                    if filtered_features:
                                        geojson_layer_data = {'type': 'FeatureCollection', 'features': filtered_features}
                                        title_suffix = f" — CN {selected_cluster} ({len(filtered_features)} village{'s' if len(filtered_features) != 1 else ''})"
                                    else:
                                        geojson_layer_data = None
                                        st.warning(f"⚠️ No GeoJSON features found for CN {selected_cluster}.")
                                        title_suffix = ""
                                else:
                                    geojson_layer_data = None
                                    title_suffix = ""
                            else:
                                title_suffix = " (unfiltered)"

                            if geojson_layer_data:
                                folium.GeoJson(
                                    geojson_layer_data,
                                    name=f"📍 GeoJSON{title_suffix}",
                                    style_function=lambda x: {
                                        'color': '#1f77b4',
                                        'weight': 2,
                                        'opacity': 0.7,
                                        'fillColor': '#1f77b4',
                                        'fillOpacity': 0.2
                                    },
                                    popup=folium.GeoJsonPopup(fields=['VILCODE11'], labels=True)
                                ).add_to(nn_map)
                        except Exception as e:
                            st.warning(f"⚠️ Error adding GeoJSON layer: {e}")
                    
                    Fullscreen(position='topleft', title='Full Screen', title_cancel='Exit Full Screen').add_to(nn_map)
                    folium.LayerControl(collapsed=False).add_to(nn_map)
                    folium_static(nn_map, width=None, height=480)

                    with st.expander("Distance Summary", expanded=False):
                        st.dataframe(
                            nn_viz[['CN', 'Distance_km', 'Count', 'Centroid_Lat', 'Centroid_Lon']],
                            use_container_width=True, hide_index=True
                        )
                else:
                    st.info("No nearby clusters found.")
            else:
                st.warning("Location columns not found.")
        
        render_nearest_neighbour_map()

    # ── CHECKSUMS ─────────────────────────────────────────────────────────────
    st.markdown("---")
    st.subheader("🔐 Checksums")
    
    # 1. Parent MD5 — from column in uploaded data (all rows share the same value)
    if 'Parent MD5' in df.columns:
        parent_md5_val = str(df['Parent MD5'].dropna().iloc[0]) if not df['Parent MD5'].dropna().empty else "(empty)"
    else:
        parent_md5_val = None
    
    # 2. Uploaded file MD5
    uploaded_md5_val = calculate_md5_checksum(uploaded_file)
    
    # 3. Output file MD5 (set when user clicks download button)
    output_md5_val = st.session_state.get('output_checksum', None)
    output_tmp_path = st.session_state.get('output_temp_path', None)
    
    chk_col1, chk_col2, chk_col3 = st.columns(3)
    
    with chk_col1:
        st.markdown("**① Parent MD5**")
        st.caption("Source reference — from `Parent MD5` column in uploaded data")
        if parent_md5_val is not None:
            st.code(parent_md5_val, language='text')
        else:
            st.warning("`Parent MD5` column not found in data")
    
    with chk_col2:
        st.markdown("**② Uploaded File MD5**")
        st.caption("Integrity check of the file uploaded to this session")
        if uploaded_md5_val:
            st.code(uploaded_md5_val, language='text')
        else:
            st.warning("Could not calculate checksum")
    
    with chk_col3:
        st.markdown("**③ Output File MD5**")
        st.caption("MD5 of the downloaded dataset; copy saved to `%TEMP%`")
        if output_md5_val:
            st.code(output_md5_val, language='text')
            st.caption(f"📁 Temp copy: `{output_tmp_path}`")
        else:
            st.info("Click **⬇ Download Complete Dataset** above to generate")

    # ── NAME CORRECTION FORM DOWNLOAD ──────────────────────────────────────
    st.markdown("---")
    st.subheader("📋 Download Name Correction Form")
    st.caption("Excel file with unique cluster identifiers for name corrections and validations")
    
    try:
        # Get DBSCAN parameters from session state
        min_neighbors_val = st.session_state.get('min_neighbors_input', st.session_state.get('min_neighbors_default_val', '10'))
        excel_bytes = generate_name_correction_form(df, df_processed, min_neighbors_val)
        if excel_bytes:
            st.download_button(
                "📥 Download Name Correction Form",
                excel_bytes,
                f"name_correction_form_{df['bi1a'].iloc[0] if len(df) > 0 else 'unknown'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    except Exception as e:
        st.error(f"Error generating Name Correction Form: {e}")

else:
    pass
    #st.info("👆 Upload a CSV or Excel file to get started.")